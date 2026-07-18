# -*- coding: utf-8 -*-
"""
Moralite — Rapor Motoru: son istihbarat/KAP sonucunu biçimli Excel'e döker.
Çıktılar `raporlar/` klasörüne yazılır; arayüzden indirilebilir.
"""
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

KLASOR = Path(__file__).resolve().parent.parent / "raporlar"
KLASOR.mkdir(exist_ok=True)

LACIVERT = "1F3864"
ACIK_MAVI = "D9E2F3"
GRI = "F2F2F2"


def _baslik_hucre(ws, satir, metin, genislik):
    ws.merge_cells(start_row=satir, start_column=1, end_row=satir, end_column=genislik)
    h = ws.cell(row=satir, column=1, value=metin)
    h.font = Font(bold=True, size=14, color="FFFFFF")
    h.fill = PatternFill("solid", fgColor=LACIVERT)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[satir].height = 26


def excel_uret(veri: dict) -> str:
    """veri: {soru, cevap, kaynaklar: [{no, baslik, url, ozet?}], tip, zaman}
    → üretilen dosya adı (raporlar/ altında)."""
    wb = Workbook()

    # ── Sayfa 1: Rapor ──
    ws = wb.active
    ws.title = "Rapor"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 100

    _baslik_hucre(ws, 1, "MORALİTE İSTİHBARAT RAPORU", 2)
    meta = [
        ("Tarih", veri.get("zaman", datetime.now().strftime("%d.%m.%Y %H:%M"))),
        ("Tür", "KAP Analizi" if veri.get("tip") == "kap" else "Web İstihbaratı"),
        ("Soru", veri.get("soru", "")),
        ("Kaynak sayısı", len(veri.get("kaynaklar") or [])),
    ]
    satir = 3
    for ad, deger in meta:
        a = ws.cell(row=satir, column=1, value=ad)
        a.font = Font(bold=True)
        a.fill = PatternFill("solid", fgColor=ACIK_MAVI)
        b = ws.cell(row=satir, column=2, value=str(deger))
        b.alignment = Alignment(wrap_text=True, vertical="top")
        satir += 1

    satir += 1
    baslik = ws.cell(row=satir, column=1, value="BULGULAR / DEĞERLENDİRME")
    baslik.font = Font(bold=True, size=11)
    satir += 1
    # Cevabı paragraflara böl; her paragraf ayrı satır (uzun metin okunur kalsın)
    paragraflar = [p.strip() for p in re.split(r"\n{1,}", veri.get("cevap", "")) if p.strip()]
    for p in paragraflar:
        ws.merge_cells(start_row=satir, start_column=1, end_row=satir, end_column=2)
        h = ws.cell(row=satir, column=1, value=p)
        h.alignment = Alignment(wrap_text=True, vertical="top")
        # kaba yükseklik tahmini: ~110 karakter/satır
        ws.row_dimensions[satir].height = max(16, 15 * (len(p) // 110 + 1))
        satir += 1

    satir += 1
    not_h = ws.cell(row=satir, column=1, value="Bu rapor yerel modelle üretilmiştir; yatırım tavsiyesi değildir.")
    not_h.font = Font(italic=True, size=9, color="808080")
    ws.merge_cells(start_row=satir, start_column=1, end_row=satir, end_column=2)

    # ── Sayfa 2: Kaynaklar ──
    wk = wb.create_sheet("Kaynaklar")
    sutunlar = [("No", 6), ("Başlık", 60), ("Özet", 60), ("Bağlantı", 60)]
    for i, (ad, gen) in enumerate(sutunlar, 1):
        h = wk.cell(row=1, column=i, value=ad)
        h.font = Font(bold=True, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor=LACIVERT)
        wk.column_dimensions[get_column_letter(i)].width = gen
    for r, k in enumerate(veri.get("kaynaklar") or [], 2):
        wk.cell(row=r, column=1, value=k.get("no"))
        wk.cell(row=r, column=2, value=k.get("baslik", "")).alignment = Alignment(wrap_text=True, vertical="top")
        wk.cell(row=r, column=3, value=(k.get("ozet") or "")[:500]).alignment = Alignment(wrap_text=True, vertical="top")
        h = wk.cell(row=r, column=4, value=k.get("url", ""))
        h.hyperlink = k.get("url", "")
        h.font = Font(color="0563C1", underline="single")
        if r % 2 == 0:
            for c in range(1, 5):
                wk.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GRI)

    ad = f"Moralite_Rapor_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    wb.save(KLASOR / ad)
    return ad
