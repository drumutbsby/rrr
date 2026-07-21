# -*- coding: utf-8 -*-
"""Model Kritiklik Degerlendirmesi (Questionnaire) - TAM Archer Field-Config."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="1F3864"; BLUE="2E5496"; GREY="F2F2F2"; GREEN="E2EFDA"
LB="DDEBF7"; ORANGE="F8CBAD"; YEL="FFE699"; GRN="C6E0B4"
thin=Side(style="thin",color="BFBFBF"); border=Border(left=thin,right=thin,top=thin,bottom=thin)

def hdr(cells,fill=NAVY):
    for c in cells:
        c.fill=PatternFill("solid",fgColor=fill); c.font=Font(bold=True,color="FFFFFF",size=9.5)
        c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True); c.border=border
def scell(c,size=9,wrap=True,top=True,fill=None,mono=False):
    c.font=Font(size=size,name="Consolas" if mono else "Calibri"); c.border=border
    c.alignment=Alignment(horizontal="left",vertical="top" if top else "center",wrap_text=wrap)
    if fill:c.fill=PatternFill("solid",fgColor=fill)
def widths(ws,ws_w):
    for i,w in enumerate(ws_w,1): ws.column_dimensions[get_column_letter(i)].width=w
def table(ws,start,headers,rows,ws_w,zebra=True,freeze=True,mono_cols=()):
    widths(ws,ws_w); r=start
    for j,h in enumerate(headers,1): ws.cell(row=r,column=j,value=h)
    hdr(ws[r]); ws.row_dimensions[r].height=28; r+=1
    for i,row in enumerate(rows):
        for j,val in enumerate(row,1):
            scell(ws.cell(row=r,column=j,value=val),fill=(GREY if zebra and i%2 else None),mono=((j-1) in mono_cols))
        r+=1
    if freeze: ws.freeze_panes=ws.cell(row=start+1,column=1)
    return r
def title_row(ws,text,span,r=1,size=11):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=span)
    c=ws.cell(row=r,column=1,value=text); c.font=Font(bold=True,size=size,color=NAVY)
    c.alignment=Alignment(wrap_text=True,vertical="center"); ws.row_dimensions[r].height=30

wb=openpyxl.Workbook()

# ===== 0. KUNYE =====
ws=wb.active; ws.title="0. Uygulama Künyesi"; ws.sheet_view.showGridLines=False
widths(ws,[34,92])
ws.merge_cells("A1:B1"); ws["A1"]="MODEL KRİTİKLİK DEĞERLENDİRMESİ – ARCHER FIELD-CONFIG"
ws["A1"].font=Font(size=18,bold=True,color=NAVY)
ws.merge_cells("A2:B2"); ws["A2"]="Questionnaire (anket) tam alan konfigürasyonu · Materyalite × Karmaşıklık → Tier · MRM Workspace · Taslak v0.1"
ws["A2"].font=Font(size=11,italic=True,color="595959")
kunye=[
 ("Uygulama tipi","QUESTIONNAIRE (Archer anket) – standart uygulamadan farklı; alanlar 'soru'dur, WEIGHTEDSCORE burada çalışır"),
 ("Uygulama adı","Model Kritiklik / Materyalite Değerlendirmesi"),
 ("Workspace","Model Riski Yönetimi (MRM)"),
 ("Anahtar alan","Değerlendirme Başlığı (Text). Tekil kimlik: Değerlendirme ID (Tracking ID)."),
 ("Ana ilişki","Değerlendirilen Model → Model Envanteri (skorlar/tier REF ile envantere geri yazılır)"),
 ("Skorlama modeli","İki boyut: Materyalite (1-5) × Karmaşıklık (1-5); her blok ağırlıkları %100'e normalize; İçsel Risk = Mat × Karm (1-25)"),
 ("Tier çıktısı","İçsel Risk eşiğinden Önerilen Tier; 'sert kurallar' (düzenleyici/sermaye kullanımı, otonomi) minimum tier'i yükseltir; 2. hat teyidi"),
 ("Questionnaire mekaniği","Her skorlu soru = Values List; SORU AĞIRLIĞI (weighting) + CEVAP başına SAYISAL DEĞER; WEIGHTEDSCORE = değer × ağırlık"),
 ("Geri yazım","Model Envanteri CALC-03/04/06 alanları bu anketten MOSTRECENTVALUE(REF(...)) ile skoru/tier'i çeker"),
 ("Ön koşul (mevcut)","Model Envanteri"),
 ("Kaynak çerçeve","PRA SS1/23 §1.3 (materyalite × karmaşıklık); OSFI E-23 (risk rating=zafiyet×materyalite, negligible)"),
]
r=4
for k,v in kunye:
    ws.cell(row=r,column=1,value=k).font=Font(bold=True,color=NAVY,size=10.5)
    c=ws.cell(row=r,column=2,value=v); c.alignment=Alignment(wrap_text=True,vertical="top"); c.font=Font(size=10.5)
    ws.row_dimensions[r].height=30 if len(v)<92 else 46; r+=1
ws.cell(row=r+1,column=1,value="SEKME REHBERİ").font=Font(bold=True,size=12,color=NAVY)
guide=[
 ("1. Field Config (Sorular)","Her soru/alan: tip, ağırlık, cevap ölçeği (numeric), zorunluluk, hesaplama/DDE, kaynak"),
 ("2. Cevap Ölçekleri","Standart 1-5 ölçekler ve modifiye ölçekleri, cevap başına sayısal değer/etiket"),
 ("3. Ağırlık & Skorlama Modeli","Blok ağırlıkları, skor formülleri, tier eşikleri, sert kurallar"),
 ("4. Hesaplanan Alanlar","WEIGHTEDSCORE / tier hesabı Archer formülleri"),
 ("5. Cross-Reference & Geri Yazım","Model Envanteri ile bağ ve skor/tier geri yazımı"),
 ("6. DDE","Koşullu sorular (AI/ML soruları, sert kural görünürlüğü)"),
 ("7. Advanced Workflow","Değerlendirme → 2. hat teyidi → yeniden değerlendirme"),
 ("8. Layout","Bölüm yerleşimi"),
 ("9. Erişim & FLS","Roller ve alan güvenliği"),
 ("10. Bildirimler","Periyodik yeniden değerlendirme"),
 ("11. Referanslar","Birincil kaynaklar"),
]
r+=2
for nm,d in guide:
    ws.cell(row=r,column=1,value=nm).font=Font(bold=True,color=BLUE,size=10)
    c=ws.cell(row=r,column=2,value=d); c.alignment=Alignment(wrap_text=True); c.font=Font(size=10); r+=1

# ===== 1. FIELD CONFIG (SORULAR) =====
ws=wb.create_sheet("1. Field Config")
title_row(ws,"KRİTİKLİK DEĞERLENDİRMESİ – SORU/ALAN KONFİGÜRASYONU  ·  Ağırlık = soru weighting  ·  Cevap ölçeği sayısal değer taşır (SELECTEDVALUENUMBER/WEIGHTEDSCORE)",10)
FH=["#","Blok / Bölüm","Soru / Alan","Soru/Alan Tipi (Archer)","Ağırlık","Cevap Ölçeği","Zor.","Hesaplama / DDE","Kaynak"]
F=[
 # KIMLIK / BAGLAM
 ["1","Bağlam","Değerlendirme ID","Tracking ID (System)","—","—","E","KRT-##### auto","Tasarım"],
 ["2","Bağlam","Değerlendirme Başlığı","Text (tek satır) · KEY","—","—","E","Ör. '[Model] – Kritiklik 2026'","Tasarım"],
 ["3","Bağlam","Değerlendirilen Model","Cross-Reference","—","→ Model Envanteri","E","XREF-01; geri yazım kaynağı","SS 1.3"],
 ["4","Bağlam","Model Tipi (getirilen)","Values List / Read-only","—","← Envanter REF","H","CALC-01; AI sorularını sürer (DDE-01)","SS 1.3c"],
 ["5","Bağlam","Düzenleyici / Sermaye Kullanımı","Values List (Evet/Hayır)","—","Evet/Hayır","E","SERT KURAL girdisi (DDE-05)","OSFI (negligible istisnası)"],
 ["6","Bağlam","Değerlendiren","Users/Groups List","—","—","E","1. hat/model sahibi","SS 2.4c"],
 ["7","Bağlam","Değerlendirme Tarihi","Date","—","—","E","","Tasarım"],
 # MATERYALITE
 ["8","Materyalite (Σ=%100)","M1 – Risk tutarı / portföy büyüklüğü / bilanço etkisi","Values List (tekli)","0,30","Materyalite 1-5","E","WEIGHTEDSCORE(M1)","SS 1.3b (nicel)"],
 ["9","Materyalite","M2 – Etkilenen müşteri / işlem sayısı","Values List (tekli)","0,15","Materyalite 1-5","E","WEIGHTEDSCORE(M2)","SS 1.3b (nicel)"],
 ["10","Materyalite","M3 – Kararlara etki (fiyatlama/karşılık/sermaye/onay)","Values List (tekli)","0,25","Materyalite 1-5","E","WEIGHTEDSCORE(M3)","SS 1.3b (nitel)"],
 ["11","Materyalite","M4 – Ödeme gücü / yasal sermaye / finansal raporlama etkisi","Values List (tekli)","0,20","Materyalite 1-5","E","WEIGHTEDSCORE(M4)","SS 1.3b (nitel)"],
 ["12","Materyalite","M5 – İtibar / müşteri / davranışsal etki","Values List (tekli)","0,10","Materyalite 1-5","E","WEIGHTEDSCORE(M5)","SS 1.3b (nitel)"],
 # KARMASIKLIK
 ["13","Karmaşıklık (Σ=%100)","K1 – Girdi veri niteliği/kalitesi (alternatif/yapılandırılmamış?)","Values List (tekli)","0,18","Karmaşıklık 1-5","E","WEIGHTEDSCORE(K1)","SS 1.3c; AI faktörü"],
 ["14","Karmaşıklık","K2 – Metodoloji / varsayım karmaşıklığı","Values List (tekli)","0,20","Karmaşıklık 1-5","E","WEIGHTEDSCORE(K2)","SS 1.3c"],
 ["15","Karmaşıklık","K3 – Uygulama bütünlüğü / entegrasyon karmaşıklığı","Values List (tekli)","0,12","Karmaşıklık 1-5","E","WEIGHTEDSCORE(K3)","SS 1.3c"],
 ["16","Karmaşıklık","K4 – Kullanım sıklığı ve yaygınlığı","Values List (tekli)","0,10","Karmaşıklık 1-5","E","WEIGHTEDSCORE(K4)","SS 1.3c"],
 ["17","Karmaşıklık","K5 – Açıklanabilirlik / şeffaflık / yorumlanabilirlik (opaklık)","Values List (tekli)","0,15","Karmaşıklık 1-5","K","WEIGHTEDSCORE(K5); AI'da zorunlu (DDE-01)","SS 1.3c(ii); EU AI Act"],
 ["18","Karmaşıklık","K6 – Tasarımcı / veri yanlılığı potansiyeli","Values List (tekli)","0,10","Karmaşıklık 1-5","K","WEIGHTEDSCORE(K6); AI'da zorunlu","SS 1.3c(ii); EU AI Act"],
 ["19","Karmaşıklık","K7 – Otonomi / dinamik yeniden kalibrasyon","Values List (tekli)","0,15","Karmaşıklık 1-5","H","WEIGHTEDSCORE(K7)","SS (dinamik model)"],
 # MODIFIYE / SERT KURAL
 ["20","Modifiye","B1 – Downstream (beslenen) model/karar bağımlılığı sayısı","Values List (tekli)","modifiye","Modifiye 0-3","H","Tier eskalasyon (CALC-06)","SS 1.2b (ara bağımlılık)"],
 ["21","Modifiye","O1 – İnsan gözetimi olmadan otomatik karara girme derecesi","Values List (tekli)","modifiye","Modifiye 0-3","K","SERT KURAL; AI'da zorunlu (DDE-01)","EU AI Act (Art.14)"],
 # SONUC / SKOR
 ["22","Sonuç (Skor)","Materyalite Skoru","Numeric (Calculated)","—","1-5","H","CALC-02 = Σ WEIGHTEDSCORE(M1..M5)","SS 1.3b"],
 ["23","Sonuç (Skor)","Karmaşıklık Skoru","Numeric (Calculated)","—","1-5","H","CALC-03 = Σ WEIGHTEDSCORE(K1..K7)","SS 1.3c"],
 ["24","Sonuç (Skor)","İçsel (Inherent) Risk Skoru","Numeric (Calculated)","—","1-25","H","CALC-04 = Mat × Karm","OSFI (zafiyet×materyalite)"],
 ["25","Sonuç (Skor)","Önerilen Tier","Text (Calculated)","—","Tier 1/2/3/İhmal","H","CALC-05 (eşik + sert kurallar)","SS 1.3; OSFI"],
 ["26","Sonuç (Skor)","Model Risk Derecesi (öneri)","Values List (Calculated/derived)","—","Yüksek/Orta/Düşük","H","İçsel risk bandından türetilir","OSFI (rating)"],
 # TEYIT / TAKIP
 ["27","Teyit ve Takip","Nihai Tier (2. hat teyidi)","Values List (tekli)","—","Tier 1/2/3/İhmal","E","2. hat yazar; override edebilir","SS 1.3e (bağımsız teyit)"],
 ["28","Teyit ve Takip","Tier Override Gerekçesi","Text (çok satır)","—","—","K","Önerilen≠Nihai ise zorunlu (DDE-06)","SS 1.3e"],
 ["29","Teyit ve Takip","2. Hat Teyit Durumu","Values List (tekli)","—","Bekliyor/Teyit/İtiraz","E","AWF","SS 1.3d/e"],
 ["30","Teyit ve Takip","Teyit Tarihi","Date","—","—","K","","SS"],
 ["31","Teyit ve Takip","Sonraki Yeniden Değerlendirme Tarihi","Date (Calculated)","—","—","H","CALC-07 (nihai tier→periyot)","SS 1.3d (periyodik)"],
 ["32","Teyit ve Takip","İlgili Bulgular","Cross-Reference","—","→ Bulgular","H","Tier/veri sorunları için","SS"],
 # SISTEM
 ["33","Sistem (System)","Kayıt İzinleri","Record Permissions","—","—","—","1. hat R/U; 2. hat teyit alanları U","SS bağımsızlık"],
 ["34","Sistem (System)","Oluşturulma / Son Güncelleme","First Published / Last Updated","—","—","—","Sistem","Platform"],
 ["35","Sistem (System)","Değişiklik Geçmişi","History Log","—","—","—","Denetlenebilirlik","OSFI (controls)"],
]
endr=table(ws,2,FH,F,[4,20,44,26,9,18,5,32,20])
for row in ws.iter_rows(min_row=3,max_row=endr-1,min_col=7,max_col=7):
    for c in row:
        c.fill=PatternFill("solid",fgColor={"E":ORANGE,"K":YEL,"H":LB}.get(c.value,"FFFFFF"))
        c.alignment=Alignment(horizontal="center",vertical="top")
for row in ws.iter_rows(min_row=3,max_row=endr-1,min_col=4,max_col=4):
    for c in row:
        if c.value and "Calculated" in str(c.value): c.fill=PatternFill("solid",fgColor=GREEN)
        elif c.value and ("Cross-Reference" in str(c.value) or "Read-only" in str(c.value)): c.fill=PatternFill("solid",fgColor=LB)

# ===== 2. CEVAP OLCEKLERI =====
ws=wb.create_sheet("2. Cevap Ölçekleri")
title_row(ws,"CEVAP ÖLÇEKLERİ (VALUES LIST + NUMERIC MAPPING)  ·  Her skorlu sorunun cevabı bir sayısal değer taşır",4)
SC=[
 ["Materyalite 1-5","1 – Çok Düşük","1","İhmal edilebilir tutar/etki"],
 ["Materyalite 1-5","2 – Düşük","2",""],
 ["Materyalite 1-5","3 – Orta","3",""],
 ["Materyalite 1-5","4 – Yüksek","4",""],
 ["Materyalite 1-5","5 – Çok Yüksek","5","Büyük tutar/kritik karar/sermaye/finansal raporlama"],
 ["Karmaşıklık 1-5","1 – Çok Basit","1","Şeffaf, standart, düşük veri karmaşıklığı"],
 ["Karmaşıklık 1-5","2 – Basit","2",""],
 ["Karmaşıklık 1-5","3 – Orta","3",""],
 ["Karmaşıklık 1-5","4 – Karmaşık","4",""],
 ["Karmaşıklık 1-5","5 – Çok Karmaşık","5","Opak/AI, alternatif veri, yüksek otonomi"],
 ["Modifiye 0-3","0 – Yok","0","Etki yok"],
 ["Modifiye 0-3","1 – Sınırlı","1",""],
 ["Modifiye 0-3","2 – Belirgin","2",""],
 ["Modifiye 0-3","3 – Yüksek","3","Çok sayıda downstream / tam otonom karar"],
 ["Evet/Hayır","Evet","1","Sert kural tetikler"],
 ["Evet/Hayır","Hayır","0",""],
 ["Tier","Tier 1","3","En yüksek yoğunluk"],
 ["Tier","Tier 2","2",""],
 ["Tier","Tier 3","1",""],
 ["Tier","İhmal Edilebilir","0","Hafif yol (düzenleyici kullanımda YASAK)"],
 ["2. Hat Teyit Durumu","Bekliyor · Teyit Edildi · İtiraz/Yeniden Değerlendir","—","AWF"],
]
r=table(ws,2,["Ölçek Adı","Cevap / Değer","Sayısal Değer","Anlam / Anchor"],SC,[22,54,14,42])
note=ws.cell(row=r+1,column=1,value="Not: M/K sorularının her biri kendi bağlamına göre 1-5 ölçeğinde 'anchor' cümlelerle detaylandırılmalıdır (ör. M1 için 1='<X mn TL', 5='>Y mld TL'). Eşikler banka risk iştahınca kalibre edilir.")
note.font=Font(italic=True,size=9,color="595959"); ws.merge_cells(start_row=r+1,start_column=1,end_row=r+1,end_column=4); ws.row_dimensions[r+1].height=30

# ===== 3. AGIRLIK & SKORLAMA MODELI =====
ws=wb.create_sheet("3. Ağırlık & Skorlama")
title_row(ws,"AĞIRLIK & SKORLAMA MODELİ  ·  Materyalite × Karmaşıklık iki boyutu; her blok Σ ağırlık = %100",5)
# Agirlik tablosu
title_row(ws,"1) SORU AĞIRLIKLARI (question weighting)",5,r=3,size=10)
W=[
 ["Materyalite","M1 Risk tutarı/portföy/bilanço","0,30",""],
 ["Materyalite","M2 Müşteri/işlem sayısı","0,15",""],
 ["Materyalite","M3 Kararlara etki","0,25",""],
 ["Materyalite","M4 Ödeme gücü/sermaye/finansal rap.","0,20",""],
 ["Materyalite","M5 İtibar/davranışsal","0,10",""],
 ["Materyalite","TOPLAM","1,00","→ Materyalite Skoru (1-5)"],
 ["Karmaşıklık","K1 Girdi veri niteliği","0,18",""],
 ["Karmaşıklık","K2 Metodoloji/varsayım","0,20",""],
 ["Karmaşıklık","K3 Uygulama/entegrasyon","0,12",""],
 ["Karmaşıklık","K4 Kullanım sıklığı","0,10",""],
 ["Karmaşıklık","K5 Açıklanabilirlik (AI)","0,15",""],
 ["Karmaşıklık","K6 Yanlılık (AI)","0,10",""],
 ["Karmaşıklık","K7 Otonomi/dinamik","0,15",""],
 ["Karmaşıklık","TOPLAM","1,00","→ Karmaşıklık Skoru (1-5)"],
]
rr=table(ws,4,["Blok","Soru","Ağırlık","Not"],W,[16,40,10,30],freeze=False)
for row in ws.iter_rows(min_row=5,max_row=rr-1,min_col=2,max_col=3):
    for c in row:
        if c.value=="TOPLAM" or c.value=="1,00": c.font=Font(bold=True,size=9)
# Tier esik tablosu
r2=rr+2
title_row(ws,"2) İÇSEL RİSK → TIER EŞİKLERİ (parametrik – banka politikası)",5,r=r2,size=10)
TT=[
 ["İçsel Risk ≥ 15","Tier 1","Tam / derinlemesine validasyon","Yıllık"],
 ["8 ≤ İçsel Risk < 15","Tier 2","Standart validasyon","2 Yıllık"],
 ["3 ≤ İçsel Risk < 8","Tier 3","Hafif / hedefli gözden geçirme","3 Yıllık"],
 ["İçsel Risk < 3","İhmal Edilebilir","Kayıt + asgari kontrol","Olay-tetikli"],
]
r3=table(ws,r2+1,["İçsel Risk (Mat×Karm, 1-25)","Önerilen Tier","Validasyon Kapsamı","Örnek Revalidasyon Frekansı"],TT,[26,18,34,26],freeze=False)
# Sert kurallar
r4=r3+1
title_row(ws,"3) SERT KURALLAR (tier eskalasyonu – skorun üzerine)",5,r=r4,size=10)
HR=[
 ["Düzenleyici/Sermaye kullanımı = Evet (IRB, İSEDES, TFRS9, piyasa/AMA)","Minimum Tier 2; 'İhmal Edilebilir' SEÇİLEMEZ","OSFI negligible istisnası; BDDK"],
 ["O1 Otonomi = 3 (tam otonom) VE müşteri etkisi yüksek","Minimum Tier 2","EU AI Act (Art.14)"],
 ["B1 Downstream bağımlılık = 3 (çok sayıda beslenen model)","Önerilen tier bir kademe yükselir","SS 1.2b"],
 ["GenAI/Temel Model = Evet","Minimum Tier 2 (ek AI/ML değerlendirmesi zorunlu)","NIST AI 600-1; WF GenAI"],
]
table(ws,r4+1,["Koşul","Etki","Kaynak"],HR,[46,34,26],freeze=False)

# ===== 4. HESAPLANAN ALANLAR =====
ws=wb.create_sheet("4. Hesaplanan Alanlar")
title_row(ws,"HESAPLANAN ALANLAR  ·  WEIGHTEDSCORE YALNIZ questionnaire'de çalışır (Values List sorusu + ağırlık + cevap sayısal değeri)",4)
C=[
 ["CALC-01","Model Tipi (getirilen)","Text",'MOSTRECENTVALUE(REF([Değerlendirilen Model],[Model Tipi]))',"AI sorularının koşullu görünümü için modelden tip çeker (DDE-01)"],
 ["CALC-02","Materyalite Skoru","Numeric",'SUM(WEIGHTEDSCORE([M1]),WEIGHTEDSCORE([M2]),WEIGHTEDSCORE([M3]),WEIGHTEDSCORE([M4]),WEIGHTEDSCORE([M5]))',"Ağırlıklar Σ=1 olduğundan sonuç 1-5 ağırlıklı ortalama"],
 ["CALC-03","Karmaşıklık Skoru","Numeric",'SUM(WEIGHTEDSCORE([K1]),WEIGHTEDSCORE([K2]),WEIGHTEDSCORE([K3]),WEIGHTEDSCORE([K4]),WEIGHTEDSCORE([K5]),WEIGHTEDSCORE([K6]),WEIGHTEDSCORE([K7]))',"1-5 ağırlıklı ortalama (AI faktörleri dahil)"],
 ["CALC-04","İçsel Risk Skoru","Numeric",'ROUND([Materyalite Skoru]*[Karmaşıklık Skoru],2)',"1-25 aralığı (zafiyet=karmaşıklık × materyalite)"],
 ["CALC-05","Önerilen Tier","Text",
  'IF(OR(CONTAINS(EXACT,VALUEOF([Düzenleyici/Sermaye Kullanımı],"Evet")),SELECTEDVALUENUMBER([O1])>=3,CONTAINS(EXACT,VALUEOF([GenAI Bayrağı],"Evet"))),MAX_TIER2, TIER_BY_SCORE)  //sözde-kod',
  "SERT KURAL: düzenleyici/otonom/GenAI ise min Tier 2; aksi halde İçsel Risk eşiğinden (≥15→T1, ≥8→T2, ≥3→T3, <3→İhmal). B1=3 ise bir kademe yükselt. Gerçek formül IF zinciriyle kurulur."],
 ["CALC-06","Model Risk Derecesi","Values List/derived",'IF([İçsel Risk Skoru]>=15,"Yüksek",IF([İçsel Risk Skoru]>=8,"Orta","Düşük"))',"İçsel risk bandından derece"],
 ["CALC-07","Sonraki Yeniden Değerlendirme Tarihi","Date",'IF(ISEMPTY([Teyit Tarihi]),"",DATEADD([Teyit Tarihi],"m",TIER_PERIYOT_AY))',"Nihai tier'e göre periyot (T1=12, T2=24, T3=36 ay); jeton teyidi"],
]
r=table(ws,2,["ID","Alan","Dönüş Tipi","Formül (Archer)","Açıklama"],C,[9,26,12,84,40],mono_cols=(3,))
for row in ws.iter_rows(min_row=3,max_row=r-1,min_col=4,max_col=4):
    for c in row: c.font=Font(name="Consolas",size=8.5)

# ===== 5. CROSS-REFERENCE & GERI YAZIM =====
ws=wb.create_sheet("5. XRef & Geri Yazım")
title_row(ws,"CROSS-REFERENCE & SKOR/TIER GERİ YAZIMI  ·  Anket hesaplar, Model Envanteri REF ile çeker",4)
X=[
 ["XREF-01","Değerlendirilen Model","Cross-Reference (external)","Model Envanteri","'Kritiklik Değerlendirmesi' (Envanter XREF-05) · Yeni"],
 ["XREF-02","İlgili Bulgular","Cross-Reference (external)","Bulgular","'Kaynak Değerlendirme' · Mevcut"],
]
r=table(ws,2,["ID","Kaynak Alan (Anket)","Tip","Hedef Uygulama","Ters (Related Records)/Yeni-Mevcut"],X,[10,26,26,26,44])
title_row(ws,"GERİ YAZIM: Model Envanteri hangi alanı bu anketten çeker",5,r=r+1,size=10)
GB=[
 ["Model Envanteri > Materyalite Skoru (CALC-03)","← MOSTRECENTVALUE(REF([Kritiklik Değerlendirmesi],[Materyalite Skoru]))"],
 ["Model Envanteri > Karmaşıklık Skoru (CALC-04)","← MOSTRECENTVALUE(REF([Kritiklik Değerlendirmesi],[Karmaşıklık Skoru]))"],
 ["Model Envanteri > İçsel Risk Skoru (CALC-05)","← Materyalite × Karmaşıklık (envanterde de hesaplanabilir)"],
 ["Model Envanteri > Tier (CALC-06)","← MOSTRECENTVALUE(REF([Kritiklik Değerlendirmesi],[Nihai Tier]))"],
]
table(ws,r+2,["Model Envanteri alanı","Geri yazım kaynağı (REF)"],GB,[44,72],freeze=False,mono_cols=(1,))

# ===== 6. DDE =====
ws=wb.create_sheet("6. DDE")
title_row(ws,"DATA-DRIVEN EVENTS (DDE)  ·  Koşullu sorular ve sert kural görünürlüğü",5)
D=[
 ["DDE-01",'[Model Tipi (getirilen)] IN {AI/ML, GenAI/Temel Model}',"ACL","K5 (açıklanabilirlik), K6 (yanlılık), O1 (otonomi) sorularını zorunlu yap; AI açıklama metnini göster","AI karmaşıklık faktörleri (SS 1.3c-ii)"],
 ["DDE-02",'[Model Tipi (getirilen)] = "Deterministik/Kural"',"ACL","Karmaşıklık sorularını sadeleştir; opsiyonel yap","Kapsam-dışı deterministik hafifletme"],
 ["DDE-03",'[Önerilen Tier] ≠ [Nihai Tier]',"ACL","[Tier Override Gerekçesi] zorunlu yap","2. hat override şeffaflığı (SS 1.3e)"],
 ["DDE-04",'[Düzenleyici/Sermaye Kullanımı] = "Evet"',"ACL","'İhmal Edilebilir' seçeneğini Nihai Tier listesinden FİLTRELE (FVL); uyarı göster","Sert kural: negligible yasak"],
 ["DDE-05",'[2. Hat Teyit Durumu] = "İtiraz/Yeniden Değerlendir"',"ACL","Yeniden değerlendirme uyarısı; skorları düzenlenebilir yap","Teyit döngüsü"],
]
table(ws,2,["ID","Kural (Koşul)","Aksiyon","Aksiyon Detayı","Amaç"],D,[9,40,10,48,32])

# ===== 7. ADVANCED WORKFLOW =====
ws=wb.create_sheet("7. Advanced Workflow")
title_row(ws,"ADVANCED WORKFLOW (AWF)  ·  Değerlendirme → 2. hat teyidi → tamamlanma",5)
A=[
 ["N0","Start","Oluşturma","(otomatik)","N1"],
 ["N1","User Action","Değerlendirme (1. hat)","Teyide Gönder","N2 (skorlar hesaplanır)"],
 ["N2","User Action","2. Hat Teyidi","Teyit Et · İtiraz/Yeniden Değerlendir","N3 / N1"],
 ["N3","Update Content","Nihai Tier Damgala","(otomatik)","Model Envanteri Tier/skor güncellenir (REF/geri yazım); N4"],
 ["N4","Stop","Tamamlandı","—","(son; sonraki yeniden değerlendirme planlanır)"],
]
table(ws,2,["Node #","Node Tipi","Node Adı","Geçiş(ler)/Buton","Sonraki Node/Aksiyon"],A,[8,20,26,36,44])

# ===== 8. LAYOUT =====
ws=wb.create_sheet("8. Layout")
title_row(ws,"LAYOUT  ·  Bölüm yerleşimi",3)
L=[
 ["Bağlam","Model ve değerlendirme bağlamı","1-7"],
 ["Materyalite","Materyalite soruları (Σ=%100)","8-12"],
 ["Karmaşıklık","Karmaşıklık soruları (Σ=%100, AI dahil)","13-19"],
 ["Modifiye","Bağımlılık ve otonomi","20-21"],
 ["Sonuç (Skor)","Hesaplanan skorlar ve önerilen tier","22-26"],
 ["Teyit ve Takip","2. hat teyidi, override, yeniden değerlendirme","27-32"],
 ["Sistem","Kayıt yönetimi","33-35"],
]
table(ws,2,["Bölüm (Section)","Açıklama","Sorular/Alanlar (#)"],L,[22,42,24])

# ===== 9. ERISIM & FLS =====
ws=wb.create_sheet("9. Erişim & FLS")
title_row(ws,"ERİŞİM & ALAN-SEVİYESİ GÜVENLİK  ·  R=Read, U=Update  ·  Skorlama 1. hat, teyit 2. hat",5)
E=[
 ["Bağlam / Materyalite / Karmaşıklık / Modifiye (sorular)","R/U","R","R","R"],
 ["Sonuç Skorları (hesaplanan)","R","R","R","R"],
 ["Nihai Tier / Override Gerekçesi / Teyit","R","R/U","R","R"],
 ["Sistem / History Log","R","R","R","R (denetim)"],
]
r=table(ws,2,["Bölüm / Alan grubu","1.Hat Sahip/Değerlendiren","2.Hat Bağımsız Validasyon","2.Hat MRM Programı","3.Hat İç Denetim"],E,[42,22,24,18,18])
for row in ws.iter_rows(min_row=3,max_row=r-1,min_col=2,max_col=5):
    for c in row:
        v=str(c.value); c.fill=PatternFill("solid",fgColor=(GRN if "U" in v else LB))
        c.alignment=Alignment(horizontal="center",vertical="center")
note=ws.cell(row=r+1,column=1,value="Not: 1. hat skorlama sorularını doldurur; Nihai Tier ve teyit alanları yalnız 2. hat bağımsız validasyona açıktır (SS1/23 1.3e – bireysel tier atamalarının bağımsız yeniden değerlendirilmesi).")
note.font=Font(italic=True,size=9,color="595959"); ws.merge_cells(start_row=r+1,start_column=1,end_row=r+1,end_column=5); ws.row_dimensions[r+1].height=30

# ===== 10. BILDIRIMLER =====
ws=wb.create_sheet("10. Bildirimler")
title_row(ws,"BİLDİRİMLER (NOTIFICATIONS)",4)
N=[
 ["Değerlendirme teyide gönderildi","2. Hat Validasyon","Anında","Teyit talebi"],
 ["Nihai Tier damgalandı","Model Sahibi, MRM Programı","Anında","Envanter güncellendi"],
 ["Sonraki Yeniden Değerlendirme yaklaşıyor (30 gün)","Model Sahibi","Günlük değerlendirme","Periyodik tiering (SS 1.3d)"],
 ["Model materyal değişti (Envanter'den tetik)","Model Sahibi, MRM","Anında","Tier yeniden değerlendirme"],
 ["Sert kural ihlali (düzenleyici model 'İhmal' seçildi)","Model Sahibi, MRM","Anında","Uyumsuz tier engeli"],
]
table(ws,2,["Tetikleyici","Alıcı","Zamanlama","Amaç"],N,[42,26,22,32])

# ===== 11. REFERANSLAR =====
ws=wb.create_sheet("11. Referanslar")
title_row(ws,"BİRİNCİL KAYNAKLAR (Kritiklik/Tiering)",3)
RR=[
 ["PRA/BoE","SS1/23 §1.3 – model tiering (materyalite × karmaşıklık)","https://www.bankofengland.co.uk/-/media/boe/files/prudential-regulation/supervisory-statement/2023/ss123.pdf"],
 ["OSFI","E-23 (2027) – model risk rating (zafiyet×materyalite), negligible","https://www.osfi-bsif.gc.ca/en/guidance/guidance-library/guideline-e-23-model-risk-management-2027"],
 ["US Fed","SR 11-7 – risk-tabanlı yoğunluk","https://www.federalreserve.gov/supervisionreg/srletters/sr1107.htm"],
 ["EU","AI Act Art.14 (insan gözetimi), veri/şeffaflık","https://artificialintelligenceact.eu/annex/3/"],
 ["Archer Help","WEIGHTEDSCORE (questionnaire)","https://help.archerirm.cloud/610-en/content/platform/calcfunctions/system_weightedscore.htm"],
 ["Archer Help","Functions/operators (SELECTEDVALUENUMBER, REF, DATEADD)","https://help.archerirm.cloud/platform_2025_04/en-us/content/platform/fields/fld_calc_functions_operators_formulas.htm"],
 ["Archer Help","Data-Driven Events","https://help.archerirm.cloud/platform_2024_11/en-us/content/platform/datadrivenevents/dde_basics.htm"],
]
table(ws,2,["Kaynak","Konu","URL"],RR,[18,48,80])

out="/home/user/rrr/mrm-archer/Model_Kritiklik_Degerlendirmesi_Archer_Field_Config.xlsx"
wb.save(out)
print("KAYDEDILDI:",out); print("Sekme:",len(wb.sheetnames)); print("Soru/alan sayisi:",len(F))
for s in wb.sheetnames: print("  -",s)
