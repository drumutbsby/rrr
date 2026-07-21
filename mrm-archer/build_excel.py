# -*- coding: utf-8 -*-
"""MRM - Archer Entegrasyonu: Excel is analizi calisma kitabi uretimi."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- Stil paleti ----------
NAVY   = "1F3864"
BLUE   = "2E5496"
LBLUE  = "D6E0F0"
GREY   = "F2F2F2"
AMBER  = "FFF2CC"
GREEN  = "E2EFDA"
WHITE  = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(row, cells, fill=NAVY, color="FFFFFF", size=10, wrap=True, height=None):
    for c in row:
        c.fill = PatternFill("solid", fgColor=fill)
        c.font = Font(bold=True, color=color, size=size, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
        c.border = border

def style_cell(c, size=10, wrap=True, top=False, fill=None, bold=False, color="000000"):
    c.font = Font(size=size, name="Calibri", bold=bold, color=color)
    c.alignment = Alignment(horizontal="left", vertical="top" if top else "center", wrap_text=wrap)
    c.border = border
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_table(ws, start_row, headers, rows, widths, hdr_fill=NAVY, zebra=True,
                col_fills=None, freeze=True):
    set_widths(ws, widths)
    r = start_row
    for j, h in enumerate(headers, 1):
        ws.cell(row=r, column=j, value=h)
    hdr(ws[r], hdr_fill)
    ws.row_dimensions[r].height = 30
    r += 1
    for i, row in enumerate(rows):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=r, column=j, value=val)
            fill = GREY if (zebra and i % 2 == 1) else None
            if col_fills and (j-1) in col_fills:
                fill = col_fills[j-1]
            style_cell(c, top=True, fill=fill)
        r += 1
    if freeze:
        ws.freeze_panes = ws.cell(row=start_row+1, column=1)
    return r

wb = openpyxl.Workbook()

# ============================================================
# 0) KAPAK / ICINDEKILER
# ============================================================
ws = wb.active
ws.title = "0. Kapak"
ws.sheet_view.showGridLines = False
set_widths(ws, [3, 40, 70, 20])
ws.merge_cells("B2:D2")
ws["B2"] = "MODEL RİSKİ YÖNETİMİ (MRM)"
ws["B2"].font = Font(size=22, bold=True, color=NAVY)
ws.merge_cells("B3:D3")
ws["B3"] = "Archer Entegrasyonu – İş Analizi Gereklilik Kitabı"
ws["B3"].font = Font(size=14, bold=True, color=BLUE)
ws.merge_cells("B4:D4")
ws["B4"] = "Hedef birim: Yazılım Geliştirme Daire Başkanlığı  |  Taslak v0.1"
ws["B4"].font = Font(size=11, italic=True, color="595959")

meta = [
    ("Doküman türü", "İş analizi / çözüm tasarımı gereklilik dokümanı (taslak)"),
    ("Kapsam", "Bankanın tüm modelleri ve algoritmaları (AI/ML/GenAI, istatistiksel, deterministik, satıcı/grup) için uçtan uca Model Riski Yönetimi süreç ve envanterinin Archer üzerinde tesisi"),
    ("Amaç", "Mevcut Archer mimarisine, uluslararası MRM çerçeveleriyle (SR 11-7, OSFI E-23, PRA SS1/23, ECB TRIM, EU AI Act, NIST AI RMF) uyumlu; Türkiye (BDDK) mevzuatı çıktığında tam uyumu sağlayacak ileriye dönük bir MRM çözümünün gereklilik tanımı"),
    ("Statü", "Regülasyon araştırmasına dayalı taslak – gözden geçirme ve onay bekliyor"),
]
r = 6
for k, v in meta:
    ws.cell(row=r, column=2, value=k).font = Font(bold=True, color=NAVY, size=11)
    c = ws.cell(row=r, column=3, value=v); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=3, end_row=r+1, end_column=4)
    ws.row_dimensions[r].height = 28
    r += 2

ws.cell(row=r+1, column=2, value="İÇİNDEKİLER (Sekmeler)").font = Font(bold=True, size=13, color=NAVY)
toc = [
    ("1. Gereklilik Matrisi", "Temel gereklilikler ve alt detaylar → Archer karşılığı, öncelik, kaynak düzenleme, faz"),
    ("2. Model Envanteri – Alanlar", "Ana model kaydı alan tasarımı (E-23 minimum standardı + SS1/23 envanter nitelikleri)"),
    ("3. Model Validasyonu – Alanlar", "Bağımsız validasyon / gözden geçirme kaydı alan tasarımı"),
    ("4. Kritiklik-Tiering Soru Seti", "Materyalite + karmaşıklık skorlama anketi ve tier eşlemesi"),
    ("5. AI-ML Ek Değerlendirme", "AI/ML/GenAI modellerine özel ek alanlar ve kontroller"),
    ("6. İzleme Metrik Kataloğu", "Sürekli izleme KPI/KRI'ları, eşikler ve tetikleyiciler"),
    ("7. PMA-İstisna Kayıtları", "Model sonrası düzeltme (PMA), kısıt ve istisna kaydı alanları"),
    ("8. Model Değişiklik – Alanlar", "Model değişiklik/versiyon yönetimi kaydı alan tasarımı"),
    ("9. Seçim Listeleri (Values List)", "Tüm açılır liste değer setleri"),
    ("10. İlişki Haritası", "Uygulamalar arası cross-reference / yeniden kullanım haritası"),
    ("11. RACI", "Rol-sorumluluk matrisi (üç savunma hattı)"),
    ("12. Yaşam Döngüsü Geçişleri", "Durum makinesi ve iş akışı geçiş kuralları"),
    ("13. Yol Haritası", "Fazlama ve teslim planı"),
    ("14. Regülasyon Eşleme", "Gereklilik ↔ düzenleme (SR11-7/E-23/SS1-23/TRIM/EU AI Act/BDDK) izlenebilirlik"),
    ("15. Referanslar", "Birincil kaynak URL'leri"),
]
r += 2
for name, desc in toc:
    ws.cell(row=r, column=2, value=name).font = Font(bold=True, color=BLUE, size=10)
    c = ws.cell(row=r, column=3, value=desc); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 22
    r += 1

# ============================================================
# 1) GEREKLILIK MATRISI
# ============================================================
ws = wb.create_sheet("1. Gereklilik Matrisi")
headers = ["ID", "Ana Gereklilik", "Alt Gereklilik / Açıklama", "Archer Karşılığı (uygulama · alan · mekanizma)", "Öncelik", "Kaynak Düzenleme", "Faz"]
R = [
 # A. MODEL TANIMI VE KAPSAM
 ["A-01","Model tanımı","Kurum genelinde tek bir 'model' tanımı benimsenir: girdi veriyi (nicel/nitel/uzman görüşü) istatistiksel/ekonomik/finansal/matematiksel teknikle işleyip çıktı üreten yöntem. Üç bileşen: girdi · işleme · raporlama/çıktı.","Politikalar ve Standartlar (MRM Politikası) · Model Envanteri: 'Model bileşenleri' referans alanları","Yüksek","SR 11-7; SS1/23 P1; OSFI E-23","F1"],
 ["A-02","Deterministik yöntem / EUC ayrımı","Deterministik/kural-tabanlı yöntemler ve karmaşık hesap tabloları (EUC) model kapsamı dışıysa; materyal + karmaşık + karar üzerinde etkiliyse MRM çerçevesinin ilgili unsurları uygulanır, aksi halde sağlam yönetim kontrolleri.","Model Envanteri: 'Deterministik yöntem' bayrağı + 'Kontrol-yalnızca' iş akışı yolu","Yüksek","SS1/23 P1 (1.1b/c)","F1"],
 ["A-03","AI/ML/GenAI bayraklama","Model tipi ve teknoloji sınıfı (AI/ML, GenAI/temel model, istatistiksel, deterministik) ayrı bayraklarla işaretlenir; AI/ML ek değerlendirme tetiklenir.","Model Envanteri: 'AI/ML', 'GenAI', 'Temel Model', 'Dinamik Model' bayrakları + DDE ile AI/ML anketini açar","Yüksek","SS1/23 (dinamik model); OSFI E-23; EU AI Act","F1"],
 ["A-04","Kapsam – tüm risk türleri","Envanter tüm risk türlerini kapsar: kredi/IRB, piyasa/VaR, operasyonel/AMA, karşı taraf/CVA, TFRS 9 BKZ, İSEDES sermaye, ALM/IRRBB, AML/suistimal, fiyatlama/değerleme, AI/GenAI.","Model Envanteri: 'Model kategorisi' seçim listesi (çok değerli)","Yüksek","OSFI E-23 (enterprise-wide); BDDK İSEDES/IRB/AMA/TFRS9","F1"],
 # B. ENVANTER
 ["B-01","Kurum geneli tekil envanter","İş kolu/tüzel kişilik ayrı envanter tutsa da kurum çapında birleşik, doğru ve güncel (evergreen) tek envanter; kullanımdaki + geliştirmedeki + emekli modeller.","Yeni uygulama: 'Model Envanteri' (kayıt seviyesi kontrol, versiyonlama)","Yüksek","SR 11-7; SS1/23 P1 (1.2); OSFI E-23","F1"],
 ["B-02","Zorunlu asgari alan seti","E-23 asgari alanları (Model ID, ad/tanım, risk derecesi, sahip, geliştirici, köken) + yüksek riskli modeller için versiyon, prod tarihi, validatör, onaylayan, bağımlılıklar, veri kaynakları, onaylı kullanımlar, kısıtlar, son/sonraki validasyon, izleme durumu.","Model Envanteri alan tasarımı (bkz. Sekme 2)","Yüksek","OSFI E-23 (min. inventory); SS1/23 (1.2c)","F1"],
 ["B-03","Model bağımlılıkları","Upstream/downstream (feeder/consumer) model bağımlılıkları; doğrudan ve dolaylı ara bağımlılıklar toplulaştırılmış model riski için izlenir.","Model Envanteri ↔ Model Envanteri self-reference (besleyen/beslenen)","Yüksek","SS1/23 (1.2b); OSFI E-23","F2"],
 ["B-04","Satıcı / grup modeli ayrımı","İç geliştirme / satıcı / grup-ana modeli ayrımı; satıcı modelde bankanın kendi kullanımını kendi çıktısıyla validasyonu ve sürekli izlemesi; grup validasyonundan yararlanma 3 koşula bağlı.","Model Envanteri: 'Köken', 'Tedarikçi' (mevcut Tedarikçi uyg. link), grup-validasyon koşul alanları","Yüksek","SS1/23 P2 (2.6); SR 11-7 (vendor)","F2"],
 ["B-05","Envanter kontrol ve denetlenebilirlik","Envanter sağlam kontrollü, denetlenebilir, güncel; emekli modeller makul süre saklanır.","History log, audit trail, kayıt izinleri; saklama politikası alanı","Orta","OSFI E-23 (evergreen); SR 11-7","F1"],
 # C. TIERING
 ["C-01","Kurum geneli tutarlı tiering","Her modele risk-tabanlı materyalite + karmaşıklık derecesi atayan tutarlı yaklaşım.","Yeni anket: 'Model Kritiklik/Materyalite Değerlendirmesi' (skorlu) → Envanter'de Tier alanını hesaplar","Yüksek","SS1/23 P1 (1.3); SR 11-7; OSFI E-23","F1"],
 ["C-02","Materyalite (nicel + nitel)","Nicel: risk tutarı/portföy büyüklüğü/müşteri sayısı. Nitel: amaç, kararlara etki, ödeme gücü/finansal performans etkisi.","Kritiklik anketi: materyalite blok soruları + ağırlıklar","Yüksek","SS1/23 (1.3b); OSFI E-23","F1"],
 ["C-03","Karmaşıklık (AI dahil)","Girdi veri niteliği/kalitesi, metodoloji/varsayım seçimi, uygulama bütünlüğü, kullanım sıklığı; ilgili yerde alternatif/yapılandırılmamış veri ve açıklanabilirlik/şeffaflık/yanlılık.","Kritiklik anketi: karmaşıklık blok soruları + AI karmaşıklık faktörleri","Yüksek","SS1/23 (1.3c); EU AI Act; NIST","F1"],
 ["C-04","İçsel risk = zafiyet × materyalite","İçsel (inherent) model risk skoru; artık (residual) risk azaltıcılardan sonra ayrı raporlanır.","Hesaplanan alan: İçsel risk skoru; ayrı 'Artık risk' alanı","Yüksek","OSFI E-23 (rating)","F1"],
 ["C-05","Tier → validasyon derinlik/frekans","Validasyon/izleme/revalidasyon kapsam-yoğunluk-frekansı tier ile orantılı (regülasyon sabit sayı dayatmaz; frekans banka politikasıdır).","Values List: Tier→frekans eşlemesi (politika parametresi); hesaplanan 'Sonraki validasyon tarihi'","Yüksek","SS1/23 (4.2b/4.4d/4.5b); OSFI E-23","F1"],
 ["C-06","Tier periyodik bağımsız gözden geçirme","Tiering yaklaşımı periyodik validasyon/kritik gözden geçirmeye; bireysel tier atamaları validasyon/revalidasyonda bağımsızca yeniden değerlendirilir.","Kritiklik anketi periyodik yeniden çalıştırma + validasyonda tier teyit alanı","Orta","SS1/23 (1.3d/1.3e)","F2"],
 ["C-07","'İhmal edilebilir risk' kategorisi","Negligible-risk modelleri tam yaşam döngüsü yönetiminden muaf tutabilen kategori; yine de envanterde.","Values List: Tier = 'İhmal edilebilir'; DDE ile hafif iş akışı","Orta","OSFI E-23","F2"],
 # D. GELISTIRME
 ["D-01","Amaç ve tasarım beyanı","Her modelin net amaç/tasarım hedefi; kavramsal sağlamlık; literatür/endüstri desteği; alternatif yaklaşım/benchmark ile kıyas.","Model Envanteri: amaç, tasarım hedefi, kavramsal dayanak alanları; dokümantasyon linkleri","Yüksek","SR 11-7; SS1/23 P3 (3.1); OSFI E-23","F2"],
 ["D-02","Veri uygunluğu ve temsil","Geliştirme verisi uygun, tutarlı, temsil edici; uygunsuz yanlılık yok; veri gizliliği uyumu; temsil zayıfsa etki değerlendirilip tier'e yansıtılır.","Model Envanteri/Validasyon: veri kaynakları, temsil değerlendirmesi, proxy, veri kalitesi test sonuçları alanları","Yüksek","SS1/23 (3.2); OSFI E-23; ECB TRIM","F2"],
 ["D-03","Geliştirme testleri","Geriye ve ileriye dönük testler + duyarlılık analizi ile operasyonel sınırlar; challenger/benchmark kıyası; materyal değişiklik ve dinamik modellerde kümülatif değişiklikte paralel çıktı analizi.","Validasyon kaydı: test türleri; 'İzleme paketi' (monitoring pack) eki","Yüksek","SR 11-7; SS1/23 (3.3); ECB TRIM","F2"],
 ["D-04","Model düzeltmeleri / uzman görüşü","Düzeltme/overlay gerekçelendirilir, envantere kaydedilir (neden + zaman içindeki hesap), bağımsız validasyona tabi; besleyen modele etkisi değerlendirilir.","Model Envanteri: 'Geliştirme düzeltmeleri' alt-formu; PMA kaydıyla ilişki","Orta","SS1/23 (3.4); SR 11-7","F2"],
 ["D-05","Geliştirme dokümantasyonu","Bağımsız uzmanın modeli anlayıp sonuçları yeniden üretebileceği kadar kapsamlı dokümantasyon; satıcı modelde bankanın kullanımını validasyona yetecek detay.","Model Envanteri: dokümantasyon türü seçim listesi + belge ekleri/URL alanları","Yüksek","SR 11-7; SS1/23 (3.5); OSFI E-23","F2"],
 ["D-06","Destekleyici sistem / uygulama ortamı","Model, iyi test edilmiş ve değişiklik-kontrollü bir bilgi sisteminde uygulanır; sistem uygunluğu periyodik yeniden değerlendirilir.","Model Envanteri ↔ mevcut Uygulamalar/Donanımlar/IT Hizmetleri linkleri; sistem test bulgu alanı","Orta","SR 11-7; SS1/23 (3.6); OSFI E-23 (deployment)","F3"],
 # E. VALIDASYON
 ["E-01","Bağımsız validasyon fonksiyonu","Geliştirme ve model sahibinden bağımsız; sermaye-modeli bankalarında ayrı raporlama hattı ile kanıtlanan bağımsızlık; board/komiteye erişim.","Yeni uygulama: 'Model Validasyonu' + Access Roles (2. hat) ayrımı","Yüksek","SR 11-7; SS1/23 P4 (4.1); ECB TRIM","F1"],
 ["E-02","Kavramsal sağlamlık değerlendirmesi","Teori/mantık/varsayım/veri kalitesi ve geliştirme kanıtının kalitesi; tüm bileşenler (girdi/işleme/çıktı); duyarlılık analizi.","Model Validasyonu: kavramsal sağlamlık bölümü (yapılandırılmış alanlar)","Yüksek","SR 11-7; SS1/23 (4.2)","F1"],
 ["E-03","Süreç doğrulama (process verification)","Girdilerin temsil ve veri kalitesine uygunluğu; hesaplama/kod/entegrasyon doğruluğu; çıktı raporlarının doğru/tam/amaca uygun olması.","Model Validasyonu: süreç doğrulama bölümü","Yüksek","SR 11-7; SS1/23 (4.3); ECB TRIM (IT/implementasyon)","F2"],
 ["E-04","Performans izleme ve çıktı analizi","Kabul edilebilir performans eşiklerine karşı izleme; benchmarking, duyarlılık, override analizi, paralel çıktı analizi/backtesting; bağımsız gözden geçirilen izleme raporları.","Model Validasyonu + Metrik Sonuçları (mevcut) entegrasyonu; eşik aşımı → Bulgu","Yüksek","SR 11-7; SS1/23 (4.4); ECB TRIM","F2"],
 ["E-05","Periyodik revalidasyon","Tier ile tutarlı frekansta bağımsız revalidasyon; önceki bulguların geçerliliği.","Model Validasyonu: 'Validasyon türü = Revalidasyon'; takvim/bildirim","Yüksek","SR 11-7; SS1/23 (4.5); ECB TRIM (yıllık)","F2"],
 ["E-06","Validasyon vs onay ayrımı","Validasyon fonksiyonu öneri üretir; onay yetkisi ayrı onay makamındadır (komite/birey). İki ayrı kayıt/rol.","Model Validasyonu: 'Validasyon sonucu (öneri)' ve 'Onay kararı' ayrı alanlar; Advanced Workflow onay adımı","Yüksek","SS1/23 (4.1b, 2.13); OSFI E-23 (approver rolü)","F1"],
 ["E-07","Etkin sorgulama (effective challenge)","Yetkin, bağımsız, teşvik ve etkiye sahip taraflarca kritik analiz; sınırlıysa telafi edici kontrol.","Model Validasyonu: etkin sorgulama notları; bağımsızlık teyidi alanı","Yüksek","SR 11-7; BDDK Rehber #16","F1"],
 ["E-08","Validasyon bulguları ve takibi","Bulgu/sınırlama/remediation dokümante edilir, sahip ve üst yönetime raporlanır, açık konular izlenir.","Mevcut 'Bulgular' uygulaması: 'Bulgu Kaynağı = Model Validasyonu' + 'Aksiyon Planları' ile remediation","Yüksek","SR 11-7; SS1/23; ECB TRIM","F1"],
 # F. IZLEME
 ["F-01","İzleme metrik seti ve eşikler","Ayrımcılık (AUC/Gini/KS), kalibrasyon, stabilite/drift (PSI/CSI), fairness, override oranı; eşikler ve çok-KPI karar kuralı.","Mevcut 'Metrikler/Metrik Sonuçları' – MRM metrik seti; Model Envanteri'ne bağlı","Yüksek","SR 11-7; ECB TRIM; NIST; EU AI Act","F2"],
 ["F-02","Backtesting / geriye dönük test","Model tahmini vs gerçekleşen; istatistiksel uygun testler; anomali → ileri analiz.","Metrik Sonuçları: backtesting metrikleri; eşik aşımı DDE","Yüksek","ECB TRIM; BDDK İç Sist. Md.26/Rehber #16","F2"],
 ["F-03","Benchmarking / kıyaslama","Temsil edici, karşılaştırılabilir dış/iç kaynaklarla kıyas (özellikle düşük temerrütlü portföyler).","Metrik/Validasyon: benchmark referans alanları","Orta","ECB TRIM; SR 11-7","F3"],
 ["F-04","Drift ve bozulma izleme","Veri driftin (PSI/CSI) ve kavram driftinin sürekli izlenmesi; bozulma eşiği → revalidasyon/retrain tetikler.","AI/ML izleme metrikleri; DDE ile Validasyon/Değişiklik tetikleme","Yüksek","EU AI Act (Art.15); NIST; EBA ML-IRB","F3"],
 # G. MITIGANTLAR
 ["G-01","Model sonrası düzeltmeler (PMA)","PMA/overlay/override kurum geneli tutarlı süreç; nasıl hesaplandığı, ne zaman azaltılıp kaldırılacağı; materyal modellerde üst yönetim/komite onayı; bağımsız gözden geçirme; PMA'lı ve PMA'sız raporlama.","Yeni uygulama/alt-form: 'PMA-İstisna Kayıtları' (bkz. Sekme 7)","Yüksek","SS1/23 P5 (5.1)","F2"],
 ["G-02","Kullanım kısıtları","Ciddi eksik/hata bulunduğunda sıkı kontrol altında kullanım veya kullanım limiti/yasağı; konu ve remediation envanterde izlenir.","Model Envanteri: 'Kullanım kısıtı/limit' alanları; durum = 'Kısıtlı'","Yüksek","SS1/23 (5.2); OSFI E-23","F2"],
 ["G-03","İstisnalar ve eskalasyon","İstisna tanımı (onaysız/validasyonsuz kullanım, amaç dışı, kalıcı eşik ihlali, tutarsız backtest); resmi onaylı eskalasyon; paydaşlara zamanında bildirim.","PMA-İstisna kaydı: istisna türü, tolerans, eskalasyon zaman damgaları; Bildirimler","Yüksek","SS1/23 (5.3)","F2"],
 # H. YONETISIM
 ["H-01","Board / üst yönetim gözetimi","Board MRM çerçevesini kurar, model risk iştahını belirler, profil vs iştah düzenli rapor alır, en materyal modellerin çıktısını sorgular.","MRM Dashboard (board paneli); Risk İştahı referans kaydı; raporlama iView'ları","Yüksek","SR 11-7; SS1/23 P2 (2.1); OSFI E-23","F3"],
 ["H-02","Sorumlu üst yönetici (SMF/hesap verebilirlik)","MRM çerçevesinden genel sorumlu, en kıdemli birey belirlenir; sorumluluk beyanı güncellenir.","Model Envanteri/Politika: 'MRM sorumlu yönetici' alanı; organizasyon linki","Yüksek","SS1/23 P2 (2.2)","F1"],
 ["H-03","MRM politikası ve prosedürler","Board onaylı politika; model & model riski tanımı, tiering, geliştirme standartları, veri kalitesi, validasyon standartları, performans izleme, mitigantlar, onay & değişiklik.","Mevcut 'Politikalar ve Standartlar' uyg.: MRM politikası kaydı + Envanter'e bağ","Yüksek","SR 11-7; SS1/23 (2.3); OSFI E-23","F1"],
 ["H-04","Roller ve sorumluluklar","Owner, developer, user, validator, approver, stakeholder rolleri; her yaşam döngüsü aşamasında tanımlı; gerekli yetkinlik/bağımsızlık.","Model Envanteri: rol referans alanları (Organizasyon/Kişi/İletişim linkleri); RACI (Sekme 11)","Yüksek","SR 11-7; SS1/23 (2.4); OSFI E-23","F1"],
 ["H-05","Üç savunma hattı & iç denetim","1. hat (sahip/geliştirici/kullanıcı), 2. hat (MRM/bağımsız validasyon), 3. hat (iç denetim çerçeve etkinliğini denetler).","Access Roles ayrımı; iç denetim gözden geçirme kaydı; mevcut Denetlenen Görüşleri","Yüksek","SR 11-7; SS1/23 (2.5); BDDK İç Sist. Yön.","F1"],
 ["H-06","Yıllık öz-değerlendirme / attestation","MRM çerçevesinin 5 ilkeye karşı yıllık öz-değerlendirmesi + remediation planı; board'a raporlama; finansal raporlama için MRM etkinlik raporu (yıllık, denetim komitesine).","Yeni anket: 'MRM Yıllık Öz-Değerlendirme'; görev/bildirim ile yıllık döngü","Orta","SS1/23 (1.5-1.8, 3.7-3.8)","F3"],
 # I. AI/ML EK
 ["I-01","Açıklanabilirlik / yorumlanabilirlik","Algoritma ailesine uygun global+lokal açıklanabilirlik; opak modelde post-hoc (SHAP/LIME) ve yöntem sınırları; kredi kararında red-gerekçesi (reason code).","AI/ML Ek Değerlendirme: açıklanabilirlik yöntemi, reason-code yeteneği, model kartı alanları","Yüksek","EU AI Act (Art.13); NIST; EY","F2"],
 ["I-02","Yanlılık ve adalet (bias/fairness)","Korumalı gruplar bazında dağıtım öncesi ve sürekli fairness testi; eğitim verisi temsil incelemesi; metrik+eşik; 4/5 (disparate impact) kuralı.","AI/ML Ek: fairness metrikleri, korumalı gruplar, disparate impact oranı, mitigasyon","Yüksek","EU AI Act (Art.10); MAS FEAT; Fairlearn","F2"],
 ["I-03","Veri lineage / provenance","Veri kökeni, toplama koşulu/tarihi, işleme/etiketleme, lisans/haklar; loglama.","AI/ML Ek: veri lineage referansı, kaynak sicili, etiketleme kalite kaydı","Orta","EU AI Act (Art.10,12); OSFI E-23","F3"],
 ["I-04","Robustluk / reproducibility / güvenlik","Pertürbasyon robustluk testi; sabit seed/ortam/versiyon ile tekrar üretilebilirlik; adversarial + (GenAI) prompt-injection/jailbreak testleri; PII sızıntı taraması.","AI/ML Ek: robustluk/adversarial/reproducibility referans alanları","Orta","EU AI Act (Art.15); NIST; WF GenAI","F3"],
 ["I-05","Retraining yönetişimi","Takvim + olay (drift/bozulma) tetikli retraining politikası; versiyon kontrolü; materyal retrain'de revalidasyon.","AI/ML Ek: retrain frekansı, son retrain, tetik tipi; Model Değişiklik kaydı","Yüksek","EBA ML-IRB; OSFI E-23","F3"],
 ["I-06","İnsan gözetimi (human-in-the-loop)","Tasarımda insan gözetimi; yüksek riskli çıktıda insan incelemesi olmadan otomatik akışa girmez; gözetim rolleri eğitimli.","AI/ML Ek: insan gözetimi bayrağı + rol; onaylı kullanım kapsamı","Yüksek","EU AI Act (Art.14); WF GenAI","F3"],
 ["I-07","Etik / temel haklar değerlendirmesi","Sonuç-doğuran/müşteriyi etkileyen kullanımda etik/sorumlu-AI incelemesi; AB yüksek-riskte FRIA.","AI/ML Ek: etik inceleme / FRIA referansı; EU AI Act risk kategorisi","Orta","EU AI Act (Art.27); MAS FEAT; NIST GOVERN","F3"],
 ["I-08","Temel model / 3. taraf bağımlılık (GenAI)","Temel model provenance/risk profili, lisans/IP/telif, barındırma/sızıntı, konsantrasyon, versiyon bildirimi.","AI/ML Ek: temel model adı/versiyon/sağlayıcı, lisans, satıcı risk notu, çıkış planı","Orta","EU AI Act (Art.25); NIST; FSB","F3"],
 ["I-09","GenAI prompt/çıktı kontrolleri","Girdi kontrolleri (zararlı prompt engelleme, şablon); çıktı guardrail'leri (hallucination/toksisite tarama); etkileşim loglama; halüsinasyon metrikleri (NLI/SelfCheckGPT).","AI/ML Ek: guardrail modelleri, halüsinasyon/toksisite metrik+eşik, prompt şablon sicili","Orta","WF GenAI; NIST AI 600-1; HKMA","F4"],
 # J. VERI KALITESI
 ["J-01","Veri kalitesi boyutları","Tamlık, doğruluk, tutarlılık, güncellik, teklik, geçerlilik, erişilebilirlik, izlenebilirlik; eşik + RAG göstergeleri.","Veri kalitesi test sonucu alanları; DQF referansı; metrik göstergeleri","Orta","ECB TRIM; BDDK İç Sist. Md.26","F3"],
 ["J-02","Veri kalitesi çerçevesi (DQF) & BCBS 239 hazırlık","Board onaylı, dokümante DQF; bağımsız değerlendirme; risk verisi toplulaştırma yeteneği (BCBS 239'a ileriye dönük hazırlık).","DQF politika kaydı; veri sahipliği (iş+IT) alanları; issue takibi (Bulgular)","Orta","ECB TRIM; BCBS 239; BDDK ECL Rehberi","F4"],
 # K. RAPORLAMA / BDDK
 ["K-01","Toplulaştırılmış model riski raporlama","Bireysel + toplu model riski; iştaha karşı profil; board/komite raporları; ısı haritası.","MRM Dashboard iView'ları; toplulaştırma raporları","Orta","SR 11-7; SS1/23 (2.1d); OSFI E-23","F3"],
 ["K-02","BDDK İSEDES / validasyon raporu üretimi","İçsel Model Validasyon raporu ve İSEDES model bölümlerini talep üzerine üretebilme; her artefaktın ilgili mevzuat maddesine eşlemesi.","Rapor şablonları/iView; Regülasyon eşleme (Sekme 14) alanları","Yüksek","BDDK İç Sist. Yön. (İSEDES); Rehber #13/#16","F2"],
 # L. ARCHER TEKNIK
 ["L-01","Seçim listeleri (Values Lists)","Model tipi, kategori, yaşam döngüsü durumu, tier, validasyon türü/sonucu, risk derecesi vb. standart listeler.","Global/uygulama Values Lists (bkz. Sekme 9)","Yüksek","Tasarım","F1"],
 ["L-02","Gelişmiş iş akışı (Advanced Workflow)","Validasyon onay akışı: Geliştirici → Bağımsız Validasyon → Model Risk Komitesi → Onay/Ret/Koşullu.","Advanced Workflow; durum makinesi (Sekme 12)","Yüksek","SS1/23 P4; OSFI E-23","F1"],
 ["L-03","Hesaplanan alanlar","Sonraki validasyon tarihi (= son + tier frekansı), içsel risk skoru, gecikme bayrağı, yaş.","Calculated fields","Yüksek","Tasarım","F1"],
 ["L-04","Veri-güdümlü olaylar (DDE)","Durum geçişi, alan görünürlüğü/koşullu zorunluluk, validasyon başarısızsa bulgu oluşturma.","Data-Driven Events","Yüksek","Tasarım","F2"],
 ["L-05","Bildirimler","Yaklaşan/geciken revalidasyon, izleme eşik aşımı, açık bulgu SLA, PMA kaldırma tetiği.","Notifications + Görev Yönetimi (mevcut)","Yüksek","Tasarım","F2"],
 ["L-06","Veri beslemeleri (Data Feeds)","MLOps/model geliştirme ortamından envanter ve metrik beslemesi (opsiyonel entegrasyon).","Data Feeds / API entegrasyonu","Düşük","Tasarım","F4"],
 ["L-07","Erişim rolleri / yetkilendirme","1./2./3. hat ayrımı; alan/kayıt seviyesi güvenlik; validasyon bağımsızlığının sistemsel garantisi.","Access Roles, record permissions, field-level security","Yüksek","SR 11-7; SS1/23; ECB TRIM","F1"],
 ["L-08","Dashboard / iView'lar","Envanter kırılımı, validasyon takvimi/gecikmeler, açık bulgular, model riski ısı haritası, AI/ML portföyü.","MRM Workspace Dashboards (iApp + iView)","Orta","Tasarım","F3"],
]
end = write_table(ws, 1, headers, R, [8, 22, 52, 46, 10, 26, 6],
                  col_fills=None)
# Oncelik renklendirme
prio_fill = {"Yüksek": "F8CBAD", "Orta": "FFE699", "Düşük": "C6E0B4"}
for row in ws.iter_rows(min_row=2, max_row=end-1, min_col=5, max_col=5):
    for c in row:
        if c.value in prio_fill:
            c.fill = PatternFill("solid", fgColor=prio_fill[c.value])
            c.alignment = Alignment(horizontal="center", vertical="top")

# ============================================================
# 2) MODEL ENVANTERI - ALANLAR
# ============================================================
def field_sheet(title, intro, rows):
    ws = wb.create_sheet(title)
    ws.merge_cells("A1:H1")
    ws["A1"] = intro
    ws["A1"].font = Font(bold=True, size=11, color=NAVY)
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 34
    headers = ["Grup", "Alan Adı", "Alan Tipi (Archer)", "Seçim Listesi / Değerler", "Zorunlu", "Hesaplanan / Çapraz Referans", "Açıklama"]
    write_table(ws, 3, headers, rows, [20, 30, 20, 30, 9, 30, 44])
    # zorunlu renk
    for row in ws.iter_rows(min_row=4, max_row=3+len(rows), min_col=5, max_col=5):
        for c in row:
            if c.value == "Evet":
                c.fill = PatternFill("solid", fgColor="F8CBAD")
            elif c.value == "Koşullu":
                c.fill = PatternFill("solid", fgColor="FFE699")
            c.alignment = Alignment(horizontal="center", vertical="top")
    return ws

inv_rows = [
 ["Kimlik","Model ID","Text (auto)","MDL-#####","Evet","Anahtar alan","Sistem üretimi tekil kimlik"],
 ["Kimlik","Model Adı","Text","","Evet","","Kısa ad"],
 ["Kimlik","Model Tanımı / Temel Özellikler","Text (area)","","Evet","","Amaç ve temel işlev özeti"],
 ["Kimlik","Model Versiyonu","Text","","Koşullu","","Yüksek riskli modellerde zorunlu"],
 ["Sınıflandırma","Model Tipi","Values List","İstatistiksel; AI/ML; GenAI/Temel Model; Deterministik/Kural; Hibrit","Evet","","Teknoloji sınıfı"],
 ["Sınıflandırma","Model Kategorisi (risk türü)","Values List (çoklu)","Kredi/IRB; Piyasa/VaR; Operasyonel/AMA; Karşı Taraf/CVA; TFRS9 BKZ; İSEDES Sermaye; ALM/IRRBB; AML/Suistimal; Fiyatlama/Değerleme; Diğer","Evet","","Kullanım alanı"],
 ["Sınıflandırma","AI/ML Bayrağı","Values List (Evet/Hayır)","Evet; Hayır","Evet","DDE → AI/ML Ek Değerlendirme","Ek anketi tetikler"],
 ["Sınıflandırma","GenAI / Temel Model Bayrağı","Values List (Evet/Hayır)","Evet; Hayır","Koşullu","","GenAI ek kontrolleri"],
 ["Sınıflandırma","Dinamik Model Bayrağı","Values List (Evet/Hayır)","Evet; Hayır","Hayır","","Otonom yeniden kalibrasyon/parametre"],
 ["Sınıflandırma","Deterministik Yöntem / EUC Bayrağı","Values List (Evet/Hayır)","Evet; Hayır","Hayır","DDE → kontrol-yalnızca yol","Model kapsamı dışı ayrım"],
 ["Köken","Model Kökeni","Values List","İç Geliştirme; Satıcı/Vendor; Grup-Ana Model","Evet","","Sahiplik kaynağı"],
 ["Köken","Tedarikçi","Cross-Reference","→ Tedarikçi (mevcut)","Koşullu","Tedarikçi uygulaması","Satıcı modelde zorunlu"],
 ["Köken","Grup Validasyonu Koşulları","Text (area)","","Koşullu","","3 koşul (SS1/23 2.6c) kanıtı"],
 ["Amaç/Tasarım","Amaç ve Onaylı Kullanım","Text (area)","","Evet","","Hedeflenen kullanım"],
 ["Amaç/Tasarım","Fiili Kullanım","Text (area)","","Orta","","Amaç dışı kullanım tespiti"],
 ["Amaç/Tasarım","Model Operasyonel Sınırları","Text (area)","","Orta","","Geçerli girdi/uygulama sınırları"],
 ["Amaç/Tasarım","Tasarım Hedefi / Kavramsal Dayanak","Text (area)","","Orta","","Literatür/endüstri referansı"],
 ["Varsayım/Sınır","Varsayımlar","Text (area)","","Orta","","Temel varsayımlar"],
 ["Varsayım/Sınır","Sınırlamalar / Yakalanamayan Riskler","Text (area)","","Orta","","Bilinen kısıtlar"],
 ["Kritiklik","Tier (Materyalite/Kritiklik)","Values List","Tier 1; Tier 2; Tier 3; İhmal Edilebilir","Evet","← Kritiklik anketinden hesaplanır","Yönlendirici tier"],
 ["Kritiklik","Materyalite Skoru","Numeric","","Hayır","Hesaplanan","Nicel+nitel"],
 ["Kritiklik","Karmaşıklık Skoru","Numeric","","Hayır","Hesaplanan","AI faktörleri dahil"],
 ["Kritiklik","İçsel (Inherent) Risk Skoru","Numeric","","Hayır","Hesaplanan (zafiyet×materyalite)","Doğal risk"],
 ["Kritiklik","Artık (Residual) Risk","Values List","Yüksek; Orta; Düşük","Hayır","","Azaltıcılar sonrası"],
 ["Roller","Model Sahibi","Cross-Reference","→ Organizasyon / Kontaklar","Evet","İş Hiyerarşisi","Hesap veren birey/birim"],
 ["Roller","Model Geliştirici","Cross-Reference","→ Kontaklar/Organizasyon","Evet","","Geliştiren"],
 ["Roller","Model Kullanıcısı","Cross-Reference","→ Kontaklar/Organizasyon","Orta","","Çıktıyı kullanan"],
 ["Roller","Model Validatörü","Cross-Reference","→ Kontaklar (2. hat)","Koşullu","","Bağımsız validatör"],
 ["Roller","Onay Makamı","Cross-Reference / Values List","Model Risk Komitesi; Üst Yönetim","Koşullu","","Onaylayan"],
 ["Roller","MRM Sorumlu Yöneticisi","Cross-Reference","→ Organizasyon","Orta","","SMF/hesap verebilirlik"],
 ["Bağımlılık","Besleyen Modeller (Upstream)","Cross-Reference (self)","→ Model Envanteri","Hayır","Self-reference","Girdi sağlayan modeller"],
 ["Bağımlılık","Beslenen Modeller (Downstream)","Cross-Reference (self)","→ Model Envanteri","Hayır","Self-reference","Çıktıyı kullanan modeller"],
 ["Sistem/Veri","Barındıran Sistem / Uygulama","Cross-Reference","→ Uygulamalar / IT Hizmetleri","Orta","BT Altyapısı","Prod ortamı"],
 ["Sistem/Veri","Donanım","Cross-Reference","→ Donanımlar","Hayır","","Altyapı"],
 ["Sistem/Veri","Girdi Veri Kaynakları","Cross-Reference / Text","→ Bilgi Varlıkları","Orta","Bilgi Güvenliği","Veri varlıkları"],
 ["Sistem/Veri","Veri Kalitesi Test Sonucu","Text/Attachment","","Hayır","","Kalite boyutları özeti"],
 ["Düzenleme","İlgili Yasal Mevzuat","Cross-Reference","→ Yasal Mevzuatlar","Orta","Uyumluluk","Regülasyon eşleme"],
 ["Düzenleme","MRM Politikası","Cross-Reference","→ Politikalar ve Standartlar","Orta","","Bağlı politika"],
 ["Yaşam Döngüsü","Model Durumu","Values List","Taslak; Geliştirme; Validasyonda; Onayda; Onaylı/Prod; İzlemede; Revalidasyonda; Koşullu Onaylı; Kısıtlı/Askıda; Emekli","Evet","Durum makinesi (Sekme 12)","Yaşam döngüsü aşaması"],
 ["Yaşam Döngüsü","Prod'a Alınma Tarihi","Date","","Koşullu","","Go-live"],
 ["Validasyon","Son Validasyon Tarihi","Date","","Koşullu","← Model Validasyonu","En son validasyon"],
 ["Validasyon","Validasyon Frekansı","Values List","Yıllık; 2 Yıllık; 3 Yıllık; Olay-Tetikli","Koşullu","Tier eşlemesi","Politika parametresi"],
 ["Validasyon","Sonraki Validasyon Tarihi","Date","","Hayır","Hesaplanan (son+frekans)","Takvim/bildirim"],
 ["Validasyon","Gecikme Bayrağı","Values List","Zamanında; Yaklaşan; Gecikmiş","Hayır","Hesaplanan","SLA izleme"],
 ["Validasyon","Açık Bulgular / Remediation","Cross-Reference","→ Bulgular / Aksiyon Planları","Hayır","Bulgu Yönetimi","Sağlık göstergesi"],
 ["Kısıt","Kullanım Kısıtı / Limit","Text (area)","","Hayır","","Kısıtlı kullanım şartları"],
 ["Kısıt","PMA / İstisna Kayıtları","Cross-Reference","→ PMA-İstisna","Hayır","","Azaltıcılar"],
 ["Dokümantasyon","Model Dokümantasyonu","Attachment / URL","","Orta","","Geliştirme belgesi"],
 ["Dokümantasyon","Model Kartı / Teknik Dosya","Attachment / URL","","Koşullu","","AI/ML (Annex IV)"],
 ["Saklama","Emeklilik Gerekçesi","Text (area)","","Koşullu","","Devre dışı nedeni"],
 ["Saklama","Saklama Süresi Sonu","Date","","Hayır","","Emekli model saklama"],
]
field_sheet("2. Model Envanteri", "MODEL ENVANTERİ – Alan Tasarımı  (OSFI E-23 asgari envanter standardı + PRA SS1/23 envanter nitelikleri; mevcut Archer uygulamalarına cross-reference'lı)", inv_rows)

# ============================================================
# 3) MODEL VALIDASYONU - ALANLAR
# ============================================================
val_rows = [
 ["Kimlik","Validasyon ID","Text (auto)","VAL-#####","Evet","Anahtar","Tekil kimlik"],
 ["Kimlik","İlgili Model","Cross-Reference","→ Model Envanteri","Evet","1..N","Validasyona konu model(ler)"],
 ["Kimlik","Validasyon Türü","Values List","İlk (Initial) Validasyon; Bağımsız Gözden Geçirme; Süreç Doğrulama; Performans İzleme; Periyodik Revalidasyon; Hedefli (materyal değişiklik)","Evet","","Çalışma tipi"],
 ["Kimlik","Kapsam (tier'e göre)","Text (area)","","Evet","Tier'den türetilir","Kapsam/yoğunluk"],
 ["Bağımsızlık","Validatör","Cross-Reference","→ Kontaklar (2. hat)","Evet","Access role ayrımı","Bağımsız birey/birim"],
 ["Bağımsızlık","Bağımsızlık Teyidi","Values List","Tam Bağımsız; Ayrı Rapor Hattı; Telafi Edici Kontrol","Evet","","Bağımsızlık kanıtı"],
 ["Kavramsal Sağlamlık","Teori/Metodoloji Değerlendirmesi","Text (area)","","Evet","","Kavramsal sağlamlık"],
 ["Kavramsal Sağlamlık","Varsayım / Duyarlılık Analizi","Text (area)","","Orta","","Varsayım testleri"],
 ["Kavramsal Sağlamlık","Veri Kalitesi / Temsil Değerlendirmesi","Text (area)","","Orta","","Veri uygunluğu"],
 ["Süreç Doğrulama","Girdi Doğrulama","Text (area)","","Orta","","Girdi temsil/kalite"],
 ["Süreç Doğrulama","Hesaplama / Kod / Entegrasyon Doğrulama","Text (area)","","Orta","","Uygulama doğruluğu"],
 ["Süreç Doğrulama","Çıktı / Raporlama Doğrulama","Text (area)","","Orta","","Çıktı doğru/tam/amaca uygun"],
 ["Çıktı Analizi","Backtesting Sonuçları","Text/Numeric","","Koşullu","Metrik Sonuçları","Geriye dönük test"],
 ["Çıktı Analizi","Benchmarking Sonuçları","Text (area)","","Orta","","Kıyaslama"],
 ["Çıktı Analizi","Override / Paralel Çıktı Analizi","Text (area)","","Orta","","İzleme testleri"],
 ["Etkin Sorgulama","Etkin Sorgulama Notları","Text (area)","","Evet","","Kritik analiz/challenge"],
 ["Sonuç","Validasyon Sonucu (öneri)","Values List","Onay Önerilir; Koşullu Onay Önerilir; Ret Önerilir","Evet","","Validasyon FONKSİYONU önerisi"],
 ["Sonuç","Validasyon Derecesi/Rating","Values List","Tatmin Edici; Kısmen Tatmin; Tatmin Etmeyen","Orta","","Kalite notu"],
 ["Sonuç","Uygulanan Koşullar / Kısıtlar","Text (area)","","Koşullu","","Koşullu onay şartları"],
 ["Onay","Onay Kararı","Values List","Onaylandı; Koşullu Onaylandı; Reddedildi; Beklemede","Evet","AYRI alan (onay makamı)","Validasyondan bağımsız karar"],
 ["Onay","Onay Makamı","Cross-Reference / Values List","Model Risk Komitesi; Üst Yönetim","Koşullu","Advanced Workflow","Karar mercii"],
 ["Onay","Onay Tarihi","Date","","Koşullu","","Karar tarihi"],
 ["Bulgular","Üretilen Bulgular","Cross-Reference","→ Bulgular","Hayır","Bulgu Yönetimi","Validasyon bulguları"],
 ["Raporlama","Validasyon Raporu","Attachment / URL","","Evet","","Rapor belgesi"],
 ["Raporlama","Rapor Tarihi","Date","","Evet","","Tamamlanma"],
 ["Raporlama","Sonraki Revalidasyon Tarihi","Date","","Hayır","Hesaplanan","Takvim"],
]
field_sheet("3. Model Validasyonu", "MODEL VALİDASYONU – Alan Tasarımı  (SR 11-7 üç bileşenli validasyon + SS1/23 P4 + ECB TRIM; validasyon ÖNERİSİ ile ONAY kararı ayrı alanlar)", val_rows)

# ============================================================
# 4) KRITIKLIK / TIERING SORU SETI
# ============================================================
ws = wb.create_sheet("4. Kritiklik-Tiering")
ws.merge_cells("A1:G1")
ws["A1"] = ("MODEL KRİTİKLİK / MATERYALİTE DEĞERLENDİRME ANKETİ  –  Skorlu; Materyalite + Karmaşıklık boyutları "
            "→ Tier (SS1/23 P1 1.3, OSFI E-23 rating). Not: Tier→frekans eşlemesi banka POLİTİKASIDIR (regülasyon sabit sayı dayatmaz).")
ws["A1"].font = Font(bold=True, size=11, color=NAVY)
ws["A1"].alignment = Alignment(wrap_text=True, vertical="center")
ws.row_dimensions[1].height = 44
q_headers = ["Blok", "No", "Soru / Faktör", "Cevap Tipi / Ölçek", "Ağırlık", "Puanlama Notu"]
q_rows = [
 ["Materyalite","M1","Modelin etkilediği risk tutarı / portföy büyüklüğü / bilanço etkisi","1-5 (çok düşük→çok yüksek)","20%","Nicel eşikler politikada tanımlanır"],
 ["Materyalite","M2","Etkilenen müşteri / işlem sayısı","1-5","10%",""],
 ["Materyalite","M3","Kararlara ve iş sonuçlarına etkisi (fiyatlama, karşılık, sermaye, onay)","1-5","15%","Nitel"],
 ["Materyalite","M4","Ödeme gücü / yasal sermaye / finansal raporlamaya etki","1-5","15%","Regülasyon kullanımı ağırlıklı"],
 ["Materyalite","M5","İtibar / müşteri / davranışsal etki","1-5","5%",""],
 ["Karmaşıklık","K1","Girdi verinin niteliği ve kalitesi (alternatif/yapılandırılmamış veri?)","1-5","8%","AI faktörü"],
 ["Karmaşıklık","K2","Metodoloji / varsayım karmaşıklığı","1-5","7%",""],
 ["Karmaşıklık","K3","Uygulama bütünlüğü / entegrasyon karmaşıklığı","1-5","5%",""],
 ["Karmaşıklık","K4","Kullanım sıklığı ve yaygınlığı","1-5","5%",""],
 ["Karmaşıklık","K5","Açıklanabilirlik / şeffaflık / yorumlanabilirlik (opaklık)","1-5","5%","AI/ML"],
 ["Karmaşıklık","K6","Tasarımcı / veri yanlılığı potansiyeli","1-5","3%","AI/ML"],
 ["Karmaşıklık","K7","Otonomi / dinamik yeniden kalibrasyon","1-5","2%","Dinamik model"],
 ["Bağımlılık","B1","Downstream model/karar bağımlılığı sayısı","1-5","Modifiye","Skoru yükseltir"],
 ["Otonomi","O1","İnsan gözetimi olmadan otomatik karara girme derecesi","1-5","Modifiye","AI/GenAI"],
]
write_table(ws, 3, q_headers, q_rows, [16, 6, 52, 26, 10, 34])
# Tier esleme tablosu
srow = 3 + len(q_rows) + 3
ws.cell(row=srow, column=1, value="TIER EŞLEME (örnek politika – parametrik)").font = Font(bold=True, size=11, color=NAVY)
tier_headers = ["Toplam Skor Aralığı", "Tier", "Validasyon Kapsamı", "Örnek Revalidasyon Frekansı (POLİTİKA)", "Örnek İzleme Frekansı"]
tier_rows = [
 ["≥ 4.0 (yüksek)","Tier 1","Tam / derinlemesine bağımsız validasyon","Yıllık","Aylık"],
 ["2.5 – 3.99","Tier 2","Standart bağımsız validasyon","2 Yıllık","Üç Aylık"],
 ["1.5 – 2.49","Tier 3","Hafif / hedefli gözden geçirme","3 Yıllık","Altı Aylık / Yıllık"],
 ["< 1.5 / kriter","İhmal Edilebilir","Kayıt + asgari kontrol","Olay-tetikli","Yıllık gözden geçirme"],
]
write_table(ws, srow+1, tier_headers, tier_rows, [22, 16, 34, 34, 24], freeze=False)

# ============================================================
# 5) AI/ML EK DEGERLENDIRME
# ============================================================
aiml_rows = [
 ["Tip/Teknik","Algoritma Ailesi","Values List","GBM; Random Forest; Sinir Ağı; Transformer/LLM; Ensemble; Lojistik Reg.; Diğer","Evet","","Teknik sınıf"],
 ["Tip/Teknik","Öğrenme Tipi","Values List","Gözetimli; Gözetimsiz; Pekiştirmeli; Öz-gözetimli","Orta","","Learning type"],
 ["Tip/Teknik","Parametre Sayısı / Mimari","Text","","Hayır","","Model boyutu"],
 ["Temel Model","Temel Model Adı / Versiyon / Sağlayıcı","Text","","Koşullu","Tedarikçi","GenAI için"],
 ["Temel Model","Barındırma Tipi","Values List","İç; Satıcı; Bulut","Koşullu","Uygulamalar","MLOps platform"],
 ["Temel Model","Özelleştirme Yöntemi","Values List","Hazır Prompt; Fine-tuned; RAG","Koşullu","","Kullanım biçimi"],
 ["Temel Model","Lisans / IP / Telif Durumu","Text","","Koşullu","","Hukuki risk"],
 ["Temel Model","Satıcı Risk Notu / Çıkış Planı","Text (area)","","Koşullu","","Konsantrasyon/çıkış"],
 ["Veri/Eğitim","Eğitim Veri Seti / Vintage","Text","","Orta","","Veri tarihi"],
 ["Veri/Eğitim","Veri Lineage Referansı","Text/URL","","Orta","Bilgi Varlıkları","Provenance"],
 ["Veri/Eğitim","Temsil Değerlendirmesi","Text (area)","","Orta","","Popülasyon vs eğitim"],
 ["Veri/Eğitim","Özellik Seti / Sayısı","Text","","Hayır","","Feature set"],
 ["Veri/Eğitim","Hassas Öznitelik / Proxy İncelemesi","Text (area)","","Koşullu","","Dolaylı yanlılık"],
 ["Veri/Eğitim","PII Var mı / Maskeleme Kontrolü","Values List (Evet/Hayır)+Text","Evet; Hayır","Koşullu","Bilgi Güvenliği","Gizlilik"],
 ["Açıklanabilirlik","Açıklanabilirlik Yöntemi","Values List","SHAP; LIME; Surrogate; Behavioral; Doğal Yorumlanabilir","Evet","","Global+lokal"],
 ["Açıklanabilirlik","Reason-Code / Red Gerekçesi Yeteneği","Values List (Evet/Hayır)","Evet; Hayır","Koşullu","","Kredi için"],
 ["Açıklanabilirlik","Model Kartı / Teknik Dosya","Attachment/URL","","Orta","","Annex IV"],
 ["Fairness","İzlenen Fairness Metrikleri","Values List (çoklu)","Disparate Impact; Statistical Parity; Equal Opportunity; Equalized Odds","Koşullu","","Adalet"],
 ["Fairness","Korumalı Gruplar","Text","","Koşullu","","Test grupları"],
 ["Fairness","Disparate Impact Oranı (son)","Numeric","","Hayır","İzleme metriği","4/5 kuralı bayrağı"],
 ["Fairness","Yanlılık Azaltımı Uygulandı mı","Text (area)","","Hayır","","Mitigasyon"],
 ["Robustluk/Güvenlik","Robustluk (Pertürbasyon) Testi","Text/URL","","Orta","","Dayanıklılık"],
 ["Robustluk/Güvenlik","Adversarial / Jailbreak Testi","Text/URL","","Koşullu","","GenAI güvenlik"],
 ["Robustluk/Güvenlik","Reproducibility Manifestosu","Text/URL","","Hayır","","Seed/ortam/versiyon"],
 ["İzleme","Drift Metrikleri + Eşik","Text","→ İzleme Metrik Kataloğu","Orta","Metrik Sonuçları","PSI/CSI/KL"],
 ["İzleme","Bozulma Tetik Tanımı","Text (area)","","Orta","DDE","Revalidasyon/retrain tetiği"],
 ["İzleme","Override / İstisna Oranı","Numeric","","Hayır","İzleme metriği","Operasyonel"],
 ["İzleme","Halüsinasyon Metrik + Eşik (GenAI)","Text","","Koşullu","","NLI/SelfCheckGPT"],
 ["İzleme","Toksisite Metrik + Eşik (GenAI)","Text","","Koşullu","","Guardrail"],
 ["İzleme","Guardrail Modelleri (GenAI)","Text (area)","","Koşullu","","Çıktı kontrolleri"],
 ["Yaşam Döngüsü","Retraining Frekansı / Son Retrain / Tetik","Text+Date","","Koşullu","Model Değişiklik","Yenileme yönetişimi"],
 ["Yaşam Döngüsü","Champion/Challenger Referansı","Text","","Hayır","","Benchmark"],
 ["Yaşam Döngüsü","İnsan Gözetimi Bayrağı + Rol","Values List (Evet/Hayır)+Text","Evet; Hayır","Koşullu","","Human-in-the-loop"],
 ["Yönetişim","Kullanım Amacı Onayı / Onaylı Kapsam","Text (area)","","Orta","","Use-case approval"],
 ["Yönetişim","Etik İnceleme / FRIA Referansı","Text/URL","","Koşullu","","AB yüksek-risk"],
 ["Yönetişim","EU AI Act Risk Kategorisi","Values List","Yüksek Risk; Sınırlı; Minimal; Yasak","Koşullu","","Regülasyon sınıfı"],
 ["Yönetişim","Bağlı Gizlilik / BilgGüv / 3.Taraf Kayıtları","Cross-Reference","→ Bilgi Varlıkları/Tedarikçi","Hayır","","Çerçeve entegrasyonu"],
]
field_sheet("5. AI-ML Ek Değerlendirme", "AI/ML & GenAI EK DEĞERLENDİRME – Alan Tasarımı  (standart alanlara EK; EU AI Act Art.10-15/27, NIST AI RMF, WF GenAI çerçevesi)", aiml_rows)

# ============================================================
# 6) IZLEME METRIK KATALOGU
# ============================================================
ws = wb.create_sheet("6. İzleme Metrik Kataloğu")
ws.merge_cells("A1:G1")
ws["A1"] = "SÜREKLİ İZLEME – METRİK KATALOĞU  (mevcut 'Metrikler / Metrik Sonuçları' uygulamasında MRM metrik seti olarak; eşik aşımı → Bulgu/Aksiyon Planı)"
ws["A1"].font = Font(bold=True, size=11, color=NAVY)
ws["A1"].alignment = Alignment(wrap_text=True, vertical="center")
ws.row_dimensions[1].height = 34
m_headers = ["Kategori", "Metrik", "Ölçtüğü", "Eşik / Konvansiyon", "Model Tipi", "Tetik"]
m_rows = [
 ["Ayrımcılık","AUC / ROC-AUC","İyi/kötü ayrım gücü","Baz çizgiye göre düşüş izlenir","Sınıflandırma","Belirgin düşüş → inceleme"],
 ["Ayrımcılık","Gini (=2·AUC−1)","Ayrım gücü (endüstri ölçeği)","Mutlak seviye + düşüş","Kredi/skorlama","Eşik altı → revalidasyon"],
 ["Ayrımcılık","KS İstatistiği","Skor CDF ayrımı","Bozulma → inceleme","Sınıflandırma",""],
 ["Ayrımcılık","Precision/Recall/F1","Sınıflandırma hata dengesi","Use-case FP/FN maliyetine göre","Sınıflandırma",""],
 ["Stabilite/Drift","PSI (Popülasyon Stabilite)","Skor/çıktı dağılım kayması","<0.10 ihmal; 0.10-0.25 incele; >0.25 aksiyon","Tümü","0.25 üstü → aksiyon"],
 ["Stabilite/Drift","CSI (Karakteristik Stabilite)","Tekil değişken kayması","PSI ile aynı bantlar","Tümü","Özellik stabilitesi"],
 ["Stabilite/Drift","KL / JS Divergence","Dağılım mesafesi (veri drift)","Baz-vs-prod","AI/ML",""],
 ["Stabilite/Drift","Kavram Drift","Girdi→çıktı ilişkisi değişimi","Alert eşiği","AI/ML","→ revalidasyon"],
 ["Kalibrasyon","ECE (Beklenen Kalib. Hatası)","Tahmini olasılık vs gözlem","Kalibrasyon eğrisi","Olasılık modelleri",""],
 ["Kalibrasyon","Kalibrasyon Drift","Zamanla kalibrasyon bozulması","Eşik","Tümü","→ model güncelleme"],
 ["Fairness","Disparate Impact / Adverse Impact","Grup lehte-sonuç oranı","4/5 (0.80) kuralı","AI/ML kredi","Bayrak"],
 ["Fairness","Statistical Parity Diff.","Gruplar arası pozitif oran farkı","Politika eşiği","AI/ML",""],
 ["Fairness","Equal Opportunity / Equalized Odds","TPR (ve FPR) pariteleri","Politika eşiği","AI/ML",""],
 ["Operasyonel","Override / İstisna Oranı","İnsan override sıklığı","Trend/eşik","Tümü","Aşırı override → decommission tetiği"],
 ["Operasyonel","Backtesting Sapması","Tahmin vs gerçekleşen","İstatistiksel test","Tümü","→ ileri analiz"],
 ["GenAI","Halüsinasyon Skoru","Olgusal tutarlılık (NLI/SelfCheckGPT)","Politika eşiği","GenAI","Guardrail"],
 ["GenAI","Toksisite Skoru","Uygunsuz/zararlı çıktı","Politika eşiği","GenAI","Guardrail"],
 ["GenAI","RAG Faithfulness / Answer Relevance","Bağlama dayanma","Politika eşiği","GenAI",""],
 ["GenAI","Query/Domain Stability","Girdi sorgu drifti","Alert","GenAI",""],
 ["GenAI","Operasyonel (feedback/latency/cost)","Kullanıcı geri bildirimi, gecikme, maliyet","Operasyonel eşik","GenAI",""],
]
write_table(ws, 3, m_headers, m_rows, [16, 30, 30, 34, 18, 30])

# ============================================================
# 7) PMA / ISTISNA
# ============================================================
pma_rows = [
 ["Kimlik","PMA/İstisna ID","Text (auto)","PMA-#####","Evet","Anahtar","Tekil kimlik"],
 ["Kimlik","İlgili Model / Portföy","Cross-Reference","→ Model Envanteri","Evet","","Konu model"],
 ["Kimlik","Kayıt Türü","Values List","PMA/Overlay; Yönetim Overlay'i; Override; Kullanım İstisnası; Performans İstisnası","Evet","","Tür"],
 ["PMA","Gerekçe / Justifikasyon","Text (area)","","Evet","","Neden"],
 ["PMA","Materyalite","Values List","Yüksek; Orta; Düşük","Evet","","Etki"],
 ["PMA","Hesaplama Metodolojisi","Text (area)","","Evet","","Nasıl hesaplandı"],
 ["PMA","Azaltma / Kaldırma Kriterleri","Text (area)","","Evet","","Ne zaman kalkar"],
 ["PMA","Uzun Süreli Kullanım Tetikleri","Text (area)","","Evet","DDE/Bildirim","Validasyon+remediation tetiği"],
 ["Onay","Onay Makamı / Destek Kanıtı","Cross-Reference/Text","Üst Yönetim; Risk Komitesi; Denetim Komitesi","Koşullu","","Materyal modelde zorunlu"],
 ["Gözden Geçirme","Bağımsız Gözden Geçirme Kaydı","Text (area)","","Evet","","Alaka, nitel akıl yürütme, girdi, çıktı, kök neden"],
 ["Gözden Geçirme","Kök Neden Analizi","Text (area)","","Orta","","Altta yatan sınırlama"],
 ["Raporlama","PMA'lı ve PMA'sız Raporlama Bayrağı","Values List (Evet/Hayır)","Evet; Hayır","Evet","","Şeffaflık"],
 ["Raporlama","Tekrarlayan Kullanım / Trend Göstergesi","Values List","Evet; Hayır","Hayır","","Tasarım kusuru işareti"],
 ["İstisna","İstisna Türü","Values List","Kullanım; Performans","Koşullu","","İstisna tipi"],
 ["İstisna","Tolerans / İhlal Eşiği","Text/Numeric","","Koşullu","","Maksimum sapma"],
 ["İstisna","Uygulanan Risk Kontrolleri / Azaltıcılar","Text (area)","","Koşullu","","Alternatif model/artan izleme"],
 ["İstisna","Paydaş Onayı","Text","","Koşullu","","Sahip/kullanıcı/üst yönetim"],
 ["İstisna","Eskalasyon Zaman Damgaları","Date/Text","","Koşullu","Bildirimler","Zamanında farkındalık"],
 ["İstisna","Geçerlilik Sonu / Süre","Date","","Koşullu","","Geçici mi"],
]
field_sheet("7. PMA-İstisna", "MODEL RİSKİ AZALTICILARI – PMA / KISIT / İSTİSNA – Alan Tasarımı  (SS1/23 P5: 5.1 PMA, 5.2 kısıt, 5.3 istisna/eskalasyon)", pma_rows)

# ============================================================
# 8) MODEL DEGISIKLIK
# ============================================================
chg_rows = [
 ["Kimlik","Değişiklik ID","Text (auto)","CHG-#####","Evet","Anahtar","Tekil kimlik"],
 ["Kimlik","İlgili Model","Cross-Reference","→ Model Envanteri","Evet","","Konu model"],
 ["Kimlik","Mevcut / Yeni Versiyon","Text","","Evet","","Versiyonlama"],
 ["Değişiklik","Değişiklik Türü","Values List","Rekalibrasyon; Yeniden Geliştirme; Metodoloji; Veri Kaynağı; Retraining; Parametre; Kod/Sistem","Evet","","Tip"],
 ["Değişiklik","Değişiklik Açıklaması","Text (area)","","Evet","","Ne değişti"],
 ["Değişiklik","Değişiklik Gerekçesi / Tetik","Values List+Text","Takvim; Drift; Performans İhlali; Regülasyon; İş İhtiyacı","Evet","","Neden"],
 ["Materyalite","Materyalite Değerlendirmesi","Values List","Materyal; Materyal Değil","Evet","","Revalidasyon eşiği"],
 ["Materyalite","Etki Analizi (besleyen/beslenen)","Text (area)","","Orta","Envanter bağımlılıkları","Downstream etki"],
 ["Validasyon","Revalidasyon Gerekli mi","Values List (Evet/Hayır)","Evet; Hayır","Evet","DDE → Model Validasyonu","Materyal değişiklikte Evet"],
 ["Validasyon","Paralel Çıktı Analizi","Text (area)","","Koşullu","","Öncesi/sonrası vs gerçek"],
 ["Onay","Onay Makamı / Kararı","Values List","Onaylandı; Reddedildi; Beklemede","Evet","Advanced Workflow","Değişiklik onayı"],
 ["Onay","Uygulama Tarihi","Date","","Koşullu","","Yürürlük"],
 ["Kayıt","Kod/Sistem Değişiklik Kontrolü","Text/URL","","Orta","Uygulamalar","QA/change control kanıtı"],
]
field_sheet("8. Model Değişiklik", "MODEL DEĞİŞİKLİK / VERSİYON YÖNETİMİ – Alan Tasarımı  (SS1/23 model change 2.3c(viii)/3.3c; OSFI E-23; materyal değişiklik → revalidasyon)", chg_rows)

# ============================================================
# 9) SECIM LISTELERI
# ============================================================
ws = wb.create_sheet("9. Seçim Listeleri")
ws.merge_cells("A1:C1")
ws["A1"] = "SEÇİM LİSTELERİ (VALUES LISTS) – Global ve uygulama düzeyi değer setleri"
ws["A1"].font = Font(bold=True, size=11, color=NAVY)
vl_headers = ["Liste Adı", "Değerler", "Kullanıldığı Uygulama/Alan"]
vl_rows = [
 ["Model Tipi","İstatistiksel; AI/ML; GenAI/Temel Model; Deterministik/Kural; Hibrit","Envanter · Model Tipi"],
 ["Model Kategorisi","Kredi/IRB; Piyasa/VaR; Operasyonel/AMA; Karşı Taraf/CVA; TFRS9 BKZ; İSEDES Sermaye; ALM/IRRBB; AML/Suistimal; Fiyatlama/Değerleme; Diğer","Envanter · Kategori (çoklu)"],
 ["Yaşam Döngüsü Durumu","Taslak; Geliştirme; Validasyonda; Onayda; Onaylı/Prod; İzlemede; Revalidasyonda; Koşullu Onaylı; Kısıtlı/Askıda; Emekli","Envanter · Model Durumu"],
 ["Tier","Tier 1; Tier 2; Tier 3; İhmal Edilebilir","Envanter · Tier"],
 ["Validasyon Türü","İlk Validasyon; Bağımsız Gözden Geçirme; Süreç Doğrulama; Performans İzleme; Periyodik Revalidasyon; Hedefli","Validasyon · Tür"],
 ["Validasyon Sonucu (öneri)","Onay Önerilir; Koşullu Onay Önerilir; Ret Önerilir","Validasyon · Sonuç"],
 ["Onay Kararı","Onaylandı; Koşullu Onaylandı; Reddedildi; Beklemede","Validasyon/Değişiklik · Onay"],
 ["Model Kökeni","İç Geliştirme; Satıcı/Vendor; Grup-Ana Model","Envanter · Köken"],
 ["Artık Risk","Yüksek; Orta; Düşük","Envanter · Residual"],
 ["Validasyon Frekansı","Yıllık; 2 Yıllık; 3 Yıllık; Olay-Tetikli","Envanter · Frekans"],
 ["Gecikme Bayrağı","Zamanında; Yaklaşan; Gecikmiş","Envanter · Hesaplanan"],
 ["Bulgu Kaynağı (ekleme)","...(mevcut) + Model Validasyonu","Bulgular · Kaynak (mevcut listeye değer eklenir)"],
 ["Algoritma Ailesi","GBM; Random Forest; Sinir Ağı; Transformer/LLM; Ensemble; Lojistik Reg.; Diğer","AI/ML Ek · Algoritma"],
 ["EU AI Act Risk Kategorisi","Yüksek Risk; Sınırlı; Minimal; Yasak","AI/ML Ek · Risk Kategorisi"],
 ["Fairness Metrikleri","Disparate Impact; Statistical Parity; Equal Opportunity; Equalized Odds","AI/ML Ek (çoklu)"],
 ["PMA Kayıt Türü","PMA/Overlay; Yönetim Overlay'i; Override; Kullanım İstisnası; Performans İstisnası","PMA-İstisna · Tür"],
 ["Değişiklik Türü","Rekalibrasyon; Yeniden Geliştirme; Metodoloji; Veri Kaynağı; Retraining; Parametre; Kod/Sistem","Model Değişiklik · Tür"],
 ["Evet/Hayır","Evet; Hayır","Muhtelif bayraklar"],
]
write_table(ws, 3, vl_headers, vl_rows, [26, 74, 40])

# ============================================================
# 10) ILISKI HARITASI
# ============================================================
ws = wb.create_sheet("10. İlişki Haritası")
ws.merge_cells("A1:E1")
ws["A1"] = "UYGULAMALAR ARASI İLİŞKİ / YENİDEN KULLANIM HARİTASI  (yeni MRM uygulamaları ↔ mevcut Archer uygulamaları)"
ws["A1"].font = Font(bold=True, size=11, color=NAVY)
rel_headers = ["Kaynak (MRM)", "İlişki", "Hedef Uygulama", "Yeni / Mevcut", "Amaç"]
rel_rows = [
 ["Model Envanteri","→ sahip birim","Organizasyon – İş Hiyerarşisi","Mevcut","Model sahipliği (GMY/Daire/Bölüm)"],
 ["Model Envanteri","→ satıcı","Tedarikçi","Mevcut","Satıcı/vendor modeller"],
 ["Model Envanteri","→ barındıran sistem","Uygulamalar / IT Hizmetleri","Mevcut","Prod ortamı bağı"],
 ["Model Envanteri","→ donanım","Donanımlar","Mevcut","Altyapı"],
 ["Model Envanteri","→ girdi verisi","Bilgi Varlıkları Envanteri","Mevcut","Veri varlıkları / lineage"],
 ["Model Envanteri","→ ilişkili risk","Risk / Risk Hiyerarşisi","Mevcut","Toplulaştırılmış model riski"],
 ["Model Envanteri","→ kontroller","Kontroller","Mevcut","Model riskini azaltan kontroller"],
 ["Model Envanteri","→ mevzuat","Yasal Mevzuatlar","Mevcut","Regülasyon eşleme"],
 ["Model Envanteri","→ politika","Politikalar ve Standartlar","Mevcut","MRM politikası / DQF"],
 ["Model Envanteri","↔ kendisi","Model Envanteri (self)","Yeni","Upstream/downstream bağımlılık"],
 ["Model Validasyonu","→ model","Model Envanteri","Yeni","Validasyona konu model"],
 ["Model Validasyonu","→ bulgu","Bulgular","Mevcut","Validasyon bulguları"],
 ["Model Validasyonu","→ görev","Görev Yönetimi","Mevcut","Validasyon görevleri"],
 ["Model İzleme","= metrik","Metrikler / Metrik Sonuçları","Mevcut","Sürekli izleme KPI/KRI"],
 ["Model İzleme","→ bulgu (eşik aşımı)","Bulgular / Aksiyon Planları","Mevcut","Eşik aşımı remediation"],
 ["Model Değişiklik","→ model / revalidasyon","Model Envanteri / Model Validasyonu","Yeni","Materyal değişiklikte revalidasyon"],
 ["PMA-İstisna","→ model","Model Envanteri","Yeni","Azaltıcılar"],
 ["PMA-İstisna","→ risk kabul","Risk Kabul Talepleri","Mevcut","İstisna/kabul akışı"],
 ["MRM Öz-Değerlendirme","→ iç denetim","Denetlenen Görüşleri","Mevcut","3. hat gözetimi"],
 ["AI/ML Ek Değerlendirme","→ model","Model Envanteri","Yeni","AI/ML ek nitelikler"],
]
write_table(ws, 3, rel_headers, rel_rows, [24, 20, 32, 12, 40])
for row in ws.iter_rows(min_row=4, max_row=3+len(rel_rows), min_col=4, max_col=4):
    for c in row:
        c.fill = PatternFill("solid", fgColor=GREEN if c.value=="Mevcut" else AMBER)
        c.alignment = Alignment(horizontal="center", vertical="top")

# ============================================================
# 11) RACI
# ============================================================
ws = wb.create_sheet("11. RACI")
ws.merge_cells("A1:H1")
ws["A1"] = "RACI – Rol / Sorumluluk Matrisi  (R=Sorumlu, A=Onaylayan/Hesap veren, C=Danışılan, I=Bilgilendirilen · Üç savunma hattı)"
ws["A1"].font = Font(bold=True, size=11, color=NAVY)
raci_headers = ["Faaliyet", "Model Sahibi (1.hat)", "Model Geliştirici (1.hat)", "Model Kullanıcı (1.hat)", "Bağımsız Validasyon (2.hat)", "Model Risk Komitesi / Onay", "İç Denetim (3.hat)", "Board / SMF"]
raci_rows = [
 ["Model kaydı & envantere giriş","A","R","I","C","I","I","I"],
 ["Kritiklik/tier belirleme","R","C","I","A (teyit)","I","I","I"],
 ["Model geliştirme & dokümantasyon","A","R","I","C","I","-","-"],
 ["Bağımsız validasyon","I","C","I","R/A","I","I","I"],
 ["Model onayı","C","I","I","R (öneri)","A","I","I"],
 ["Prod'a alma / deployment","A","R","C","C","I","-","I"],
 ["Sürekli izleme","A/R","C","R","C","I","I","I"],
 ["Periyodik revalidasyon","C","C","I","R/A","I","I","I"],
 ["Bulgu remediation","A/R","R","C","C (doğrulama)","I","I","I"],
 ["PMA / istisna onayı","R","C","I","C (bağımsız gözden geçirme)","A","I","I"],
 ["Model değişiklik onayı","R","R","I","C","A","I","I"],
 ["Model emekliye ayırma","A/R","C","I","C","A","I","I"],
 ["MRM çerçeve etkinliği denetimi","I","I","I","C","I","R/A","I"],
 ["Model risk iştahı & raporlama","I","I","I","C","C","I","A/R"],
 ["Yıllık öz-değerlendirme / attestation","C","C","I","R","C","C","A"],
]
write_table(ws, 3, raci_headers, raci_rows, [34, 15, 15, 14, 18, 18, 14, 14])
raci_fill = {"A":"F8CBAD","R":"C6E0B4","A/R":"C6E0B4","R/A":"C6E0B4","C":"FFE699","I":"DDEBF7","-":"F2F2F2"}
for row in ws.iter_rows(min_row=4, max_row=3+len(raci_rows), min_col=2, max_col=8):
    for c in row:
        c.fill = PatternFill("solid", fgColor=raci_fill.get(c.value, "FFFFFF"))
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(bold=True, size=10)

# ============================================================
# 12) YASAM DONGUSU GECISLERI
# ============================================================
ws = wb.create_sheet("12. Yaşam Döngüsü")
ws.merge_cells("A1:F1")
ws["A1"] = "MODEL YAŞAM DÖNGÜSÜ – DURUM MAKİNESİ & İŞ AKIŞI GEÇİŞ KURALLARI  (Advanced Workflow + DDE ile uygulanır)"
ws["A1"].font = Font(bold=True, size=11, color=NAVY)
lc_headers = ["Sıra", "Durum", "Girişte Tetiklenen", "Sonraki Durum(lar)", "Geçiş Koşulu / Kontrol", "Sorumlu"]
lc_rows = [
 ["1","Taslak / Tanımlama","Model ID üretimi; kritiklik anketi başlar","Geliştirme","Amaç & kapsam tanımlı","Model Sahibi"],
 ["2","Geliştirme","Geliştirme dokümantasyonu, veri, testler","Bağımsız Validasyonda","Geliştirme testleri + doküman tamam","Geliştirici"],
 ["3","Bağımsız Validasyonda","Model Validasyonu kaydı açılır","Onayda / Geliştirme(geri)","Validasyon raporu + öneri hazır","Bağımsız Validasyon"],
 ["4","Onayda","Advanced Workflow onay adımı","Onaylı/Prod · Koşullu Onaylı · Reddedildi","Onay makamı kararı","Model Risk Komitesi"],
 ["5","Onaylı / Prod'da","Prod tarihi; izleme metrikleri devreye","İzlemede","Deployment testleri tamam","Sahip + BT"],
 ["6","İzlemede","İzleme metrikleri periyodik; eşik DDE","Revalidasyonda · Kısıtlı · Değişiklik","Frekans dolumu / eşik aşımı / değişiklik","Sahip + Validasyon"],
 ["7","Revalidasyonda","Revalidasyon kaydı","İzlemede · Kısıtlı · Emekli","Revalidasyon sonucu","Bağımsız Validasyon"],
 ["8","Koşullu Onaylı","PMA/kısıt kaydı zorunlu; tetik izleme","Onaylı/Prod · Kısıtlı","Koşullar giderildi / süre","Sahip + Validasyon"],
 ["9","Kısıtlı / Askıda","Kullanım limiti; bulgu/istisna izleme","İzlemede · Emekli","Ciddi eksik giderildi / karar","Onay makamı"],
 ["10","Emekli / Devre Dışı","Emeklilik gerekçesi; saklama süresi; downstream bildirimi","(son)","Ciddi hata/tekrar ihlal/obsolescence","Sahip + Onay"],
]
write_table(ws, 3, lc_headers, lc_rows, [6, 22, 34, 30, 34, 20])

# ============================================================
# 13) YOL HARITASI
# ============================================================
ws = wb.create_sheet("13. Yol Haritası")
ws.merge_cells("A1:E1")
ws["A1"] = "UYGULAMA YOL HARİTASI / FAZLAMA  (MVP → tam süreç; her faz teslim edilebilir bir bütün)"
ws["A1"].font = Font(bold=True, size=11, color=NAVY)
rm_headers = ["Faz", "Kapsam", "Teslim Edilebilirler", "Temel Gereklilikler", "Bağımlılık"]
rm_rows = [
 ["F1 – Temel Envanter & Yönetişim (MVP)","Model Envanteri uygulaması, tiering anketi, temel roller, durum makinesi, seçim listeleri, erişim rolleri","Kurum geneli model envanteri; kritiklik/tier; MRM politikası kaydı; board paneli iskeleti","A-*, B-01/02/05, C-01..05, E-01/06/07/08, H-02..05, L-01/02/03/07","—"],
 ["F2 – Validasyon & Bulgu Döngüsü","Model Validasyonu uygulaması, Bulgular entegrasyonu, geliştirme alanları, bağımlılıklar, satıcı model","Bağımsız validasyon iş akışı; validasyon→bulgu→aksiyon; BDDK İSEDES/validasyon raporu","B-03/04, D-*, E-02..05, F-02, G-01..03, K-02, L-04/05","F1"],
 ["F3 – İzleme & Raporlama","Metrikler/Metrik Sonuçları MRM seti, dashboard'lar, toplulaştırılmış raporlama, öz-değerlendirme","İzleme metrik kataloğu; model riski ısı haritası; yıllık öz-değerlendirme","D-06, F-01/03/04, H-01/06, J-01, K-01, L-08","F2"],
 ["F4 – AI/ML İleri & Entegrasyon","AI/ML ek değerlendirme derinleştirme, GenAI guardrail metrikleri, DQF/BCBS239, data feeds","GenAI izleme; DQF; MLOps veri beslemesi (opsiyonel)","I-09, J-02, L-06 (+ I-* olgunlaştırma)","F3"],
]
r = write_table(ws, 3, rm_headers, rm_rows, [30, 40, 40, 30, 12])
for i, row in enumerate(ws.iter_rows(min_row=4, max_row=3+len(rm_rows), min_col=1, max_col=1)):
    for c in row:
        c.fill = PatternFill("solid", fgColor=["DDEBF7","D6E0F0","BDD7EE","9DC3E6"][i])
        c.font = Font(bold=True, size=10, color=NAVY)

# ============================================================
# 14) REGULASYON ESLEME
# ============================================================
ws = wb.create_sheet("14. Regülasyon Eşleme")
ws.merge_cells("A1:G1")
ws["A1"] = "REGÜLASYON İZLENEBİLİRLİK EŞLEMESİ  (gereklilik alanları ↔ düzenleme; ✔ = kapsıyor. BDDK sütunu ileriye dönük uyum içindir)"
ws["A1"].font = Font(bold=True, size=11, color=NAVY)
reg_headers = ["Gereklilik Alanı", "SR 11-7 (Fed/OCC)", "OSFI E-23 (2027)", "PRA SS1/23", "ECB TRIM/GtIM", "EU AI Act / NIST", "BDDK (mevcut/ileriye dönük)"]
reg_rows = [
 ["Model tanımı (3 bileşen)","✔","✔","✔ P1","✔","—","İç Sist. Yön. Md.3 (Validasyon tanımı)"],
 ["Kurum geneli model envanteri","✔","✔ (min. şema)","✔ P1","kısmi","—","İleriye dönük; İSEDES model kaydı"],
 ["Risk-tabanlı tiering/materyalite","✔","✔ (rating)","✔ P1","kısmi","NIST MAP","İyi Uyg. Rehberi (orantılılık)"],
 ["Bağımsız validasyon & effective challenge","✔","✔ (review)","✔ P4","✔ (yıllık)","—","İSEDES bağımsız ekip; Rehber #16"],
 ["Kavramsal sağlamlık","✔","✔","✔ P4","✔","—","Rehber #16"],
 ["Süreç doğrulama","✔","✔","✔ P4","✔ (IT test)","—","İç Sist. Md.26"],
 ["Backtesting / çıktı analizi","✔","✔","✔ P4","✔ (Art.185b)","MEASURE","Md.26; Rehber #16 (geriye dönük test)"],
 ["Benchmarking / kıyaslama","✔","✔","✔ P4","✔ (Art.185c)","—","Rehber #16 (kıyaslama)"],
 ["Sürekli izleme & eşikler","✔","✔ (monitoring)","✔ P4","✔ (threshold)","MEASURE/MANAGE","İleriye dönük"],
 ["Model değişiklik yönetimi","✔","✔","✔ (2.3c viii)","✔ (material change)","—","Rehber #16 (güncelleme/iptal)"],
 ["PMA / azaltıcılar / istisna","kısmi","✔ (overlay)","✔ P5","✔ (override)","—","İleriye dönük"],
 ["Yönetişim / board / SMF","✔","✔ P1","✔ P2","✔","NIST GOVERN","İç Sist. Yön. (üç hat)"],
 ["Üç savunma hattı / iç denetim","✔","✔","✔ (2.5)","✔ (audit)","—","İç kontrol/denetim/risk yön."],
 ["Veri kalitesi / yönetişimi","✔","✔ (data)","✔ (3.2)","✔ (7+ boyut)","EU AI Act Art.10","Md.26; ECL Rehberi; BCBS239 (ileriye dönük)"],
 ["Yaşam döngüsü aşamaları","kısmi","✔ (6 aşama)","✔","kısmi","NIST","İleriye dönük"],
 ["AI/ML açıklanabilirlik/yanlılık","—","✔","✔ (1.3c)","—","✔ EU AI Act/NIST","İleriye dönük"],
 ["AI/ML izleme (drift/fairness)","—","✔ (drift)","kısmi","—","✔","İleriye dönük"],
 ["GenAI / temel model / guardrail","—","kısmi","kısmi","—","✔ NIST AI 600-1","İleriye dönük"],
 ["Stres testi bağlantısı","—","kısmi","kısmi","kısmi","—","İç Sist. Md.43; Rehber #11"],
 ["TFRS9 BKZ model yönetişimi","—","—","kısmi","—","—","Karşılıklar Yön.; ECL Rehberi (Dok.943)"],
]
write_table(ws, 3, reg_headers, reg_rows, [34, 15, 16, 14, 16, 18, 34])
for row in ws.iter_rows(min_row=4, max_row=3+len(reg_rows), min_col=2, max_col=6):
    for c in row:
        if c.value == "✔":
            c.fill = PatternFill("solid", fgColor="C6E0B4")
        elif c.value == "kısmi":
            c.fill = PatternFill("solid", fgColor="FFE699")
        elif c.value == "—":
            c.fill = PatternFill("solid", fgColor="F2F2F2")
        c.alignment = Alignment(horizontal="center", vertical="top")

# ============================================================
# 15) REFERANSLAR
# ============================================================
ws = wb.create_sheet("15. Referanslar")
ws.merge_cells("A1:C1")
ws["A1"] = "BİRİNCİL KAYNAK REFERANSLARI"
ws["A1"].font = Font(bold=True, size=11, color=NAVY)
ref_headers = ["Çerçeve / Düzenleme", "Belge", "URL"]
ref_rows = [
 ["US Fed","SR 11-7 Supervisory Guidance on MRM","https://www.federalreserve.gov/supervisionreg/srletters/sr1107.htm"],
 ["US OCC","Bulletin 2011-12 (ve revize 2026-13)","https://www.occ.gov/news-issuances/bulletins/2011/bulletin-2011-12.html"],
 ["OSFI (Kanada)","Guideline E-23 Model Risk Management (2027)","https://www.osfi-bsif.gc.ca/en/guidance/guidance-library/guideline-e-23-model-risk-management-2027"],
 ["PRA / Bank of England","SS1/23 Model risk management principles for banks","https://www.bankofengland.co.uk/prudential-regulation/publication/2023/may/model-risk-management-principles-for-banks-ss"],
 ["ECB","Guide for the Targeted Review of Internal Models (TRIM)","https://www.bankingsupervision.europa.eu/ecb/pub/pdf/trim_guide.en.pdf"],
 ["ECB","Guide to Internal Models (revize Temmuz 2025)","https://www.bankingsupervision.europa.eu/ecb/pub/pdf/ssm.supervisory_guide202507.en.pdf"],
 ["EU","AI Act (Reg. 2024/1689) – Annex III / Art.9-15,27","https://artificialintelligenceact.eu/annex/3/"],
 ["NIST","AI Risk Management Framework (AI 100-1)","https://nvlpubs.nist.gov/nistpubs/ai/nist.ai.100-1.pdf"],
 ["NIST","Generative AI Profile (AI 600-1)","https://doi.org/10.6028/NIST.AI.600-1"],
 ["EBA","Machine Learning for IRB models (discussion + follow-up)","https://www.eba.europa.eu/discussion-paper-machine-learning-irb-models"],
 ["FSB","Financial Stability Implications of AI (2024)","https://www.fsb.org/2024/11/the-financial-stability-implications-of-artificial-intelligence/"],
 ["EY","Model Risk Management for AI/ML","https://www.ey.com/en_us/insights/banking-capital-markets/understand-model-risk-management-for-ai-and-machine-learning"],
 ["Wells Fargo","MRM for Generative AI (arXiv 2503.15668)","https://arxiv.org/pdf/2503.15668"],
 ["BDDK","Bankaların İç Sistemleri ve İSEDES Hakkında Yönetmelik (RG 11/7/2014, 29057)","https://www.mevzuat.gov.tr/mevzuat?MevzuatNo=19864&MevzuatTur=7&MevzuatTertip=5"],
 ["BDDK","İçsel Derec./İleri Ölçüm Değerlendirme, Validasyon ve Kurumsal Yönetim Rehberi (Dok.1041, #16)","https://www.bddk.org.tr/Mevzuat/DokumanGetir/1041"],
 ["BDDK","TFRS 9 Beklenen Kredi Zararı Hesaplama Rehberi (Dok.943)","https://www.bddk.org.tr/Mevzuat/DokumanGetir/943"],
 ["BDDK","Kredilerin Sınıflandırılması ve Karşılıklar Yönetmeliği (RG 22/6/2016, 29750)","https://www.mevzuat.gov.tr/File/GeneratePdf?mevzuatNo=22599&mevzuatTur=KurumVeKurulusYonetmeligi&mevzuatTertip=5"],
]
write_table(ws, 3, ref_headers, ref_rows, [26, 60, 70])

# ---------- Kaydet ----------
out = "/home/user/rrr/mrm-archer/MRM_Archer_Is_Analizi_Gereklilik_Kitabi.xlsx"
wb.save(out)
print("KAYDEDILDI:", out)
print("Sekme sayisi:", len(wb.sheetnames))
for s in wb.sheetnames:
    print(" -", s)
