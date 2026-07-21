# -*- coding: utf-8 -*-
"""Model Envanteri - TAM Archer Field-Config dokumu (Excel)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="1F3864"; BLUE="2E5496"; GREY="F2F2F2"; AMBER="FFF2CC"; GREEN="E2EFDA"
LB="DDEBF7"; ORANGE="F8CBAD"; YEL="FFE699"; GRN="C6E0B4"
thin=Side(style="thin",color="BFBFBF"); border=Border(left=thin,right=thin,top=thin,bottom=thin)

def hdr(cells,fill=NAVY):
    for c in cells:
        c.fill=PatternFill("solid",fgColor=fill); c.font=Font(bold=True,color="FFFFFF",size=9.5)
        c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True); c.border=border
def sc(c,size=9,wrap=True,top=True,fill=None,bold=False,color="000000",mono=False):
    c.font=Font(size=size,name="Consolas" if mono else "Calibri",bold=bold,color=color)
    c.alignment=Alignment(horizontal="left",vertical="top" if top else "center",wrap_text=wrap); c.border=border
    if fill:c.fill=PatternFill("solid",fgColor=fill)
def widths(ws,ws_w):
    for i,w in enumerate(ws_w,1): ws.column_dimensions[get_column_letter(i)].width=w
def table(ws,start,headers,rows,ws_w,zebra=True,freeze=True,mono_cols=()):
    widths(ws,ws_w); r=start
    for j,h in enumerate(headers,1): ws.cell(row=r,column=j,value=h)
    hdr(ws[r]); ws.row_dimensions[r].height=28; r+=1
    for i,row in enumerate(rows):
        for j,val in enumerate(row,1):
            c=ws.cell(row=r,column=j,value=val)
            sc(c,fill=(GREY if zebra and i%2 else None),mono=((j-1) in mono_cols))
        r+=1
    if freeze: ws.freeze_panes=ws.cell(row=start+1,column=1)
    return r
def title_row(ws,text,span,r=1,size=12):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=span)
    c=ws.cell(row=r,column=1,value=text); c.font=Font(bold=True,size=size,color=NAVY)
    c.alignment=Alignment(wrap_text=True,vertical="center"); ws.row_dimensions[r].height=30

wb=openpyxl.Workbook()

# =============== 0. KUNYE / KAPAK ===============
ws=wb.active; ws.title="0. Uygulama Künyesi"; ws.sheet_view.showGridLines=False
widths(ws,[34,90])
ws.merge_cells("A1:B1"); ws["A1"]="MODEL ENVANTERİ – ARCHER FIELD-CONFIG DÖKÜMÜ"
ws["A1"].font=Font(size=20,bold=True,color=NAVY)
ws.merge_cells("A2:B2"); ws["A2"]="Uygulama (Application) tam alan konfigürasyonu · MRM Workspace · Taslak v0.1"
ws["A2"].font=Font(size=12,italic=True,color="595959")
kunye=[
 ("Uygulama adı","Model Envanteri (Model Inventory)"),
 ("Workspace / Solution","Model Riski Yönetimi (MRM) – yeni workspace"),
 ("Uygulama tipi","Standart uygulama (leveled DEĞİL) · kayıt-seviyesi güvenlik açık"),
 ("Anahtar alan (Key Field)","Model Adı (Text). Tekil sistem kimliği: Model ID (Tracking ID)."),
 ("Kayıt sayısı beklentisi","Kurum geneli tüm modeller (yüzler mertebesi) – performans için indeksli anahtar alanlar"),
 ("İş akışı","Advanced Workflow (Sekme 6) – yaşam döngüsü + validasyon/onay/sertifikasyon"),
 ("Ön koşul uygulamalar (mevcut)","Organizasyon/İş Hiyerarşisi (=Business Unit), Bulgular (=Issues Management), Tedarikçi, Uygulamalar/Donanımlar/IT Hizmetleri, Bilgi Varlıkları, Kontroller, Yasal Mevzuatlar, Politikalar ve Standartlar, Metrikler/Metrik Sonuçları"),
 ("Ön koşul uygulamalar (yeni)","Model Validasyonu, Model Kritiklik Değerlendirmesi (questionnaire), AI/ML Ek Değerlendirmesi (questionnaire), Model Değişiklik, PMA-İstisna"),
 ("Referans model","Archer 'Model Risk Management App-Pack' (Platform 6.10+) yapısıyla hizalı; lisanslı ise App-Pack tabanı, değilse core alan tipleriyle özel kurulum (bkz. Sekme 10)"),
 ("Not – DATEADD birim jetonu","Sonraki validasyon hesabında ay-kaydırma jetonu ('m') canlı örnekte/Archer fonksiyon referansında teyit edilmeli."),
]
r=4
for k,v in kunye:
    ws.cell(row=r,column=1,value=k).font=Font(bold=True,color=NAVY,size=10.5)
    c=ws.cell(row=r,column=2,value=v); c.alignment=Alignment(wrap_text=True,vertical="top"); c.font=Font(size=10.5)
    ws.row_dimensions[r].height=30 if len(v)<90 else 46; r+=1
ws.cell(row=r+1,column=1,value="SEKME REHBERİ").font=Font(bold=True,size=12,color=NAVY)
guide=[
 ("1. Field Config","Her alan: Archer alan tipi, konfigürasyon, zorunluluk, seçim listesi, cross-ref hedefi, hesaplama, kaynak"),
 ("2. Seçim Listeleri","Values List değerleri + skorlama/tarih için sayısal eşleme (numeric value mapping)"),
 ("3. Hesaplanan Alanlar","Archer formül sözdizimi ile calculated field tanımları"),
 ("4. Cross-Reference Haritası","Kaynak alan → hedef uygulama → ters (Related Records) alan"),
 ("5. DDE","Data-Driven Events: kural → aksiyon (koşullu görünüm/zorunluluk/set)"),
 ("6. Advanced Workflow","Yaşam döngüsü / onay / sertifikasyon iş akışı node & transition"),
 ("7. Layout","Sekme (tab) ve bölüm (section) yerleşimi + alan sırası"),
 ("8. Erişim & FLS","Access Roles ve alan-seviyesi güvenlik (üç savunma hattı)"),
 ("9. Bildirimler","Notification tetikleyicileri"),
 ("10. App-Pack Hizalama","Archer MRM App-Pack ve mevcut uygulamalarla eşleme + build kararı"),
 ("11. Referanslar","Birincil kaynaklar"),
]
r+=2
for n,d in guide:
    ws.cell(row=r,column=1,value=n).font=Font(bold=True,color=BLUE,size=10)
    c=ws.cell(row=r,column=2,value=d); c.alignment=Alignment(wrap_text=True); c.font=Font(size=10); r+=1

# =============== 1. FIELD CONFIG ===============
ws=wb.create_sheet("1. Field Config")
title_row(ws,"MODEL ENVANTERİ – ALAN KONFİGÜRASYONU (FIELD CONFIG)  ·  Zorunluluk: E=Evet, K=Koşullu (DDE), H=Hayır  ·  Tip = birebir Archer alan tipi",11)
FH=["#","Bölüm / Tab","Alan Adı (TR)","Alan Adı (EN)","Archer Alan Tipi","Zor.","Anahtar/Tekil","Konfigürasyon / Değer / Hedef","Hesaplama / DDE","Kaynak"]
# rows: num, tab, tr, en, type, req, key, config, calc_dde, source
F=[
 # KIMLIK
 ["1","Genel > Kimlik","Model ID","Model ID","Tracking ID (System)","E","Tekil (auto)","MDL-##### otomatik; salt-okunur","—","E-23; SR 11-7"],
 ["2","Genel > Kimlik","Model Adı","Model Name","Text (tek satır)","E","KEY FIELD · Unique","Max 250 kr; benzersiz kısıt açık","—","E-23; SR; Archer"],
 ["3","Genel > Kimlik","Model Tanımı / Temel Özellikler","Model Description","Text (çok satır/rich)","E","—","Rich text; ~4000 kr","—","E-23; SS 1.2c"],
 ["4","Genel > Kimlik","Model Versiyonu","Model Version","Text (tek satır)","K","—","Materyal modelde zorunlu (DDE)","DDE-08","E-23; Archer"],
 # SINIFLANDIRMA
 ["5","Genel > Sınıflandırma","Model Tipi","Model Type","Values List (tekli)","E","—","VL: Model Tipi (bkz S2)","—","OSFI; Archer; SAS"],
 ["6","Genel > Sınıflandırma","Model Kategorisi","Model Category","Values List (çoklu)","E","—","VL: Model Kategorisi (çok-seçim)","—","OSFI (enterprise); BDDK"],
 ["7","Genel > Sınıflandırma","Öncelik","Priority","Values List (tekli)","H","—","VL: Öncelik [Yüksek/Orta/Düşük]","—","Archer App-Pack"],
 ["8","Genel > Sınıflandırma","AI/ML Bayrağı","AI/ML Flag","Values List (Evet/Hayır)","E","—","VL: Evet/Hayır","DDE-01 (AI/ML bölümü+anket)","EU AI Act; OP"],
 ["9","Genel > Sınıflandırma","GenAI / Temel Model Bayrağı","GenAI Flag","Values List (Evet/Hayır)","K","—","VL: Evet/Hayır","DDE-02","NIST AI 600-1; WF"],
 ["10","Genel > Sınıflandırma","Dinamik Model Bayrağı","Dynamic Model Flag","Values List (Evet/Hayır)","H","—","Otonom yeniden kalibrasyon","—","SS1/23 (dinamik)"],
 ["11","Genel > Sınıflandırma","Deterministik / EUC Bayrağı","Deterministic/EUC Flag","Values List (Evet/Hayır)","H","—","Model kapsamı dışı ayrım","DDE-03 (kontrol-yalnızca)","SS1/23 (1.1b/c)"],
 # AMAC/KULLANIM
 ["12","Genel > Amaç ve Kullanım","Amaç / Model Kullanımı","Purpose of the Model","Text (çok satır)","E","—","Hedeflenen kullanım","—","E-23; SS; SR; Archer"],
 ["13","Genel > Amaç ve Kullanım","İlgili Ürün / Portföy","Product / Portfolio","Text / Values List","H","—","Ürün-portföy kapsamı","—","SS 1.2c(i); SR"],
 ["14","Genel > Amaç ve Kullanım","Onaylı Kullanımlar","Approved Uses","Text (çok satır)","E","—","Yetkilendirilmiş kullanım","—","E-23"],
 ["15","Genel > Amaç ve Kullanım","Fiili Kullanım","Actual Use","Text (çok satır)","H","—","Amaç dışı kullanım tespiti","—","SS 1.2c(i)"],
 ["16","Genel > Amaç ve Kullanım","Kullanım Kısıtları","Use Restrictions","Text (çok satır)","H","—","Yasak/kısıt kullanımlar","—","SR 11-7"],
 ["17","Genel > Amaç ve Kullanım","Operasyonel Sınırlar","Operating Boundaries","Text (çok satır)","K","—","Kabul edilen girdi/uygulama aralığı","DDE-08 (material)","SS 1.2c(i)"],
 # ROLLER
 ["18","Genel > Sahiplik ve Roller","Sahip İş Birimi","Business Unit Owner","Cross-Reference","E","—","→ Organizasyon/İş Hiyerarşisi (=Business Unit)","XREF-01","E-23; SR; Archer"],
 ["19","Genel > Sahiplik ve Roller","Model Sahibi (kişi)","Model Owner","Users/Groups List","E","—","Kişi seçimi; kayıt izni kaynağı","→ Record Perms","E-23; SR"],
 ["20","Genel > Sahiplik ve Roller","Model Geliştirici","Model Developer","Users/Groups List","E","—","Kişi/grup","—","E-23; SR"],
 ["21","Genel > Sahiplik ve Roller","Model Kullanıcı(ları)","Model Users","Users/Groups List","H","—","Kişi/grup (çoklu)","—","SR; SAS"],
 ["22","Genel > Sahiplik ve Roller","Model Validatörü","Model Reviewer/Validator","Users/Groups List","K","—","2. hat; bağımsızlık FLS ile","DDE-08","E-23; SS"],
 ["23","Genel > Sahiplik ve Roller","Model Onaylayan","Model Approver","Users/Groups List","K","—","Onay makamı (komite/birey)","AWF","E-23"],
 ["24","Genel > Sahiplik ve Roller","MRM Program Sorumlusu","Model Risk Program Manager","Users/Groups List","H","—","MRM 2. hat kontağı","—","Archer; SN"],
 # KOKEN/BARINDIRMA
 ["25","Genel > Köken ve Barındırma","Model Kökeni","Model Origin","Values List (tekli)","E","—","VL: [İç Geliştirme; Satıcı; Grup-Ana]","DDE-04/05","E-23"],
 ["26","Genel > Köken ve Barındırma","Tedarikçi","Vendor / Provider","Cross-Reference","K","—","→ Tedarikçi (satıcı modelde zorunlu)","XREF-02; DDE-04","E-23; SAS"],
 ["27","Genel > Köken ve Barındırma","Grup Validasyonu Koşulları","Group Validation Conditions","Text (çok satır)","K","—","SS1/23 2.6c 3 koşul kanıtı","DDE-05","SS1/23 (2.6c)"],
 ["28","Genel > Köken ve Barındırma","Barındıran Sistem / Uygulama","Model Implementation","Cross-Reference","H","—","→ Uygulamalar / IT Hizmetleri","XREF-03","Archer; OP; E-23"],
 ["29","Genel > Köken ve Barındırma","Donanım / Altyapı","Hosting Hardware","Cross-Reference","H","—","→ Donanımlar","XREF-04","OSFI (deployment)"],
 # KRITIKLIK
 ["30","Kritiklik ve Risk","Kritiklik Değerlendirmesi","Criticality Assessment","Cross-Reference","K","—","→ Model Kritiklik Değerlendirmesi (questionnaire)","XREF-05","SS1/23 (1.3)"],
 ["31","Kritiklik ve Risk","Materyalite Skoru","Materiality Score","Numeric (Calculated)","H","—","Ankettan REF ile çekilir","CALC-03","SS 1.3b"],
 ["32","Kritiklik ve Risk","Karmaşıklık Skoru","Complexity Score","Numeric (Calculated)","H","—","Ankettan REF (AI faktörleri dahil)","CALC-04","SS 1.3c"],
 ["33","Kritiklik ve Risk","İçsel (Inherent) Risk Skoru","Inherent Risk Score","Numeric (Calculated)","H","—","zafiyet × materyalite","CALC-05","OSFI E-23"],
 ["34","Kritiklik ve Risk","Tier / Kritiklik","Materiality Tier","Values List (tekli)","E","—","VL: [Tier1; Tier2; Tier3; İhmal Edilebilir]","CALC-06 / anket set","SS; SN; Archer"],
 ["35","Kritiklik ve Risk","Model Risk Derecesi","Model Risk Rating","Values List (tekli)","E","—","VL: [Yüksek; Orta; Düşük]","—","E-23; SR; SN"],
 ["36","Kritiklik ve Risk","Artık (Residual) Risk","Residual Risk","Values List (tekli)","H","—","VL: [Yüksek; Orta; Düşük]","—","OSFI E-23"],
 # METODOLOJI/VERI
 ["37","Metodoloji ve Veri","Model Metodolojisi","Model Methodology","Text (çok satır)","H","—","Yaklaşım/teknik","—","SS (karmaşıklık); SAS"],
 ["38","Metodoloji ve Veri","Hedef Değişken","Target Variable","Text (tek satır)","K","—","AI/ML/istatistiksel","DDE-01","SAS"],
 ["39","Metodoloji ve Veri","Algoritma Ailesi","Algorithm Family","Values List (tekli)","K","—","AI/ML Ek anketinde detay; özet burada","DDE-01","SAS; OP; EU AI Act"],
 ["40","Metodoloji ve Veri","Girdi Veri Kaynakları (açıklama)","Data Sources & Description","Text (çok satır)","K","—","Materyal modelde zorunlu","DDE-08","E-23; SR; SAS"],
 ["41","Metodoloji ve Veri","Girdi Veri Varlıkları","Input Data Assets","Cross-Reference","H","—","→ Bilgi Varlıkları Envanteri","XREF-06","ECB TRIM (veri); OP"],
 ["42","Metodoloji ve Veri","Model Çıktıları / Raporlama","Model Outputs","Text (çok satır)","H","—","Çıktı ve raporlama","—","SR 11-7"],
 ["43","Metodoloji ve Veri","Besleyen Modeller (Upstream)","Upstream Models","Cross-Reference (internal, self)","H","—","→ Model Envanteri (kendisi)","XREF-07","E-23; SS 1.2b"],
 ["44","Metodoloji ve Veri","Beslenen Modeller (Downstream)","Downstream Models","Related Records","H","—","XREF-07'nin ters aynası (salt görünüm)","XREF-07 (ters)","E-23; SS 1.2b"],
 ["45","Metodoloji ve Veri","Kaynak Kod / Dokümantasyon","Source Code / Docs","Attachment + External Links","H","—","Dosya + URL","—","SAS; OP; SR (3.5)"],
 ["46","Metodoloji ve Veri","Model Kartı / Teknik Dosya","Model Card / Technical File","Attachment","K","—","AI/ML (EU AI Act Annex IV)","DDE-01","EU AI Act; NIST"],
 # VARSAYIM/BULGU
 ["47","Varsayım, Sınır, Bulgu","Varsayımlar","Assumptions","Text (çok satır)","H","—","Temel varsayımlar","—","SS; SR; SAS"],
 ["48","Varsayım, Sınır, Bulgu","Sınırlamalar (istisna/ek gereklilik dahil)","Limitations","Text (çok satır)","E","—","Bilinen kısıtlar","—","E-23; SS; SR"],
 ["49","Varsayım, Sınır, Bulgu","Yakalanamayan Riskler","Risks Not Captured","Text (çok satır)","H","—","Kapsanmayan risk","—","SS 1.2c(ii)"],
 ["50","Varsayım, Sınır, Bulgu","Telafi Edici Kontroller","Compensating Controls","Cross-Reference","H","—","→ Kontroller","XREF-08","SR 11-7"],
 ["51","Varsayım, Sınır, Bulgu","Validasyon Bulguları / Sağlık Göstergesi","Validation Findings / Health","Values List / Text","H","—","VL: [Sağlıklı; Uyarı; Sorunlu] + not","—","SS 1.2c(iii)"],
 ["52","Varsayım, Sınır, Bulgu","Göstergelerin Son Güncelleme Tarihi","Health Last Updated","Date","H","—","İzleme tazeliği","—","SS 1.2c(iii)"],
 ["53","Varsayım, Sınır, Bulgu","Açık Bulgular / Aksiyon Planları","Open Findings / Actions","Cross-Reference","H","—","→ Bulgular (=Issues Management)","XREF-09","SS; SR; OP"],
 ["54","Varsayım, Sınır, Bulgu","Politika İstisnaları / Onay Durumu","Policy Exceptions / Approval","Values List / Text","H","—","VL: [Yok; Beklemede; Onaylı İstisna]","—","SR 11-7; Archer"],
 ["55","Varsayım, Sınır, Bulgu","İlgili Yasal Mevzuat","Related Regulation","Cross-Reference","H","—","→ Yasal Mevzuatlar","XREF-10","BDDK eşleme"],
 ["56","Varsayım, Sınır, Bulgu","MRM Politikası","MRM Policy","Cross-Reference","H","—","→ Politikalar ve Standartlar","XREF-11","SS1/23 (2.3)"],
 # YASAM DONGUSU
 ["57","Yaşam Döngüsü ve Takvim","Model Durumu","Lifecycle Status","Values List (tekli)","E","—","VL: Yaşam Döngüsü (10 değer)","AWF Update Content; DDE-06/07","SS; SR; E-23"],
 ["58","Yaşam Döngüsü ve Takvim","Geliştirme Tarihi","Date Developed","Date","H","—","","—","SR 11-7"],
 ["59","Yaşam Döngüsü ve Takvim","Prod'a Alınma Tarihi","Production Deployment Date","Date","K","—","Go-live","DDE-06","E-23; SR"],
 ["60","Yaşam Döngüsü ve Takvim","Son Validasyon Tarihi","Last Validation Date","Date","K","—","← Model Validasyonu (MOSTRECENTVALUE)","CALC-01 girdi","E-23; SS; SR"],
 ["61","Yaşam Döngüsü ve Takvim","Validasyon Frekansı","Validation Frequency","Values List (tekli)","K","—","VL: Frekans (ay sayısı map: 12/24/36/0)","Tier eşlemesi (DDE-09)","SS 1.2c(iv)"],
 ["62","Yaşam Döngüsü ve Takvim","Sonraki Validasyon Tarihi","Next Validation Date","Date (Calculated)","H","—","Son + frekans (ay)","CALC-01","E-23; SR"],
 ["63","Yaşam Döngüsü ve Takvim","Gecikme Bayrağı","Overdue Flag","Text/Values List (Calculated)","H","—","Zamanında/Yaklaşan/Gecikmiş","CALC-02","İzleme SLA"],
 ["64","Yaşam Döngüsü ve Takvim","İzleme Durumu","Monitoring Status","Values List (tekli)","H","—","VL: [Normal; Uyarı; Eşik Aşımı; İstisna]","← Metrik Sonuçları/DDE","E-23 (min. field)"],
 ["65","Yaşam Döngüsü ve Takvim","Sertifikasyon Durumu","Certification Status","Values List (tekli)","H","—","VL: [Sertifikalı; Beklemede; Süresi Geçmiş]","AWF (Certification)","Archer App-Pack"],
 ["66","Yaşam Döngüsü ve Takvim","Sertifikasyon Frekansı","Certification Frequency","Values List (tekli)","H","—","VL: [Yıllık; 6 Aylık]","—","Archer App-Pack"],
 ["67","Yaşam Döngüsü ve Takvim","Son Sertifikasyon Tarihi","Last Certification Date","Date","H","—","","—","Archer App-Pack"],
 ["68","Yaşam Döngüsü ve Takvim","Validasyon Kayıtları","Validation Records","Cross-Reference","H","—","→ Model Validasyonu","XREF-12","SS1/23 P4"],
 ["69","Yaşam Döngüsü ve Takvim","Değişiklik Talepleri","Change Requests","Cross-Reference","H","—","→ Model Değişiklik","XREF-13","Archer; SS (2.3c viii)"],
 ["70","Yaşam Döngüsü ve Takvim","PMA / İstisna Kayıtları","PMA / Exceptions","Cross-Reference","H","—","→ PMA-İstisna","XREF-14","SS1/23 P5"],
 ["71","Yaşam Döngüsü ve Takvim","AI/ML Ek Değerlendirmesi","AI/ML Supplement","Cross-Reference","K","—","→ AI/ML Ek Değerlendirme","XREF-15; DDE-01","EU AI Act; NIST"],
 # EMEKLILIK
 ["72","Emeklilik ve Saklama","Emeklilik Gerekçesi","Decommission Rationale","Text (çok satır)","K","—","Emekli durumda zorunlu","DDE-07","E-23; SS"],
 ["73","Emeklilik ve Saklama","Emeklilik Tarihi","Decommission Date","Date","K","—","","DDE-07","E-23; SS"],
 ["74","Emeklilik ve Saklama","Saklama Süresi Sonu","Retention End Date","Date","H","—","Emekli model saklama","—","E-23 (retention)"],
 # SISTEM
 ["75","Sistem (System)","Kayıt İzinleri","Record Permissions","Record Permissions","—","—","Satır-seviyesi güvenlik; sahip+MRM+denetim","XREF/roller","SR; SS; ECB (bağımsızlık)"],
 ["76","Sistem (System)","Oluşturulma Tarihi","First Published","First Published Date","—","—","Sistem","—","Platform"],
 ["77","Sistem (System)","Son Güncelleme","Last Updated","Last Updated Date","—","—","Sistem","—","Platform"],
 ["78","Sistem (System)","Kayıt Durumu","Record Status","Record Status","—","—","Yeni/Güncellendi","—","Platform"],
 ["79","Sistem (System)","Değişiklik Geçmişi","History Log","History Log","—","—","Tüm alan değişiklik izi (denetlenebilirlik)","—","E-23 (evergreen/controls)"],
]
endr=table(ws,2,FH,F,[4,20,30,26,24,5,14,40,22,18],mono_cols=())
# Zorunluluk renk
for row in ws.iter_rows(min_row=3,max_row=endr-1,min_col=6,max_col=6):
    for c in row:
        c.fill=PatternFill("solid",fgColor={"E":ORANGE,"K":YEL,"H":LB}.get(c.value,"FFFFFF"))
        c.alignment=Alignment(horizontal="center",vertical="top")
# tip renk
for row in ws.iter_rows(min_row=3,max_row=endr-1,min_col=5,max_col=5):
    for c in row:
        if c.value and "Calculated" in str(c.value): c.fill=PatternFill("solid",fgColor=GREEN)
        elif c.value and ("Cross-Reference" in str(c.value) or "Related" in str(c.value)): c.fill=PatternFill("solid",fgColor=LB)

# =============== 2. SECIM LISTELERI ===============
ws=wb.create_sheet("2. Seçim Listeleri")
title_row(ws,"SEÇİM LİSTELERİ (VALUES LISTS)  ·  Sayısal Değer Eşlemesi (Numeric Value Mapping) skorlama ve tarih hesabı içindir",4)
VLH=["Liste Adı","Değer","Sayısal Değer (map)","Not"]
VL=[
 ["Model Tipi","İstatistiksel","—",""],["Model Tipi","AI/ML","—",""],["Model Tipi","GenAI/Temel Model","—",""],
 ["Model Tipi","Deterministik/Kural","—",""],["Model Tipi","Hibrit","—",""],
 ["Model Kategorisi","Kredi/IRB · Piyasa/VaR · Operasyonel/AMA · Karşı Taraf/CVA · TFRS9 BKZ · İSEDES Sermaye · ALM/IRRBB · AML/Suistimal · Fiyatlama/Değerleme · Diğer","—","Çok-seçim"],
 ["Öncelik","Yüksek","3",""],["Öncelik","Orta","2",""],["Öncelik","Düşük","1",""],
 ["Evet/Hayır","Evet","1","Bayraklar (AI/ML, GenAI, Dinamik, Deterministik)"],["Evet/Hayır","Hayır","0",""],
 ["Model Kökeni","İç Geliştirme","—",""],["Model Kökeni","Satıcı/Vendor","—","→ Tedarikçi zorunlu (DDE-04)"],["Model Kökeni","Grup-Ana Model","—","→ Grup koşulları (DDE-05)"],
 ["Tier / Kritiklik","Tier 1","3","En yüksek yoğunluk"],["Tier / Kritiklik","Tier 2","2",""],["Tier / Kritiklik","Tier 3","1",""],["Tier / Kritiklik","İhmal Edilebilir","0","Hafif yol"],
 ["Model Risk Derecesi","Yüksek","3",""],["Model Risk Derecesi","Orta","2",""],["Model Risk Derecesi","Düşük","1",""],
 ["Artık Risk","Yüksek","3",""],["Artık Risk","Orta","2",""],["Artık Risk","Düşük","1",""],
 ["Validasyon Frekansı","Yıllık","12","DATEADD ay girdisi"],["Validasyon Frekansı","2 Yıllık","24",""],["Validasyon Frekansı","3 Yıllık","36",""],["Validasyon Frekansı","Olay-Tetikli","0","Takvimsiz"],
 ["Yaşam Döngüsü Durumu","Taslak · Geliştirme · Validasyonda · Onayda · Onaylı/Prod · İzlemede · Revalidasyonda · Koşullu Onaylı · Kısıtlı/Askıda · Emekli","—","AWF ile yönetilir"],
 ["İzleme Durumu","Normal","0",""],["İzleme Durumu","Uyarı","1",""],["İzleme Durumu","Eşik Aşımı","2","→ Bulgu (DDE)"],["İzleme Durumu","İstisna","3",""],
 ["Validasyon Bulguları / Sağlık","Sağlıklı · Uyarı · Sorunlu","—","Health indicator"],
 ["Politika İstisnaları / Onay","Yok · Beklemede · Onaylı İstisna","—",""],
 ["Sertifikasyon Durumu","Sertifikalı · Beklemede · Süresi Geçmiş","—","Archer App-Pack"],
 ["Sertifikasyon Frekansı","Yıllık · 6 Aylık","—","Archer App-Pack"],
 ["Algoritma Ailesi","GBM · Random Forest · Sinir Ağı · Transformer/LLM · Ensemble · Lojistik Reg. · Diğer","—","AI/ML"],
]
table(ws,2,VLH,VL,[26,66,16,34])

# =============== 3. HESAPLANAN ALANLAR ===============
ws=wb.create_sheet("3. Hesaplanan Alanlar")
title_row(ws,"HESAPLANAN ALANLAR (CALCULATED FIELDS)  ·  Archer formül sözdizimi  ·  Not: DATEADD ay-jetonu ('m') canlı örnekte teyit edilmeli",4)
CH=["ID","Alan","Dönüş Tipi","Formül (Archer)","Açıklama"]
C=[
 ["CALC-01","Sonraki Validasyon Tarihi","Date",
  'IF(ISEMPTY([Son Validasyon Tarihi]),"",DATEADD([Son Validasyon Tarihi],"m",SELECTEDVALUENUMBER([Validasyon Frekansı])))',
  "Frekans (ay) kadar ileri; boş tarihte hata vermemesi için ISEMPTY koruması"],
 ["CALC-02","Gecikme Bayrağı","Text",
  'IF(ISEMPTY([Sonraki Validasyon Tarihi]),"—",IF([Sonraki Validasyon Tarihi]<TODAY(),"Gecikmiş",IF([Sonraki Validasyon Tarihi]<DATEADD(TODAY(),"d",30),"Yaklaşan","Zamanında")))',
  "30 gün içinde 'Yaklaşan'; geçmişse 'Gecikmiş'"],
 ["CALC-03","Materyalite Skoru","Numeric",
  'MOSTRECENTVALUE(REF([Kritiklik Değerlendirmesi],[Materyalite Skoru]))',
  "Kritiklik anketinden (WEIGHTEDSCORE ile hesaplanan) skor çekilir"],
 ["CALC-04","Karmaşıklık Skoru","Numeric",
  'MOSTRECENTVALUE(REF([Kritiklik Değerlendirmesi],[Karmaşıklık Skoru]))',
  "AI faktörleri dahil karmaşıklık skoru"],
 ["CALC-05","İçsel Risk Skoru","Numeric",
  'ROUND([Materyalite Skoru]*[Karmaşıklık Skoru],2)',
  "zafiyet (karmaşıklık) × materyalite; alternatif ağırlıklı toplam kullanılabilir"],
 ["CALC-06","Tier (otomatik öneri)","Text",
  'IF([İçsel Risk Skoru]>=P1,"Tier 1",IF([İçsel Risk Skoru]>=P2,"Tier 2",IF([İçsel Risk Skoru]>=P3,"Tier 3","İhmal Edilebilir")))',
  "P1/P2/P3 politika eşikleri; nihai tier validasyonda teyit edilir (manuel geçersiz kılınabilir)"],
 ["—","Son Validasyon Tarihi (besleme)","Date",
  'MOSTRECENTVALUE(REF([Validasyon Kayıtları],[Rapor Tarihi]))',
  "Model Validasyonu'ndaki en güncel tamamlanmış validasyon tarihini çeker"],
]
r=table(ws,2,CH,C,[10,26,10,84,44],mono_cols=(3,))
for row in ws.iter_rows(min_row=3,max_row=r-1,min_col=4,max_col=4):
    for c in row: c.font=Font(name="Consolas",size=8.5)

# =============== 4. CROSS-REFERENCE HARITASI ===============
ws=wb.create_sheet("4. Cross-Reference Haritası")
title_row(ws,"CROSS-REFERENCE / RELATED RECORDS HARİTASI  ·  Her XREF ters yönde otomatik Related Records alanı üretir",5)
XH=["ID","Kaynak Alan (Model Envanteri)","Tip","Hedef Uygulama","Ters (Related Records) alanı / Yeni-Mevcut"]
X=[
 ["XREF-01","Sahip İş Birimi","Cross-Reference (external)","Organizasyon / İş Hiyerarşisi (=Business Unit)","'Sahip Olunan Modeller' · Mevcut"],
 ["XREF-02","Tedarikçi","Cross-Reference (external)","Tedarikçi","'Sağlanan Modeller' · Mevcut"],
 ["XREF-03","Barındıran Sistem / Uygulama","Cross-Reference (external)","Uygulamalar / IT Hizmetleri","'Barındırılan Modeller' · Mevcut"],
 ["XREF-04","Donanım / Altyapı","Cross-Reference (external)","Donanımlar","'İlişkili Modeller' · Mevcut"],
 ["XREF-05","Kritiklik Değerlendirmesi","Cross-Reference (external)","Model Kritiklik Değerlendirmesi (questionnaire)","'Değerlendirilen Model' · Yeni"],
 ["XREF-06","Girdi Veri Varlıkları","Cross-Reference (external)","Bilgi Varlıkları Envanteri","'Kullanan Modeller' · Mevcut"],
 ["XREF-07","Besleyen Modeller (Upstream)","Cross-Reference (INTERNAL/self)","Model Envanteri (kendisi)","'Beslenen Modeller (Downstream)' = Related Records · Yeni"],
 ["XREF-08","Telafi Edici Kontroller","Cross-Reference (external)","Kontroller","'İlişkili Modeller' · Mevcut"],
 ["XREF-09","Açık Bulgular / Aksiyon Planları","Cross-Reference (external)","Bulgular (=Issues Management)","'Kaynak Model' · Mevcut (Bulgu Kaynağı=Model Validasyonu)"],
 ["XREF-10","İlgili Yasal Mevzuat","Cross-Reference (external)","Yasal Mevzuatlar","'İlgili Modeller' · Mevcut"],
 ["XREF-11","MRM Politikası","Cross-Reference (external)","Politikalar ve Standartlar","'Kapsanan Modeller' · Mevcut"],
 ["XREF-12","Validasyon Kayıtları","Cross-Reference (external)","Model Validasyonu","'İlgili Model' · Yeni"],
 ["XREF-13","Değişiklik Talepleri","Cross-Reference (external)","Model Değişiklik","'İlgili Model' · Yeni"],
 ["XREF-14","PMA / İstisna Kayıtları","Cross-Reference (external)","PMA-İstisna","'İlgili Model' · Yeni"],
 ["XREF-15","AI/ML Ek Değerlendirmesi","Cross-Reference (external)","AI/ML Ek Değerlendirme (questionnaire)","'Değerlendirilen Model' · Yeni"],
]
table(ws,2,XH,X,[10,32,26,34,44])

# =============== 5. DDE ===============
ws=wb.create_sheet("5. DDE")
title_row(ws,"DATA-DRIVEN EVENTS (DDE)  ·  Aksiyon tipleri: ACL=Apply Conditional Layout · SVL=Set Values List · SD=Set Date · FVL=Filter Values",5)
DH=["ID","Kural (Koşul)","Aksiyon Tipi","Aksiyon","Amaç"]
D=[
 ["DDE-01",'[AI/ML Bayrağı] = "Evet"',"ACL","'Metodoloji ve Veri' AI alanlarını (Algoritma, Hedef Değişken, Model Kartı) + 'AI/ML Ek Değerlendirmesi' XREF'i GÖSTER ve zorunlu yap","AI/ML ek nitelik ve anketi zorunlu kılar"],
 ["DDE-02",'[GenAI Bayrağı] = "Evet"',"ACL","GenAI'ye özel alanları (temel model/guardrail) AI/ML Ek anketinde göster","GenAI kontrolleri"],
 ["DDE-03",'[Deterministik/EUC Bayrağı] = "Evet"',"ACL","Validasyon/AI sekmelerini GİZLE; 'Kontrol-Yalnızca Not' bölümünü göster; validasyon alanlarının zorunluluğunu kaldır","Kapsam-dışı deterministik yol (SS1/23 1.1c)"],
 ["DDE-04",'[Model Kökeni] = "Satıcı/Vendor"',"ACL","[Tedarikçi] alanını zorunlu yap ve göster","Satıcı model kaydı"],
 ["DDE-05",'[Model Kökeni] = "Grup-Ana Model"',"ACL","[Grup Validasyonu Koşulları] zorunlu yap","SS1/23 2.6c 3 koşul"],
 ["DDE-06",'[Model Durumu] = "Onaylı/Prod"',"ACL","[Prod'a Alınma Tarihi] zorunlu yap","Go-live tarihi"],
 ["DDE-07",'[Model Durumu] = "Emekli"',"ACL","[Emeklilik Gerekçesi] + [Emeklilik Tarihi] zorunlu yap","Devre dışı bırakma kaydı"],
 ["DDE-08",'[Tier] IN {Tier 1, Tier 2}  (materyal)',"ACL","[Model Versiyonu],[Operasyonel Sınırlar],[Model Validatörü],[Girdi Veri Kaynakları] zorunlu yap","OSFI non-negligible ek alanları"],
 ["DDE-09",'[Tier] değişince',"SVL","[Validasyon Frekansı] öneri: Tier1→Yıllık, Tier2→2 Yıllık, Tier3→3 Yıllık (kullanıcı geçersiz kılabilir)","Tier→frekans politika eşlemesi"],
 ["DDE-10",'[Model Durumu] = "Koşullu Onaylı"',"ACL","[PMA/İstisna Kayıtları] zorunlu yap; uyarı göster","Azaltıcı olmadan koşullu onay engellenir"],
 ["DDE-11",'[İzleme Durumu] = "Eşik Aşımı"',"ACL","'Bulgu oluştur' uyarısı + [Açık Bulgular] bölümünü öne çıkar","İzleme→bulgu döngüsü"],
]
table(ws,2,DH,D,[9,34,10,52,34])

# =============== 6. ADVANCED WORKFLOW ===============
ws=wb.create_sheet("6. Advanced Workflow")
title_row(ws,"ADVANCED WORKFLOW (AWF)  ·  Yaşam döngüsü + validasyon/onay + sertifikasyon  ·  User Action geçişleri buton olur",5)
AH=["Node #","Node Tipi","Node Adı","Geçiş(ler) / Buton","Sonraki Node / Aksiyon"]
A=[
 ["N0","Start","Kayıt Oluşturma","(otomatik)","N1"],
 ["N1","User Action","Taslak","Geliştirmeye Gönder","N2 (Update: Durum=Geliştirme)"],
 ["N2","User Action","Geliştirme","Validasyona Gönder","N3 (Durum=Validasyonda)"],
 ["N3","User Action","Bağımsız Validasyonda","Validasyonu Tamamla · Geliştiriciye İade","N4 / N2"],
 ["N4","Evaluate Content","Validasyon Sonucu Yönlendirme","(kurala göre)","Onay Önerilir→N5 · Ret→N2"],
 ["N5","User Action","Onayda (Model Risk Komitesi)","Onayla · Koşullu Onayla · Reddet","N6 / N6c / N2"],
 ["N6","Update Content","Onaylı/Prod'a Al","(otomatik)","Durum=Onaylı/Prod; N7"],
 ["N6c","Update Content","Koşullu Onay","(otomatik)","Durum=Koşullu Onaylı; PMA zorunlu (DDE-10); N7"],
 ["N7","Wait for Content Update","İzlemede","(izleme/eşik/frekans)","N8 · N9 · N10"],
 ["N8","Routing","Revalidasyon Tetiği","Frekans doldu / eşik aşıldı","N3"],
 ["N9","User Action","Sertifikasyon","Sertifika Ver · Reddet","Update: Sertifikasyon Durumu"],
 ["N10","User Action","Emekliye Ayır","Emekliliği Onayla","N11 (Durum=Emekli)"],
 ["N11","Stop","Emekli / Devre Dışı","—","(son; saklama süresi izlenir)"],
]
table(ws,2,AH,A,[8,20,30,40,40])

# =============== 7. LAYOUT ===============
ws=wb.create_sheet("7. Layout")
title_row(ws,"LAYOUT  ·  Sekme (Tab) ve Bölüm (Section) yerleşimi",3)
LH=["Sekme (Tab)","Bölüm (Section)","İçerdiği Alanlar (#)"]
L=[
 ["Genel Bilgiler","Kimlik","1-4"],
 ["Genel Bilgiler","Sınıflandırma","5-11"],
 ["Genel Bilgiler","Amaç ve Kullanım","12-17"],
 ["Genel Bilgiler","Sahiplik ve Roller","18-24"],
 ["Genel Bilgiler","Köken ve Barındırma","25-29"],
 ["Kritiklik ve Risk","Kritiklik Değerlendirmesi","30-36"],
 ["Metodoloji ve Veri","Metodoloji / Veri / Bağımlılıklar","37-46"],
 ["Varsayım, Sınır, Bulgu","Varsayım-Sınırlama-Bulgu","47-56"],
 ["Yaşam Döngüsü","Durum ve Validasyon Takvimi","57-64"],
 ["Yaşam Döngüsü","Sertifikasyon","65-67"],
 ["Yaşam Döngüsü","İlişkili Kayıtlar","68-71"],
 ["Emeklilik","Emeklilik ve Saklama","72-74"],
 ["Sistem","Kayıt Yönetimi","75-79"],
]
table(ws,2,LH,L,[26,40,26])

# =============== 8. ERISIM & FLS ===============
ws=wb.create_sheet("8. Erişim & FLS")
title_row(ws,"ERİŞİM ROLLERİ & ALAN-SEVİYESİ GÜVENLİK (FLS)  ·  R=Read, U=Update, —=Erişim yok  ·  Validasyon bağımsızlığı sistemsel garanti",6)
EH=["Bölüm / Alan grubu","1.Hat Sahip/Geliştirici","2.Hat Bağımsız Validasyon","2.Hat MRM Programı","Model Risk Komitesi","3.Hat İç Denetim"]
E=[
 ["Kimlik / Sınıflandırma / Amaç","R/U","R","R/U","R","R"],
 ["Sahiplik ve Roller","R/U","R","R/U","R","R"],
 ["Kritiklik ve Risk (skor/tier)","R","R/U (teyit)","R/U","R","R"],
 ["Metodoloji / Veri","R/U","R","R","R","R"],
 ["Varsayım / Bulgu / Sağlık göstergesi","R","R/U","R","R","R"],
 ["Validasyon Kayıtları (XREF)","R","R/U","R","R","R"],
 ["Onay / Model Durumu geçişi","R","R (öneri)","R","R/U (onay)","R"],
 ["PMA / İstisna","R/U (öneri)","R/U (bağımsız gözden geçirme)","R","R/U (onay)","R"],
 ["Sertifikasyon","R/U","R","R/U","R","R"],
 ["Sistem / History Log","R","R","R","R","R (denetim)"],
]
r=table(ws,2,EH,E,[34,18,20,16,18,16])
for row in ws.iter_rows(min_row=3,max_row=r-1,min_col=2,max_col=6):
    for c in row:
        v=str(c.value)
        c.fill=PatternFill("solid",fgColor=(GRN if "U" in v else (LB if v=="R" else "F2F2F2")))
        c.alignment=Alignment(horizontal="center",vertical="center")
para=ws.cell(row=r+1,column=1,value="Not: Bağımsız validasyon fonksiyonunun geliştirme/sahiplik alanlarını salt-okur görmesi; skor/tier teyit ve validasyon kayıtlarında yazma yetkisine sahip olması, SR 11-7/SS1/23/ECB bağımsızlık ilkesinin sistemsel karşılığıdır.")
para.font=Font(italic=True,size=9,color="595959"); ws.merge_cells(start_row=r+1,start_column=1,end_row=r+1,end_column=6)
ws.row_dimensions[r+1].height=30

# =============== 9. BILDIRIMLER ===============
ws=wb.create_sheet("9. Bildirimler")
title_row(ws,"BİLDİRİMLER (NOTIFICATIONS)",4)
NH=["Tetikleyici","Alıcı","Zamanlama","Amaç"]
N=[
 ["Sonraki Validasyon Tarihi yaklaşıyor (30 gün)","Model Sahibi, MRM Programı","Günlük değerlendirme","Zamanında revalidasyon"],
 ["Gecikme Bayrağı = Gecikmiş","Model Sahibi, MRM Programı, Komite","Anında + haftalık özet","Gecikmiş validasyon eskalasyonu"],
 ["İzleme Durumu = Eşik Aşımı","Model Sahibi, Validasyon","Anında","İzleme→bulgu döngüsü"],
 ["Açık Bulgu SLA aşımı","Bulgu Sahibi, MRM","Anında","Remediation takibi"],
 ["Sertifikasyon Durumu = Süresi Geçmiş","Model Sahibi","Anında","Periyodik sertifikasyon"],
 ["Model Durumu = Koşullu Onaylı (PMA tetiği)","Model Sahibi, Validasyon","Anında","Uzun süreli PMA → validasyon/remediation"],
 ["Onay için bekliyor (AWF N5)","Model Risk Komitesi","Anında","Onay akışı"],
]
table(ws,2,NH,N,[40,28,22,34])

# =============== 10. APP-PACK HIZALAMA ===============
ws=wb.create_sheet("10. App-Pack Hizalama")
title_row(ws,"ARCHER MRM APP-PACK & MEVCUT UYGULAMA HİZALAMASI  ·  Build kararı",4)
GH=["Bu tasarımdaki alan/uygulama","Archer MRM App-Pack karşılığı","Mevcut Archer uygulaması ile yeniden kullanım","Karar"]
G=[
 ["Model Envanteri","Model Inventory","—","Yeni uygulama (App-Pack tabanı ya da core)"],
 ["Sahip İş Birimi","Business Unit Owner","Organizasyon / İş Hiyerarşisi (=Business Unit)","MEVCUT'a bağla"],
 ["Açık Bulgular / Aksiyon Planları","(Issues Management use case)","Bulgular + Aksiyon Planları","MEVCUT'a bağla (yeni bulgu uygulaması AÇMA)"],
 ["Model Validasyonu","Model Validation","—","Yeni uygulama"],
 ["Model Değişiklik","Model Change Request (Change Type/Significance)","—","Yeni uygulama"],
 ["Sertifikasyon (alanlar 65-67)","Model Certification (Certification Frequency)","—","Envanter içinde alan + AWF"],
 ["Kritiklik Değerlendirmesi","(tiering)","—","Yeni questionnaire (WEIGHTEDSCORE)"],
 ["AI/ML Ek Değerlendirmesi","(App-Pack'te yok)","—","Yeni questionnaire (EU AI Act/NIST)"],
 ["Metrik izleme","(App-Pack'te sınırlı)","Metrikler / Metrik Sonuçları","MEVCUT'u yeniden kullan"],
 ["Öncelik / Model Implementation","Priority / Model Implementation","Uygulamalar / IT Hizmetleri","App-Pack alan adlarıyla hizala"],
]
r=table(ws,2,GH,G,[34,34,34,30])
note=ws.cell(row=r+1,column=1,value="KARAR NOTU: Banka Archer MRM App-Pack'i (Platform 6.10+) lisanslıyorsa, yeni Model Envanteri/Validasyon/Değişiklik/Sertifikasyon yapısını App-Pack tabanına hizalayıp Business Unit ve Issues Management cross-reference'larını yeniden kullanmak; paralel yapı kurmaktan kaçınmak önerilir. Lisans yoksa yukarıdaki core alan tipleriyle özel kurulum yapılır — tasarım her iki senaryoya da uygundur.")
note.font=Font(italic=True,size=9,color="595959"); ws.merge_cells(start_row=r+1,start_column=1,end_row=r+2,end_column=4)
ws.row_dimensions[r+1].height=46

# =============== 11. REFERANSLAR ===============
ws=wb.create_sheet("11. Referanslar")
title_row(ws,"BİRİNCİL KAYNAKLAR (Field-Config)",3)
RH=["Kaynak","Konu","URL"]
RR=[
 ["Archer Help","Field types & configuration","https://help.archerirm.cloud/platform_2024_09/en-us/content/platform/fields/fld_basics.htm"],
 ["Archer Help","Cross-Reference fields","https://help.archerirm.cloud/platform_2025_04/en-us/content/platform/fields/fld_xref_basics.htm"],
 ["Archer Help","Calculated functions/operators","https://help.archerirm.cloud/platform_2025_04/en-us/content/platform/fields/fld_calc_functions_operators_formulas.htm"],
 ["Archer Help","References in formulas (REF)","https://help.archerirm.cloud/platform_2024_09/en-us/content/platform/fields/fld_calc_references.htm"],
 ["Archer Help","WEIGHTEDSCORE","https://help.archerirm.cloud/610-en/content/platform/calcfunctions/system_weightedscore.htm"],
 ["Archer Help","Data-Driven Events","https://help.archerirm.cloud/platform_2024_11/en-us/content/platform/datadrivenevents/dde_basics.htm"],
 ["Archer Help","Advanced Workflow","https://help.archerirm.cloud/platform_2024_11/en-us/content/platform/advancedworkflow/adv_wrkflw_basics.htm"],
 ["Archer Exchange","Model Risk Management App-Pack","https://help.archerirm.cloud/exchange/content/exchange/apppacks/model_risk_management.htm"],
 ["OSFI","E-23 (2027) App.1 min. model inventory","https://www.osfi-bsif.gc.ca/en/guidance/guidance-library/guideline-e-23-model-risk-management-2027"],
 ["PRA/BoE","SS1/23 §1.2c inventory content","https://www.bankofengland.co.uk/-/media/boe/files/prudential-regulation/supervisory-statement/2023/ss123.pdf"],
 ["US Fed","SR 11-7 model inventory","https://www.federalreserve.gov/supervisionreg/srletters/sr1107.htm"],
 ["IBM","OpenPages Model Risk Governance","https://www.ibm.com/docs/en/openpages/8.3.0?topic=objects-openpages-model-risk-governance"],
 ["SAS","Model Risk Management / Model Manager","https://www.sas.com/en_us/software/model-risk-management.html"],
 ["ServiceNow","MRM + AI inventory data model","https://www.servicenow.com/community/cmdb-blog/understanding-the-ai-inventory-data-model-in-servicenow-product/ba-p/3546771"],
]
table(ws,2,RH,RR,[20,44,80])

out="/home/user/rrr/mrm-archer/Model_Envanteri_Archer_Field_Config.xlsx"
wb.save(out)
print("KAYDEDILDI:",out); print("Sekme:",len(wb.sheetnames))
for s in wb.sheetnames: print("  -",s)
print("Alan sayisi:",len(F))
