# -*- coding: utf-8 -*-
"""Model Validasyonu - TAM Archer Field-Config dokumu (Excel)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="1F3864"; BLUE="2E5496"; GREY="F2F2F2"; AMBER="FFF2CC"; GREEN="E2EFDA"
LB="DDEBF7"; ORANGE="F8CBAD"; YEL="FFE699"; GRN="C6E0B4"
thin=Side(style="thin",color="BFBFBF"); border=Border(left=thin,right=thin,top=thin,bottom=thin)

def hdr(cells,fill=NAVY):
    for c in cells:
        c.fill=PatternFill("solid",fgColor=fill); c.font=Font(bold=True,color="FFFFFF",size=9.5)
        c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True); c.border=border
def scell(c,size=9,wrap=True,top=True,fill=None,mono=False):
    c.font=Font(size=size,name="Consolas" if mono else "Calibri"); c.border=border
    c.alignment=Alignment(horizontal="left",vertical="top" if top else "center",wrap_text=wrap)
    if fill:c.fill=PatternFill("solid",fgColor=fill)
def widths(ws,ws_w):
    for i,w in enumerate(ws_w,1): ws.column_dimensions[get_column_letter(i)].width=w
def table(ws,start,headers,rows,ws_w,zebra=True,freeze=True,mono_cols=()):
    widths(ws,ws_w); r=start
    for j,h in enumerate(headers,1): ws.cell(row=r,column=j,value=h)
    hdr(ws[r]); ws.row_dimensions[r].height=28; r+=1
    for i,row in enumerate(rows):
        for j,val in enumerate(row,1):
            scell(ws.cell(row=r,column=j,value=val),fill=(GREY if zebra and i%2 else None),mono=((j-1) in mono_cols))
        r+=1
    if freeze: ws.freeze_panes=ws.cell(row=start+1,column=1)
    return r
def title_row(ws,text,span,r=1,size=11):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=span)
    c=ws.cell(row=r,column=1,value=text); c.font=Font(bold=True,size=size,color=NAVY)
    c.alignment=Alignment(wrap_text=True,vertical="center"); ws.row_dimensions[r].height=30

wb=openpyxl.Workbook()

# ===== 0. KUNYE =====
ws=wb.active; ws.title="0. Uygulama Künyesi"; ws.sheet_view.showGridLines=False
widths(ws,[34,92])
ws.merge_cells("A1:B1"); ws["A1"]="MODEL VALİDASYONU – ARCHER FIELD-CONFIG DÖKÜMÜ"
ws["A1"].font=Font(size=20,bold=True,color=NAVY)
ws.merge_cells("A2:B2"); ws["A2"]="Bağımsız validasyon kaydı tam alan konfigürasyonu · MRM Workspace · Taslak v0.1"
ws["A2"].font=Font(size=12,italic=True,color="595959")
kunye=[
 ("Uygulama adı","Model Validasyonu (Model Validation)"),
 ("Workspace","Model Riski Yönetimi (MRM)"),
 ("Uygulama tipi","Standart uygulama · kayıt-seviyesi güvenlik açık · bağımsızlık FLS ile garanti"),
 ("Anahtar alan","Validasyon Başlığı (Text). Tekil sistem kimliği: Validasyon ID (Tracking ID)."),
 ("Ana ilişki","İlgili Model → Model Envanteri (bir validasyon 1..N model kapsayabilir; genelde 1)"),
 ("İş akışı","Advanced Workflow: Planlandı → Devam → Bulgu Gözden Geçirme → Onayda → Tamamlandı"),
 ("Tasarım ilkesi","Validasyon ÖNERİSİ ile ONAY kararı AYRI alanlardır (SS1/23 4.1b, 2.13); onay makamı validasyondan bağımsızdır"),
 ("Metodolojik omurga","SR 11-7 üç bileşen (kavramsal sağlamlık · süreç doğrulama · çıktı analizi) + SS1/23 P4 (4.1-4.5) + ECB TRIM test kataloğu"),
 ("Ön koşul uygulamalar (mevcut)","Model Envanteri, Bulgular (=Issues Management), Aksiyon Planları, Metrik Sonuçları, Organizasyon (2.hat birim), Tedarikçi (dış uzman kuruluş)"),
 ("Referans model","Archer 'Model Risk Management App-Pack' > Model Validation yapısıyla hizalı"),
 ("Not – DATEADD birim jetonu","Revalidasyon tarihi hesabında ay-kaydırma jetonu ('m') canlı Archer örneğinde teyit edilmeli."),
]
r=4
for k,v in kunye:
    ws.cell(row=r,column=1,value=k).font=Font(bold=True,color=NAVY,size=10.5)
    c=ws.cell(row=r,column=2,value=v); c.alignment=Alignment(wrap_text=True,vertical="top"); c.font=Font(size=10.5)
    ws.row_dimensions[r].height=30 if len(v)<92 else 46; r+=1
ws.cell(row=r+1,column=1,value="SEKME REHBERİ").font=Font(bold=True,size=12,color=NAVY)
guide=[
 ("1. Field Config","Her alan: Archer tipi, konfig, zorunluluk, seçim listesi, cross-ref, hesaplama, kaynak"),
 ("2. Seçim Listeleri","Values List değerleri + sayısal eşleme"),
 ("3. Hesaplanan Alanlar","Archer formül sözdizimi"),
 ("4. Cross-Reference Haritası","Kaynak alan → hedef uygulama → ters alan"),
 ("5. Validasyon Test Kataloğu","ECB TRIM + SR 11-7 test görevleri, sıklık (ilk/yıllık/periyodik), alan eşlemesi"),
 ("6. DDE","Data-Driven Events: kural → aksiyon"),
 ("7. Advanced Workflow","Validasyon/onay iş akışı node & transition"),
 ("8. Layout","Sekme ve bölüm yerleşimi"),
 ("9. Erişim & FLS","Access Roles ve alan-seviyesi güvenlik (bağımsızlık)"),
 ("10. Bildirimler","Notification tetikleyicileri"),
 ("11. App-Pack Hizalama","Archer MRM App-Pack + mevcut uygulama eşleme"),
 ("12. Referanslar","Birincil kaynaklar"),
]
r+=2
for nm,d in guide:
    ws.cell(row=r,column=1,value=nm).font=Font(bold=True,color=BLUE,size=10)
    c=ws.cell(row=r,column=2,value=d); c.alignment=Alignment(wrap_text=True); c.font=Font(size=10); r+=1

# ===== 1. FIELD CONFIG =====
ws=wb.create_sheet("1. Field Config")
title_row(ws,"MODEL VALİDASYONU – ALAN KONFİGÜRASYONU  ·  Zorunluluk: E=Evet, K=Koşullu (DDE), H=Hayır  ·  Tip = birebir Archer alan tipi",10)
FH=["#","Bölüm / Tab","Alan Adı (TR)","Alan Adı (EN)","Archer Alan Tipi","Zor.","Anahtar/Tekil","Konfigürasyon / Değer / Hedef","Hesaplama / DDE","Kaynak"]
F=[
 # KIMLIK
 ["1","Genel > Kimlik","Validasyon ID","Validation ID","Tracking ID (System)","E","Tekil (auto)","VAL-##### otomatik; salt-okunur","—","Tasarım"],
 ["2","Genel > Kimlik","Validasyon Başlığı","Validation Title","Text (tek satır)","E","KEY FIELD","Ör. '[Model] – İlk Validasyon 2026'","—","Tasarım"],
 ["3","Genel > Kimlik","İlgili Model","Related Model","Cross-Reference","E","—","→ Model Envanteri (1..N)","XREF-01","SS1/23 P4; SR"],
 ["4","Genel > Kimlik","Model Tier (getirilen)","Model Tier (lookup)","Values List / Read-only","H","—","← Model Envanteri (REF)","CALC-01","SS (tier→kapsam)"],
 ["5","Genel > Kimlik","Validasyon Türü","Validation Type","Values List (tekli)","E","—","VL: Validasyon Türü (6 değer)","DDE-01/06","SR; SS1/23; ECB"],
 ["6","Genel > Kimlik","Validasyon Gerekçesi / Tetik","Trigger","Values List (tekli)","E","—","VL: [Yeni Model; Materyal Değişiklik; Periyodik; Performans İhlali; Veri Değişikliği; Regülasyon]","—","OSFI E-23 (review triggers)"],
 ["7","Genel > Kimlik","Kapsam (tier'e göre)","Scope","Text (çok satır)","E","—","Tier ile orantılı kapsam/yoğunluk","DDE-02","SS 4.2b; ECB"],
 ["8","Genel > Kimlik","Planlanan Başlangıç / Bitiş","Planned Start/End","Date x2","H","—","Takvim","—","Tasarım"],
 ["9","Genel > Kimlik","Fiili Başlangıç / Bitiş","Actual Start/End","Date x2","H","—","Süre hesabı","CALC-02","Tasarım"],
 # BAGIMSIZLIK
 ["10","Genel > Bağımsızlık","Validatör","Validator","Users/Groups List","E","—","2. hat; FLS ile bağımsızlık","—","SS 4.1; SR; ECB"],
 ["11","Genel > Bağımsızlık","Validasyon Birimi","Validation Unit","Cross-Reference","H","—","→ Organizasyon (2. hat birim)","XREF-02","ECB (ayrı hat)"],
 ["12","Genel > Bağımsızlık","Bağımsızlık Teyidi","Independence Attestation","Values List (tekli)","E","—","VL: [Tam Bağımsız-Ayrı Hat; Aynı Yönetici+Ek Kontrol; Aynı Birim-Ayrı Personel; Dış Uzman Kuruluş]","DDE-03","ECB TRIM (3 opsiyon); BDDK"],
 ["13","Genel > Bağımsızlık","Dış Uzman Kuruluş","External Expert Firm","Cross-Reference","K","—","→ Tedarikçi (dış validasyonda zorunlu)","XREF-03; DDE-03","BDDK İSEDES (uzman kuruluş)"],
 # KAVRAMSAL SAGLAMLIK
 ["14","Kavramsal Sağlamlık","Teori / Metodoloji Değerlendirmesi","Conceptual Soundness","Text (çok satır)","E","—","Teori/mantık/varsayım kalitesi","—","SR 11-7; SS 4.2"],
 ["15","Kavramsal Sağlamlık","Amaca Uygunluk","Fit-for-Purpose","Text (çok satır)","H","—","Kullanım amacına uygunluk","—","SS 4.2"],
 ["16","Kavramsal Sağlamlık","Varsayım Değerlendirmesi","Assumptions Review","Text (çok satır)","H","—","Varsayımların geçerliliği","—","SR; SS"],
 ["17","Kavramsal Sağlamlık","Duyarlılık Analizi (validatör)","Sensitivity Analysis","Text (çok satır)","H","—","Girdi değişimine tepki","—","SS 4.2; SR"],
 ["18","Kavramsal Sağlamlık","Veri Kalitesi / Temsil Değerlendirmesi","Data Quality/Representativeness","Text (çok satır)","H","—","Verinin uygunluğu/temsili","—","ECB TRIM; SS 3.2"],
 ["19","Kavramsal Sağlamlık","Geliştirme Kanıtı Kalitesi","Development Evidence","Text (çok satır)","H","—","Kanıt/dokümantasyon eleştirisi","—","SS 4.2"],
 ["20","Kavramsal Sağlamlık","Nitel Bilgi / Uzman Görüşü Değerlendirmesi","Qualitative Review","Text (çok satır)","H","—","Uzman görüşü sistematiği","—","SS 4.2"],
 ["21","Kavramsal Sağlamlık","Kavramsal Sağlamlık Sonucu","Conceptual Result","Values List (tekli)","H","—","VL: [Tatmin; Kısmen; Tatmin Etmeyen]","—","SR; SS"],
 # SUREC DOGRULAMA
 ["22","Süreç Doğrulama","Girdi Doğrulama","Input Verification","Text (çok satır)","H","—","Girdi temsil/veri kalitesi","DDE-06","SS 4.3; ECB"],
 ["23","Süreç Doğrulama","Hesaplama / Kod QA","Calculation / Code QA","Text (çok satır)","H","—","Kod/hesap doğruluğu","DDE-04 (ilk)","SS 4.3; ECB (code QA)"],
 ["24","Süreç Doğrulama","Entegrasyon / Uygulama Doğrulama","Implementation Verification","Text (çok satır)","H","—","Sistem entegrasyonu","—","SS 4.3"],
 ["25","Süreç Doğrulama","IT Uygulaması Dokümana Uygun mu","IT Reproduces Documented Model","Values List (Evet/Hayır/Kısmi)","H","—","ECB: IT birebir dokümante modeli üretmeli","DDE-04 (ilk)","ECB TRIM"],
 ["26","Süreç Doğrulama","Çıktı / Raporlama Doğrulama","Output/Reporting Verification","Text (çok satır)","H","—","Çıktı doğru/tam/amaca uygun","—","SS 4.3"],
 ["27","Süreç Doğrulama","Süreç Doğrulama Sonucu","Process Result","Values List (tekli)","H","—","VL: [Tatmin; Kısmen; Tatmin Etmeyen]","—","SS 4.3"],
 # CIKTI ANALIZI
 ["28","Çıktı Analizi & Testler","Backtesting / Geriye Dönük Test","Back-testing","Text + Numeric","K","—","Tahmin vs gerçekleşen; istatistiksel test","DDE-02","ECB (Art.185b); SR; BDDK"],
 ["29","Çıktı Analizi & Testler","Ayrım Gücü (AUC/Gini/KS)","Discriminatory Power","Text + Numeric","K","—","Genel + risk faktörü + alt segment","DDE-02","ECB TRIM"],
 ["30","Çıktı Analizi & Testler","Temsil (Representativeness) Analizi","Representativeness","Text (çok satır)","H","—","Veri seti güncel obligor/pozisyon temsili","—","ECB (Art.174c)"],
 ["31","Çıktı Analizi & Testler","Override Analizi","Overrides Analysis","Text (çok satır)","H","—","Override izleme/değerlendirme","—","ECB (Art.172-3); SS 4.4"],
 ["32","Çıktı Analizi & Testler","Stabilite Analizi","Stability Analysis","Text (çok satır)","H","—","Derece/parametre/tasarım stabilitesi","—","ECB TRIM"],
 ["33","Çıktı Analizi & Testler","Benchmarking / Kıyaslama","Benchmarking","Text (çok satır)","H","—","Dış/iç karşılaştırılabilir kaynak","—","ECB (Art.185c); SR"],
 ["34","Çıktı Analizi & Testler","Paralel Çıktı Analizi","Parallel Outcomes","Text (çok satır)","K","—","Öncesi/sonrası vs gerçek (değişiklik)","DDE-01","SS 3.3c/4.4"],
 ["35","Çıktı Analizi & Testler","Eşik / Threshold İhlali","Threshold Breach","Values List (tekli)","H","—","VL: [İhlal Yok; Uyarı; İhlal]","DDE-05","ECB (thresholds)"],
 ["36","Çıktı Analizi & Testler","İlgili İzleme Metrikleri","Related Monitoring Metrics","Cross-Reference","H","—","→ Metrik Sonuçları","XREF-04","SS 4.4; ECB"],
 ["37","Çıktı Analizi & Testler","İlk-Validasyon Ek Görevleri","Initial-Only Tasks","Text (çok satır)","K","—","Replikasyon, kod QA, IT birebir üretim","DDE-04","ECB TRIM (initial only)"],
 # ETKIN SORGULAMA / BULGU / SONUC
 ["38","Sonuç ve Bulgular","Etkin Sorgulama Notları","Effective Challenge","Text (çok satır)","E","—","Kritik/bağımsız analiz","—","SR 11-7; BDDK #16"],
 ["39","Sonuç ve Bulgular","Model Sınırlamaları (validatör tespiti)","Identified Limitations","Text (çok satır)","H","—","Validatörün tespit ettiği kısıtlar","—","SR; SS"],
 ["40","Sonuç ve Bulgular","Üretilen Bulgular","Findings Raised","Cross-Reference","H","—","→ Bulgular (=Issues Management)","XREF-05; DDE-05","SS; SR; ECB"],
 ["41","Sonuç ve Bulgular","Bulgu Sayısı (özet)","Findings Count","Numeric (Calculated)","H","—","İlişkili bulgu adedi","CALC-03","Raporlama"],
 ["42","Sonuç ve Bulgular","Validasyon Sonucu (ÖNERİ)","Validation Recommendation","Values List (tekli)","E","—","VL: [Onay Önerilir; Koşullu Onay Önerilir; Ret Önerilir]","AWF","SS 4.1b (ÖNERİ)"],
 ["43","Sonuç ve Bulgular","Validasyon Derecesi / Rating","Validation Rating","Values List (tekli)","H","—","VL: [Tatmin Edici; Kısmen; Tatmin Etmeyen]","—","SR; SS"],
 ["44","Sonuç ve Bulgular","Önerilen Koşullar / Kısıtlar","Recommended Conditions","Text (çok satır)","K","—","Koşullu onayda zorunlu","DDE-07","SS 5.2 (kısıt)"],
 ["45","Sonuç ve Bulgular","Validasyon Raporu","Validation Report","Attachment + URL","E","—","Rapor belgesi","—","SR 3.5; SS 4; ECB"],
 ["46","Sonuç ve Bulgular","Rapor Tarihi","Report Date","Date","E","—","Tamamlanma tarihi","CALC-04 girdi","SS; ECB (yıllık döngü)"],
 # ONAY (AYRI)
 ["47","Onay (Onay Makamı)","Onaya Sunuldu mu","Submitted for Approval","Values List (Evet/Hayır)","H","—","AWF ile set","AWF","SS 4.1b"],
 ["48","Onay (Onay Makamı)","Onay Makamı","Approval Authority","Values List / Cross-Ref","K","—","VL: [Model Risk Komitesi; Üst Yönetim]","AWF","OSFI (approver)"],
 ["49","Onay (Onay Makamı)","Onay Kararı","Approval Decision","Values List (tekli)","E","—","VL: [Onaylandı; Koşullu Onaylandı; Reddedildi; Beklemede] – AYRI alan","AWF; DDE-08","SS 4.1b (KARAR)"],
 ["50","Onay (Onay Makamı)","Onay Tarihi","Approval Date","Date","K","—","Karar tarihi","—","OSFI"],
 ["51","Onay (Onay Makamı)","Onay Koşulları","Approval Conditions","Text (çok satır)","K","—","Koşullu onayda zorunlu","DDE-08","SS"],
 ["52","Onay (Onay Makamı)","Üst Yönetime Raporlandı mı","Reported to Senior Mgmt","Values List (Evet/Hayır)","H","—","Sonuç/eşik aşımı üst yönetime","—","ECB TRIM; SS 2.1"],
 # TAKIP
 ["53","Takip ve Revalidasyon","Sonraki Revalidasyon Tarihi","Next Revalidation Date","Date (Calculated)","H","—","Rapor + model frekansı","CALC-04","SS 4.5; ECB (yıllık)"],
 ["54","Takip ve Revalidasyon","Açık Aksiyonlar / Remediation","Remediation Actions","Cross-Reference","H","—","→ Aksiyon Planları","XREF-06","SS; SR; ECB"],
 ["55","Takip ve Revalidasyon","Validasyon Durumu (iş akışı)","Validation Status","Values List (tekli)","E","—","VL: [Planlandı; Devam Ediyor; Bulgu Gözden Geçirme; Onayda; Tamamlandı; İptal]","AWF Update","Tasarım"],
 # SISTEM
 ["56","Sistem (System)","Kayıt İzinleri","Record Permissions","Record Permissions","—","—","Bağımsızlık FLS; sahip R, validasyon R/U","XREF/roller","ECB (bağımsızlık)"],
 ["57","Sistem (System)","Oluşturulma Tarihi","First Published","First Published Date","—","—","Sistem","—","Platform"],
 ["58","Sistem (System)","Son Güncelleme","Last Updated","Last Updated Date","—","—","Sistem","—","Platform"],
 ["59","Sistem (System)","Kayıt Durumu","Record Status","Record Status","—","—","Yeni/Güncellendi","—","Platform"],
 ["60","Sistem (System)","Değişiklik Geçmişi","History Log","History Log","—","—","Alan değişiklik izi (denetlenebilirlik)","—","ECB (verifiable)"],
]
endr=table(ws,2,FH,F,[4,20,32,28,24,5,12,42,20,18])
for row in ws.iter_rows(min_row=3,max_row=endr-1,min_col=6,max_col=6):
    for c in row:
        c.fill=PatternFill("solid",fgColor={"E":ORANGE,"K":YEL,"H":LB}.get(c.value,"FFFFFF"))
        c.alignment=Alignment(horizontal="center",vertical="top")
for row in ws.iter_rows(min_row=3,max_row=endr-1,min_col=5,max_col=5):
    for c in row:
        if c.value and "Calculated" in str(c.value): c.fill=PatternFill("solid",fgColor=GREEN)
        elif c.value and ("Cross-Reference" in str(c.value) or "Read-only" in str(c.value)): c.fill=PatternFill("solid",fgColor=LB)

# ===== 2. SECIM LISTELERI =====
ws=wb.create_sheet("2. Seçim Listeleri")
title_row(ws,"SEÇİM LİSTELERİ (VALUES LISTS)  ·  Sayısal eşleme skorlama/rating içindir",4)
VL=[
 ["Validasyon Türü","İlk (Initial) Validasyon","—","Tam kapsam + ilk-only görevler (DDE-04)"],
 ["Validasyon Türü","Bağımsız Gözden Geçirme","—",""],
 ["Validasyon Türü","Süreç Doğrulama","—",""],
 ["Validasyon Türü","Performans İzleme","—",""],
 ["Validasyon Türü","Periyodik Revalidasyon","—","Genelde daha az detaylı"],
 ["Validasyon Türü","Hedefli (Materyal Değişiklik)","—","Değişiklik materyalitesiyle orantılı"],
 ["Validasyon Gerekçesi/Tetik","Yeni Model · Materyal Değişiklik · Periyodik · Performans İhlali · Veri Değişikliği · Regülasyon","—","OSFI review triggers"],
 ["Bağımsızlık Teyidi","Tam Bağımsız-Ayrı Hat","3","G-SII/O-SII için beklenen"],
 ["Bağımsızlık Teyidi","Aynı Yönetici+Ek Kontrol","2","ECB opsiyon (b)"],
 ["Bağımsızlık Teyidi","Aynı Birim-Ayrı Personel","1","ECB opsiyon (c) – G-SII/O-SII HARİÇ"],
 ["Bağımsızlık Teyidi","Dış Uzman Kuruluş","2","BDDK İSEDES uzman kuruluş → Tedarikçi zorunlu"],
 ["Kavramsal/Süreç/Rating Sonucu","Tatmin Edici","3",""],
 ["Kavramsal/Süreç/Rating Sonucu","Kısmen Tatmin","2",""],
 ["Kavramsal/Süreç/Rating Sonucu","Tatmin Etmeyen","1",""],
 ["IT Dokümana Uygunluk","Evet · Kısmi · Hayır","—","ECB IT reproduces"],
 ["Eşik / Threshold İhlali","İhlal Yok","0",""],["Eşik / Threshold İhlali","Uyarı","1",""],["Eşik / Threshold İhlali","İhlal","2","→ Bulgu (DDE-05)"],
 ["Validasyon Sonucu (ÖNERİ)","Onay Önerilir","3",""],["Validasyon Sonucu (ÖNERİ)","Koşullu Onay Önerilir","2","→ Koşullar zorunlu"],["Validasyon Sonucu (ÖNERİ)","Ret Önerilir","1",""],
 ["Onay Kararı","Onaylandı · Koşullu Onaylandı · Reddedildi · Beklemede","—","AYRI alan (onay makamı)"],
 ["Onay Makamı","Model Risk Komitesi · Üst Yönetim","—",""],
 ["Validasyon Durumu","Planlandı · Devam Ediyor · Bulgu Gözden Geçirme · Onayda · Tamamlandı · İptal","—","AWF"],
 ["Evet/Hayır","Evet · Hayır","—","Muhtelif"],
]
table(ws,2,["Liste Adı","Değer","Sayısal Değer","Not"],VL,[26,60,14,40])

# ===== 3. HESAPLANAN ALANLAR =====
ws=wb.create_sheet("3. Hesaplanan Alanlar")
title_row(ws,"HESAPLANAN ALANLAR  ·  Archer formül sözdizimi  ·  DATEADD ay-jetonu ('m') teyit edilmeli",4)
C=[
 ["CALC-01","Model Tier (getirilen)","Text",'MOSTRECENTVALUE(REF([İlgili Model],[Tier]))',"Validasyon kapsam/yoğunluğunu yönlendirmek için modelden tier'i çeker"],
 ["CALC-02","Süre (gün)","Numeric",'IF(OR(ISEMPTY([Fiili Başlangıç]),ISEMPTY([Fiili Bitiş])),"",DATEDIF([Fiili Başlangıç],[Fiili Bitiş],"d"))',"Validasyon süresi"],
 ["CALC-03","Bulgu Sayısı","Numeric",'COUNT(REF([Üretilen Bulgular],[Bulgu ID]))',"İlişkili bulgu adedi (Archer REF aggregation; alternatif: manuel/rollup)"],
 ["CALC-04","Sonraki Revalidasyon Tarihi","Date",'IF(ISEMPTY([Rapor Tarihi]),"",DATEADD([Rapor Tarihi],"m",SELECTEDVALUENUMBER(MOSTRECENTVALUE(REF([İlgili Model],[Validasyon Frekansı])))))',"Rapor tarihine modelin validasyon frekansını (ay) ekler. Not: referanslı VL üzerinde SELECTEDVALUENUMBER kullanımı canlı örnekte teyit edilmeli; alternatif olarak frekans bu kayda REF ile sayısal alan çekilip DATEADD uygulanır."],
]
r=table(ws,2,["ID","Alan","Dönüş Tipi","Formül (Archer)","Açıklama"],C,[9,26,10,86,42],mono_cols=(3,))
for row in ws.iter_rows(min_row=3,max_row=r-1,min_col=4,max_col=4):
    for c in row: c.font=Font(name="Consolas",size=8.5)

# ===== 4. CROSS-REFERENCE HARITASI =====
ws=wb.create_sheet("4. Cross-Reference Haritası")
title_row(ws,"CROSS-REFERENCE / RELATED RECORDS HARİTASI",5)
X=[
 ["XREF-01","İlgili Model","Cross-Reference (external)","Model Envanteri","'Validasyon Kayıtları' (Envanter XREF-12'nin karşılığı) · Mevcut"],
 ["XREF-02","Validasyon Birimi","Cross-Reference (external)","Organizasyon / İş Hiyerarşisi","'Yürütülen Validasyonlar' · Mevcut"],
 ["XREF-03","Dış Uzman Kuruluş","Cross-Reference (external)","Tedarikçi","'Yürütülen Validasyonlar' · Mevcut"],
 ["XREF-04","İlgili İzleme Metrikleri","Cross-Reference (external)","Metrik Sonuçları","'İlgili Validasyon' · Mevcut"],
 ["XREF-05","Üretilen Bulgular","Cross-Reference (external)","Bulgular (=Issues Management)","'Kaynak Validasyon' · Mevcut (Bulgu Kaynağı=Model Validasyonu)"],
 ["XREF-06","Açık Aksiyonlar / Remediation","Cross-Reference (external)","Aksiyon Planları","'İlgili Validasyon' · Mevcut"],
]
table(ws,2,["ID","Kaynak Alan (Model Validasyonu)","Tip","Hedef Uygulama","Ters (Related Records) / Yeni-Mevcut"],X,[10,30,26,30,46])

# ===== 5. VALIDASYON TEST KATALOGU =====
ws=wb.create_sheet("5. Validasyon Test Kataloğu")
title_row(ws,"VALİDASYON TEST KATALOĞU  ·  ECB TRIM + SR 11-7 + SS1/23  ·  Sıklık ve alan eşlemesi",6)
T=[
 ["Kavramsal sağlamlık değerlendirmesi","İlk + materyal değişiklik","Kavramsal Sağlamlık","Teori/mantık/varsayım/veri kalitesi","SR 11-7; SS 4.2","Zorunlu"],
 ["Süreç doğrulama (process verification)","Her validasyon","Süreç Doğrulama","Girdi/hesaplama/çıktı doğrulama","SS 4.3; ECB","Zorunlu"],
 ["Backtesting / geriye dönük test","En az yıllık","Çıktı Analizi (28)","Tahmin vs gerçekleşen","ECB Art.185b; SR; BDDK","Tier'e göre"],
 ["Ayrım gücü (discriminatory power)","En az yıllık","Çıktı Analizi (29)","AUC/Gini/KS; genel+faktör+segment","ECB TRIM","Tier'e göre"],
 ["Temsil (representativeness) analizi","En az yıllık","Çıktı Analizi (30)","Veri seti güncel obligor/pozisyon temsili","ECB Art.174c","Tier'e göre"],
 ["Override analizi","En az yıllık","Çıktı Analizi (31)","Override izleme/değerlendirme","ECB Art.172-3; SS 4.4","Tier'e göre"],
 ["Stabilite analizi","En az yıllık","Çıktı Analizi (32)","Derece/parametre/tasarım stabilitesi","ECB TRIM","Tier'e göre"],
 ["Girdi verisinin değerlendirilmesi","En az yıllık","Süreç/Çıktı","Güvenilir veri; güncel bilgi; temerrütler","ECB TRIM","Zorunlu"],
 ["Benchmarking / kıyaslama","Periyodik","Çıktı Analizi (33)","Dış/iç karşılaştırılabilir kaynak (özellikle düşük temerrüt)","ECB Art.185c; SR","Periyodik"],
 ["Duyarlılık testi / paralel çıktı","İzleme + değişiklik","Çıktı Analizi (17,34)","Operasyonel sınır; öncesi/sonrası","SS 4.4/3.3c","Tier'e göre"],
 ["Nitel analizler","Her validasyon","Kavramsal Sağlamlık (20)","Varsayım, uzman görüşü, doğru kullanım, mevzuat","ECB TRIM","Zorunlu"],
 ["Model geliştirme replikasyonu","YALNIZ ilk + materyal değişiklik","İlk-Only Görevler (37)","Tasarım/varsayım/metodolojiye meydan okuma","ECB TRIM (initial)","İlk-only"],
 ["Bilgisayar kodu QA","YALNIZ ilk + materyal değişiklik","Kod QA (23) / İlk-Only (37)","Kod kalite güvencesi","ECB TRIM (initial)","İlk-only"],
 ["IT uygulaması birebir üretim teyidi","YALNIZ ilk + materyal değişiklik","IT Uygunluk (25)","IT dokümante modeli birebir üretmeli","ECB TRIM (initial)","İlk-only"],
 ["Eşik/threshold değerlendirmesi","Her ilgili test","Eşik İhlali (35)","Backtest/ayrım/temsil/override/stabilite/benchmark için eşik","ECB TRIM","Zorunlu"],
]
table(ws,2,["Test / Görev","Sıklık","Alan eşlemesi (#)","Ne ölçer","Kaynak","Uygulanabilirlik"],T,[34,24,20,34,20,16])

# ===== 6. DDE =====
ws=wb.create_sheet("6. DDE")
title_row(ws,"DATA-DRIVEN EVENTS (DDE)  ·  ACL=Conditional Layout · SVL=Set Values List · SD=Set Date",5)
D=[
 ["DDE-01",'[Validasyon Türü] IN {İlk, Hedefli, Performans İzleme}',"ACL","'Paralel Çıktı Analizi' ve ilgili test alanlarını göster/zorunlu yap","Değişiklik/izleme testleri"],
 ["DDE-02",'[Model Tier] = "Tier 1"',"ACL","Backtesting, Ayrım Gücü ve derinlemesine kapsam alanlarını zorunlu yap","Tier→validasyon yoğunluğu"],
 ["DDE-03",'[Bağımsızlık Teyidi] = "Dış Uzman Kuruluş"',"ACL","[Dış Uzman Kuruluş] zorunlu yap ve göster","BDDK uzman kuruluş"],
 ["DDE-04",'[Validasyon Türü] = "İlk Validasyon"',"ACL","[İlk-Validasyon Ek Görevleri],[Kod QA],[IT Uygunluk] göster/zorunlu yap","ECB initial-only görevler"],
 ["DDE-05",'[Eşik İhlali] = "İhlal"',"ACL","'Bulgu oluştur' uyarısı; [Üretilen Bulgular] bölümünü öne çıkar/zorunlu","İzleme→bulgu döngüsü"],
 ["DDE-06",'[Validasyon Türü] = "Süreç Doğrulama"',"ACL","Süreç Doğrulama bölümünü öne çıkar; ağır çıktı-analizi alanlarını gizle","Kapsam sadeleştirme"],
 ["DDE-07",'[Validasyon Sonucu] = "Koşullu Onay Önerilir"',"ACL","[Önerilen Koşullar / Kısıtlar] zorunlu yap","Koşullu öneri gerekçesi"],
 ["DDE-08",'[Onay Kararı] = "Koşullu Onaylandı"',"ACL","[Onay Koşulları] + [Onay Tarihi] zorunlu yap","Koşullu onay kaydı"],
]
table(ws,2,["ID","Kural (Koşul)","Aksiyon","Aksiyon Detayı","Amaç"],D,[9,40,10,48,32])

# ===== 7. ADVANCED WORKFLOW =====
ws=wb.create_sheet("7. Advanced Workflow")
title_row(ws,"ADVANCED WORKFLOW (AWF)  ·  Validasyon → bulgu gözden geçirme → onay  ·  User Action geçişleri buton",5)
A=[
 ["N0","Start","Kayıt Oluşturma","(otomatik)","N1 (Durum=Planlandı)"],
 ["N1","User Action","Planlandı","Validasyonu Başlat · İptal","N2 / NX"],
 ["N2","User Action","Devam Ediyor","Validasyonu Tamamla","N3 (Durum=Bulgu Gözden Geçirme)"],
 ["N3","User Action","Bulgu Gözden Geçirme","Öneriyi Sun · Ek Test Gerekli","N4 / N2"],
 ["N4","Update Content","Öneri Damgala","(otomatik)","Set: Onaya Sunuldu=Evet; Durum=Onayda; N5"],
 ["N5","User Action","Onayda (Model Risk Komitesi)","Onayla · Koşullu Onayla · Reddet","N6 / N6c / N2"],
 ["N6","Update Content","Onay","(otomatik)","Onay Kararı=Onaylandı; Model Envanteri güncelle (durum+son validasyon+sonraki tarih); N7"],
 ["N6c","Update Content","Koşullu Onay","(otomatik)","Onay Kararı=Koşullu; Onay Koşulları zorunlu (DDE-08); N7"],
 ["N7","Stop","Tamamlandı","—","(son; revalidasyon takvimi tetiklenir)"],
 ["NX","Stop","İptal","—","(son)"],
]
table(ws,2,["Node #","Node Tipi","Node Adı","Geçiş(ler) / Buton","Sonraki Node / Aksiyon"],A,[8,20,28,38,44])

# ===== 8. LAYOUT =====
ws=wb.create_sheet("8. Layout")
title_row(ws,"LAYOUT  ·  Sekme (Tab) ve Bölüm (Section) yerleşimi",3)
L=[
 ["Genel Bilgiler","Kimlik","1-9"],
 ["Genel Bilgiler","Bağımsızlık","10-13"],
 ["Değerlendirme","Kavramsal Sağlamlık","14-21"],
 ["Değerlendirme","Süreç Doğrulama","22-27"],
 ["Değerlendirme","Çıktı Analizi & Testler","28-37"],
 ["Sonuç","Etkin Sorgulama ve Bulgular","38-46"],
 ["Onay","Onay (Onay Makamı)","47-52"],
 ["Takip","Takip ve Revalidasyon","53-55"],
 ["Sistem","Kayıt Yönetimi","56-60"],
]
table(ws,2,["Sekme (Tab)","Bölüm (Section)","İçerdiği Alanlar (#)"],L,[24,40,26])

# ===== 9. ERISIM & FLS =====
ws=wb.create_sheet("9. Erişim & FLS")
title_row(ws,"ERİŞİM ROLLERİ & ALAN-SEVİYESİ GÜVENLİK (FLS)  ·  R=Read, U=Update, —=yok  ·  Bağımsızlık sistemsel garanti",6)
E=[
 ["Kimlik / Bağımsızlık","R","R/U","R","R","R"],
 ["Kavramsal Sağlamlık","R","R/U","R","R","R"],
 ["Süreç Doğrulama","R","R/U","R","R","R"],
 ["Çıktı Analizi & Testler","R","R/U","R","R","R"],
 ["Etkin Sorgulama / Sonuç / ÖNERİ","—","R/U","R","R","R"],
 ["Üretilen Bulgular (XREF)","R","R/U","R","R","R"],
 ["Onay Kararı / Koşulları","R","R","R","R/U (onay)","R"],
 ["Takip / Revalidasyon","R","R/U","R","R","R"],
 ["Sistem / History Log","R","R","R","R","R (denetim)"],
]
r=table(ws,2,["Bölüm / Alan grubu","1.Hat Sahip/Geliştirici","2.Hat Bağımsız Validasyon","2.Hat MRM Programı","Model Risk Komitesi","3.Hat İç Denetim"],E,[34,18,22,16,18,16])
for row in ws.iter_rows(min_row=3,max_row=r-1,min_col=2,max_col=6):
    for c in row:
        v=str(c.value)
        c.fill=PatternFill("solid",fgColor=(GRN if "U" in v else (LB if v=="R" else "F2F2F2")))
        c.alignment=Alignment(horizontal="center",vertical="center")
note=ws.cell(row=r+1,column=1,value="KRİTİK: 1. hat (geliştirici/sahip) 'Etkin Sorgulama / Sonuç / ÖNERİ' alanlarını GÖREMEZ/DEĞİŞTİREMEZ; validasyon kaydının bu bölümleri yalnız bağımsız validasyona açıktır. Onay alanları yalnız Model Risk Komitesine yazılabilir. Bu, SR 11-7/SS1/23/ECB bağımsızlık ilkesinin sistemsel karşılığıdır.")
note.font=Font(italic=True,size=9,color="595959"); ws.merge_cells(start_row=r+1,start_column=1,end_row=r+1,end_column=6); ws.row_dimensions[r+1].height=44

# ===== 10. BILDIRIMLER =====
ws=wb.create_sheet("10. Bildirimler")
title_row(ws,"BİLDİRİMLER (NOTIFICATIONS)",4)
N=[
 ["Validasyon planlandı / atandı","Validatör","Anında","Çalışma başlangıcı"],
 ["Öneri sunuldu (AWF N4)","Model Risk Komitesi","Anında","Onay bekliyor"],
 ["Onay kararı verildi","Model Sahibi, MRM Programı","Anında","Sonuç bildirimi"],
 ["Eşik İhlali = İhlal","Model Sahibi, Validasyon","Anında","Bulgu/remediation"],
 ["Bulgu / Aksiyon SLA aşımı","Aksiyon Sahibi, MRM","Anında","Remediation takibi"],
 ["Sonraki Revalidasyon yaklaşıyor (30 gün)","Model Sahibi, Validasyon","Günlük değerlendirme","Zamanında revalidasyon"],
 ["Validasyon süresi planı aştı","Validasyon Yöneticisi","Haftalık","Kapasite/gecikme"],
]
table(ws,2,["Tetikleyici","Alıcı","Zamanlama","Amaç"],N,[40,28,22,34])

# ===== 11. APP-PACK HIZALAMA =====
ws=wb.create_sheet("11. App-Pack Hizalama")
title_row(ws,"ARCHER MRM APP-PACK & MEVCUT UYGULAMA HİZALAMASI",4)
G=[
 ["Model Validasyonu","Model Validation","—","Yeni uygulama (App-Pack tabanı ya da core)"],
 ["İlgili Model","(Model Inventory ilişkisi)","Model Envanteri","Yeni MRM uygulamasına bağla"],
 ["Üretilen Bulgular / Remediation","(Issues Management)","Bulgular + Aksiyon Planları","MEVCUT'a bağla (yeni bulgu uygulaması AÇMA)"],
 ["İlgili İzleme Metrikleri","(App-Pack'te sınırlı)","Metrik Sonuçları","MEVCUT'u yeniden kullan"],
 ["Dış Uzman Kuruluş","(vendor ilişkisi)","Tedarikçi","MEVCUT'a bağla (BDDK uzman kuruluş)"],
 ["Onay Kararı / ÖNERİ ayrımı","(validation outcome/approval)","—","İki AYRI alan olarak kur (SS1/23 4.1b)"],
 ["Validasyon Test Kataloğu","(App-Pack'te yapılandırılabilir)","—","ECB TRIM/SR 11-7 test setiyle kur (Sekme 5)"],
]
r=table(ws,2,["Bu tasarımdaki alan","Archer MRM App-Pack karşılığı","Mevcut uygulama yeniden kullanım","Karar"],G,[32,32,32,30])
note=ws.cell(row=r+1,column=1,value="KARAR NOTU: App-Pack lisanslıysa Model Validation yapısına hizalanın ve Issues Management/Metrik cross-reference'larını yeniden kullanın. Lisans yoksa core alan tipleriyle kurun. Her iki senaryoda da validasyon ÖNERİSİ ile onay KARARI iki ayrı alan ve iki ayrı rol olmalıdır.")
note.font=Font(italic=True,size=9,color="595959"); ws.merge_cells(start_row=r+1,start_column=1,end_row=r+2,end_column=4); ws.row_dimensions[r+1].height=44

# ===== 12. REFERANSLAR =====
ws=wb.create_sheet("12. Referanslar")
title_row(ws,"BİRİNCİL KAYNAKLAR (Validasyon Field-Config)",3)
RR=[
 ["US Fed","SR 11-7 – validasyon (kavramsal sağlamlık/izleme/çıktı analizi)","https://www.federalreserve.gov/supervisionreg/srletters/sr1107.htm"],
 ["PRA/BoE","SS1/23 Principle 4 – bağımsız validasyon (4.1-4.5)","https://www.bankofengland.co.uk/-/media/boe/files/prudential-regulation/supervisory-statement/2023/ss123.pdf"],
 ["ECB","TRIM – validasyon test kataloğu, eşikler, veri kalitesi","https://www.bankingsupervision.europa.eu/ecb/pub/pdf/trim_guide.en.pdf"],
 ["ECB","Guide to Internal Models (revize Temmuz 2025)","https://www.bankingsupervision.europa.eu/ecb/pub/pdf/ssm.supervisory_guide202507.en.pdf"],
 ["OSFI","E-23 (2027) – model review tetikleri ve rolleri","https://www.osfi-bsif.gc.ca/en/guidance/guidance-library/guideline-e-23-model-risk-management-2027"],
 ["BDDK","İçsel Derec./İleri Ölçüm Validasyon ve Kurumsal Yönetim Rehberi (Dok.1041)","https://www.bddk.org.tr/Mevzuat/DokumanGetir/1041"],
 ["BDDK","İç Sistemler & İSEDES Yönetmeliği (Md.3 Validasyon, Md.26)","https://www.mevzuat.gov.tr/mevzuat?MevzuatNo=19864&MevzuatTur=7&MevzuatTertip=5"],
 ["Archer Help","Field types / Calculated / DDE / Advanced Workflow","https://help.archerirm.cloud/platform_2024_09/en-us/content/platform/fields/fld_basics.htm"],
 ["Archer Exchange","Model Risk Management App-Pack (Model Validation)","https://help.archerirm.cloud/exchange/content/exchange/apppacks/model_risk_management.htm"],
]
table(ws,2,["Kaynak","Konu","URL"],RR,[18,50,78])

out="/home/user/rrr/mrm-archer/Model_Validasyonu_Archer_Field_Config.xlsx"
wb.save(out)
print("KAYDEDILDI:",out); print("Sekme:",len(wb.sheetnames)); print("Alan sayisi:",len(F))
for s in wb.sheetnames: print("  -",s)
